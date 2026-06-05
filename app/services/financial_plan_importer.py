import csv
import hashlib
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.vehicles import Vehicle
from app.services.spreadsheets import normalize_header


PLAN_DOCUMENT_TYPE = "finance_rental_plan"
PLAN_CLASSIFICATION = "finance"
PLAN_SOURCE = "financial_plan_preview"
PLAN_TARGET_SUBPATH = "02_Financeiro/Planos de renda"

DEFAULT_PLAN_ROOT = Path(
    r"C:\Users\andre\OneDrive - D'accord Invest - Serviços Partilhados SA"
    r"\CARFAST - GESTÃO - GESTÃO\Documentação CarFast v2\Planos de renda 19-05"
)
DEFAULT_SALES_DEBT_MAP = Path(
    r"C:\Users\andre\OneDrive - D'accord Invest - Serviços Partilhados SA"
    r"\Descargas OneDrive\frota_03062026_mapa_vendas_step2_venda_divida.xlsx"
)

PLATE_RE = re.compile(r"\b([A-Z]{2}[-\s]?\d{2}[-\s]?[A-Z]{2}|\d{2}[-\s]?[A-Z]{2}[-\s]?\d{2})\b", re.I)
CONTRACT_RE = re.compile(r"(?<!\d)(\d{5,14})(?!\d)")
IGNORED_DIR_PARTS = {"_DUPLICADOS_EXATOS"}
SUPPORTED_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".csv"}


@dataclass
class PlanCandidate:
    entity: str
    relative_path: str
    source_path: str
    file_name: str
    suffix: str
    file_hash: str
    bytes: int
    duplicate_count: int
    contract_number: str
    explicit_plate: str
    sold_hint: bool


def normalize_plate(value: Any) -> str:
    text = str(value or "").strip().upper()
    return re.sub(r"[^A-Z0-9]", "", text)


def normalize_contract(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    digits = re.sub(r"\D", "", text)
    return digits


def normalize_entity(value: Any) -> str:
    text = str(value or "").strip().upper()
    text = re.sub(r"\s+", " ", text)
    aliases = {
        "SANTANDER CONSUMER": "SANTANDER",
        "CGD LOCARENT": "CGD_LOCARENT",
        "CGD-LOCARENT": "CGD_LOCARENT",
        "VW BANK": "VWBFS",
        "VOLKSWAGEN BANK": "VWBFS",
        "LEASE PLAN": "LEASEPLAN",
    }
    return aliases.get(text, text.replace(" ", "_"))


def extract_contract_from_name(name: str) -> str:
    matches = CONTRACT_RE.findall(name)
    return max(matches, key=len) if matches else ""


def extract_plate_from_name(name: str) -> str:
    match = PLATE_RE.search(name)
    return normalize_plate(match.group(1)) if match else ""


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def plan_candidates_from_manifest(root: Path) -> tuple[list[PlanCandidate], list[dict[str, Any]], list[dict[str, Any]]]:
    manifest = root / "_manifesto_planos_renda_19-05.csv"
    ignored_duplicates: list[dict[str, Any]] = []
    ignored_support_files: list[dict[str, Any]] = []
    candidates: list[PlanCandidate] = []
    rows = read_csv_rows(manifest) if manifest.exists() else []
    seen_hashes: set[str] = set()

    if rows:
        for row in rows:
            rel = row.get("CopiedRelativePath") or row.get("SourceRelativePath") or ""
            rel_parts = set(Path(rel).parts)
            suffix = Path(rel).suffix.lower()
            path_parts = Path(rel).parts
            entity = normalize_entity(row.get("Entity") or (path_parts[0] if path_parts else ""))
            is_duplicate = str(row.get("CopiedAsDuplicate") or "").strip().lower() == "true"
            duplicate_dir = bool(rel_parts.intersection(IGNORED_DIR_PARTS))
            file_hash = (row.get("Hash") or "").strip().upper()
            if suffix not in SUPPORTED_EXTENSIONS:
                continue
            if entity == "_GERAL":
                ignored_support_files.append({"relative_path": rel, "hash": file_hash, "reason": "ficheiro_apoio_geral"})
                continue
            if is_duplicate or duplicate_dir or (file_hash and file_hash in seen_hashes):
                ignored_duplicates.append({"relative_path": rel, "hash": file_hash, "reason": "duplicado"})
                continue
            if file_hash:
                seen_hashes.add(file_hash)
            source_path = root / rel
            name = Path(rel).name
            candidates.append(
                PlanCandidate(
                    entity=entity,
                    relative_path=rel,
                    source_path=str(source_path),
                    file_name=name,
                    suffix=suffix,
                    file_hash=file_hash,
                    bytes=int(float(row.get("Bytes") or 0)),
                    duplicate_count=int(float(row.get("DuplicateCountForHash") or 1)),
                    contract_number=extract_contract_from_name(name),
                    explicit_plate=extract_plate_from_name(name),
                    sold_hint="vendido" in name.casefold() or "sold" in name.casefold(),
                )
            )
        return candidates, ignored_duplicates, ignored_support_files

    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        rel = str(path.relative_to(root))
        relative_parts = path.relative_to(root).parts
        entity = normalize_entity(relative_parts[0] if relative_parts else "")
        if entity == "_GERAL":
            ignored_support_files.append({"relative_path": rel, "hash": "", "reason": "ficheiro_apoio_geral"})
            continue
        if set(relative_parts).intersection(IGNORED_DIR_PARTS):
            ignored_duplicates.append({"relative_path": rel, "hash": "", "reason": "pasta duplicados"})
            continue
        digest = file_sha256(path)
        if digest in seen_hashes:
            ignored_duplicates.append({"relative_path": rel, "hash": digest, "reason": "hash repetido"})
            continue
        seen_hashes.add(digest)
        candidates.append(
            PlanCandidate(
                entity=entity,
                relative_path=rel,
                source_path=str(path),
                file_name=path.name,
                suffix=path.suffix.lower(),
                file_hash=digest,
                bytes=path.stat().st_size,
                duplicate_count=1,
                contract_number=extract_contract_from_name(path.name),
                explicit_plate=extract_plate_from_name(path.name),
                sold_hint="vendido" in path.name.casefold() or "sold" in path.name.casefold(),
            )
        )
    return candidates, ignored_duplicates, ignored_support_files


def row_lookup(headers: list[Any]) -> dict[str, int]:
    lookup: dict[str, int] = {}
    for idx, header in enumerate(headers):
        if header not in (None, ""):
            lookup[normalize_header(header)] = idx
    return lookup


def value_by_names(row: tuple[Any, ...], lookup: dict[str, int], names: list[str]) -> Any:
    for name in names:
        idx = lookup.get(normalize_header(name))
        if idx is not None and idx < len(row):
            value = row[idx]
            if value not in (None, ""):
                return value
    return None


def load_sales_debt_map(path: Path) -> dict[tuple[str, str], set[str]]:
    matches: dict[tuple[str, str], set[str]] = defaultdict(set)
    if not path.exists():
        return matches
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Mapa_Base"] if "Mapa_Base" in wb.sheetnames else wb[wb.sheetnames[0]]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        lookup = row_lookup(headers)
        for row in ws.iter_rows(min_row=2, values_only=True):
            plate = normalize_plate(value_by_names(row, lookup, ["matricula", "matrícula", "plate"]))
            entity = normalize_entity(value_by_names(row, lookup, ["entidade_divida", "entidade", "banco"]))
            contract = normalize_contract(value_by_names(row, lookup, ["contrato_divida", "numero_objeto", "nº contrato"]))
            if plate and entity and contract:
                matches[(entity, contract)].add(plate)
    finally:
        wb.close()
    return matches


def load_general_contract_map(root: Path) -> dict[tuple[str, str], set[str]]:
    matches: dict[tuple[str, str], set[str]] = defaultdict(set)
    path = root / "_GERAL" / "Frota_Bancos_NumContrato.xlsx"
    if not path.exists():
        return matches
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            header_row = next(ws.iter_rows(min_row=1, max_row=1))
            headers = [cell.value for cell in header_row]
            lookup = row_lookup(headers)
            for row in ws.iter_rows(min_row=2, values_only=True):
                contract = normalize_contract(value_by_names(row, lookup, ["Nº Contrato", "contrato", "numero_contrato"]))
                plate = normalize_plate(value_by_names(row, lookup, ["platenr", "matricula", "matrícula", "plate"]))
                entity = normalize_entity(row[0] if row else "")
                if entity and contract and plate:
                    matches[(entity, contract)].add(plate)
    finally:
        wb.close()
    return matches


def load_vehicle_index(db: Session | None) -> dict[str, Vehicle]:
    if db is None:
        return {}
    vehicles = db.scalars(select(Vehicle)).all()
    return {normalize_plate(vehicle.plate): vehicle for vehicle in vehicles if vehicle.plate}


def choose_match(
    candidate: PlanCandidate,
    vehicle_by_plate: dict[str, Vehicle],
    sales_map: dict[tuple[str, str], set[str]],
    general_map: dict[tuple[str, str], set[str]],
) -> dict[str, Any]:
    explicit_vehicle = vehicle_by_plate.get(candidate.explicit_plate) if candidate.explicit_plate else None
    contract_key = (candidate.entity, candidate.contract_number)
    sales_plates = sales_map.get(contract_key, set())
    general_plates = general_map.get(contract_key, set())
    contract_plates = sales_plates or general_plates
    existing_contract_plates = {plate for plate in contract_plates if plate in vehicle_by_plate}

    if explicit_vehicle and not contract_plates:
        return {"status": "auto_associado", "reason": "matricula_nome", "plate": candidate.explicit_plate}
    if explicit_vehicle and candidate.explicit_plate in contract_plates:
        return {"status": "auto_associado", "reason": "matricula_nome_contrato", "plate": candidate.explicit_plate}
    if explicit_vehicle and contract_plates and candidate.explicit_plate not in contract_plates:
        return {
            "status": "conflito",
            "reason": "matricula_nome_difere_contrato",
            "plate": candidate.explicit_plate,
            "contract_plates": sorted(contract_plates),
        }
    if len(existing_contract_plates) == 1:
        return {"status": "auto_associado", "reason": "contrato_entidade", "plate": next(iter(existing_contract_plates))}
    if len(contract_plates) == 1:
        return {
            "status": "pre_associado_sem_db",
            "reason": "contrato_tem_matricula_mas_db_nao_validada",
            "plate": next(iter(contract_plates)),
        }
    if len(contract_plates) > 1:
        return {"status": "conflito", "reason": "contrato_ambiguuo", "plate": "", "contract_plates": sorted(contract_plates)}
    if candidate.contract_number:
        return {"status": "sem_correspondencia", "reason": "contrato_nao_encontrado", "plate": ""}
    return {"status": "sem_correspondencia", "reason": "sem_chave_segura", "plate": ""}


def preview_financial_plan_import(
    db: Session | None,
    *,
    plan_root: Path = DEFAULT_PLAN_ROOT,
    sales_debt_map: Path = DEFAULT_SALES_DEBT_MAP,
) -> dict[str, Any]:
    candidates, ignored_duplicates, ignored_support_files = plan_candidates_from_manifest(plan_root)
    sales_map = load_sales_debt_map(sales_debt_map)
    general_map = load_general_contract_map(plan_root)
    vehicle_by_plate = load_vehicle_index(db)

    rows: list[dict[str, Any]] = []
    for candidate in candidates:
        match = choose_match(candidate, vehicle_by_plate, sales_map, general_map)
        plate = match.get("plate") or ""
        vehicle = vehicle_by_plate.get(plate)
        normalized_name = (
            f"{plate} - {candidate.entity} - Contrato {candidate.contract_number} - Plano de renda{candidate.suffix}"
            if plate and candidate.contract_number
            else ""
        )
        target_path = (
            f"01_Documentação Operacional/Viaturas/{plate}/{PLAN_TARGET_SUBPATH}/{normalized_name}"
            if normalized_name
            else ""
        )
        rows.append(
            {
                "status": match["status"],
                "reason": match["reason"],
                "entity": candidate.entity,
                "contract_number": candidate.contract_number,
                "explicit_plate": candidate.explicit_plate,
                "matched_plate": plate,
                "vehicle_id": vehicle.id if vehicle else "",
                "vehicle_vin": vehicle.vin if vehicle else "",
                "vehicle_unit": vehicle.rentway_unit_nr if vehicle else "",
                "sold_hint": "sim" if candidate.sold_hint else "",
                "suffix": candidate.suffix,
                "relative_path": candidate.relative_path,
                "source_path": candidate.source_path,
                "hash": candidate.file_hash,
                "bytes": candidate.bytes,
                "duplicate_count": candidate.duplicate_count,
                "contract_plates": ";".join(match.get("contract_plates", [])),
                "document_type": PLAN_DOCUMENT_TYPE,
                "classification": PLAN_CLASSIFICATION,
                "supplier_name": candidate.entity,
                "target_path_fase2": target_path,
                "normalized_name_fase2": normalized_name,
            }
        )

    summary = Counter(row["status"] for row in rows)
    summary.update({"duplicados_ignorados": len(ignored_duplicates)})
    summary.update({"ficheiros_apoio_ignorados": len(ignored_support_files)})
    extension_counts = Counter(row["suffix"] for row in rows)
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "plan_root": str(plan_root),
        "sales_debt_map": str(sales_debt_map),
        "rows": rows,
        "ignored_duplicates": ignored_duplicates,
        "ignored_support_files": ignored_support_files,
        "summary": dict(summary),
        "extension_counts": dict(extension_counts),
        "vehicle_count": len(vehicle_by_plate),
        "sales_contract_keys": len(sales_map),
        "general_contract_keys": len(general_map),
    }


def write_preview_report(report: dict[str, Any], output_dir: Path) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    rows = report["rows"]
    preview_path = output_dir / "preview_planos_financeiros.csv"
    duplicates_path = output_dir / "duplicados_ignorados.csv"
    support_path = output_dir / "ficheiros_apoio_ignorados.csv"
    summary_path = output_dir / "resumo_preview.csv"
    proposal_path = output_dir / "proposta_tecnica_importador_planos_financeiros.md"

    preview_fields = [
        "status",
        "reason",
        "entity",
        "contract_number",
        "explicit_plate",
        "matched_plate",
        "vehicle_id",
        "vehicle_vin",
        "vehicle_unit",
        "sold_hint",
        "suffix",
        "relative_path",
        "source_path",
        "hash",
        "bytes",
        "duplicate_count",
        "contract_plates",
        "document_type",
        "classification",
        "supplier_name",
        "target_path_fase2",
        "normalized_name_fase2",
    ]
    with preview_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=preview_fields)
        writer.writeheader()
        writer.writerows(rows)

    with duplicates_path.open("w", encoding="utf-8-sig", newline="") as handle:
        fields = ["relative_path", "hash", "reason"]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(report["ignored_duplicates"])

    with support_path.open("w", encoding="utf-8-sig", newline="") as handle:
        fields = ["relative_path", "hash", "reason"]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(report["ignored_support_files"])

    with summary_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["indicador", "valor"])
        writer.writerow(["gerado_em", report["generated_at"]])
        writer.writerow(["pasta_fonte", report["plan_root"]])
        writer.writerow(["ficheiro_mapa_venda_divida", report["sales_debt_map"]])
        writer.writerow(["viaturas_carregadas_da_app", report["vehicle_count"]])
        writer.writerow(["contratos_mapa_venda_divida", report["sales_contract_keys"]])
        writer.writerow(["contratos_apoio_geral", report["general_contract_keys"]])
        for key, value in report["summary"].items():
            writer.writerow([key, value])
        for suffix, value in report["extension_counts"].items():
            writer.writerow([f"extensao_{suffix}", value])

    proposal_path.write_text(
        "\n".join(
            [
                "# Proposta tecnica - importador de planos financeiros",
                "",
                "Objetivo: criar um fluxo dentro do modulo Documentos para pre-validar e depois criar registos `Document` ligados a viaturas.",
                "",
                "Fase 1 - preview/dry-run:",
                "- ler manifesto de planos e ignorar `_DUPLICADOS_EXATOS`/hash repetido;",
                "- extrair entidade, contrato e matricula do nome/caminho;",
                "- cruzar contrato+entidade com mapa venda/divida e ficheiros de apoio `_GERAL`;",
                "- cruzar matricula com a frota da app;",
                "- gerar relatorio de automaticos, sem correspondencia, conflitos e duplicados ignorados.",
                "",
                "Fase 2 - aplicacao controlada:",
                "- criar `Document` com `vehicle_id`, `plate`, `classification=finance`, `document_type=finance_rental_plan`, `supplier_name` e `contract_number`;",
                "- nao copiar ficheiros no primeiro apply; guardar `storage_path` de origem e `folder_path` previsto;",
                "- so copiar/mover para 365 depois de validacao do arquivo final e permissoes.",
                "",
                "Regra de seguranca: qualquer conflito entre matricula explicita e contrato fica para revisao manual.",
            ]
        ),
        encoding="utf-8",
    )

    return {
        "preview": str(preview_path),
        "duplicates": str(duplicates_path),
        "support_files": str(support_path),
        "summary": str(summary_path),
        "proposal": str(proposal_path),
    }
