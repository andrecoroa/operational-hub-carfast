from __future__ import annotations

import csv
from collections.abc import Iterator
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
import json
from pathlib import Path
import re
import unicodedata
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.models.documents import Document, DocumentEvent, VehicleDocumentRecord
from app.models.vehicles import Vehicle
from app.services.document_workflow import get_or_create_workflow_state


VIN_RE = re.compile(r"(?<![A-Z0-9])([A-HJ-NPR-Z0-9]{17})(?![A-Z0-9])", re.IGNORECASE)


def _key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").strip().lower())
    return "".join(char for char in text if char.isalnum() and not unicodedata.combining(char))


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _document_key(value: Any) -> str:
    return _key(value)


def _supplier_nif(value: Any) -> str:
    digits = re.sub(r"\D", "", _text(value))
    return digits if len(digits) == 9 else ""


def _invoice_key(number: Any, supplier_nif: Any) -> str:
    number_key = _document_key(number)
    nif_key = _supplier_nif(supplier_nif)
    return f"{nif_key}:{number_key}" if nif_key else number_key


def _plate_key(value: Any) -> str:
    return _key(value).upper()


def _vin(value: Any) -> str:
    raw = _text(value).upper()
    match = VIN_RE.search(raw)
    if match:
        return match.group(1).upper()
    compact = re.sub(r"[^A-Z0-9]", "", raw)
    return compact if len(compact) == 17 else ""


def _date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _decimal(value: Any) -> str | None:
    text = _text(value).replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return format(Decimal(text), "f")
    except InvalidOperation:
        return None


def _rows(path: Path) -> Iterator[tuple[str, int, dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row_number, row in enumerate(reader, start=2):
                yield "CSV", row_number, {_key(key): value for key, value in row.items()}
        return
    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError("A leitura de .xls requer a dependência xlrd.") from exc
        workbook = xlrd.open_workbook(path)
        for sheet in workbook.sheets():
            if sheet.nrows < 2:
                continue
            headers = [_key(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
            for row_index in range(1, sheet.nrows):
                yield sheet.name, row_index + 1, {
                    headers[col]: sheet.cell_value(row_index, col)
                    for col in range(sheet.ncols)
                    if headers[col]
                }
        return
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            headers = [_key(value) for value in next(iterator, ())]
            for row_number, values in enumerate(iterator, start=2):
                yield sheet.title, row_number, {
                    headers[index]: value
                    for index, value in enumerate(values)
                    if index < len(headers) and headers[index]
                }
    finally:
        workbook.close()


def _first(row: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    for alias in aliases:
        value = row.get(_key(alias))
        if value not in (None, ""):
            return value
    return None


def _vehicle_maps(db: Session) -> tuple[dict[str, Vehicle], dict[str, Vehicle], dict[str, Vehicle]]:
    by_vin: dict[str, Vehicle] = {}
    by_plate: dict[str, Vehicle] = {}
    by_unit: dict[str, Vehicle] = {}
    for vehicle in db.scalars(select(Vehicle)).all():
        if vehicle.vin:
            by_vin[_vin(vehicle.vin)] = vehicle
        if vehicle.plate:
            by_plate[_plate_key(vehicle.plate)] = vehicle
        if vehicle.rentway_unit_nr:
            by_unit[_key(vehicle.rentway_unit_nr)] = vehicle
    return by_vin, by_plate, by_unit


INVOICE_EXTRACTION_ACTIONS = {
    "invoice.ocr.extracted",
    "invoice.ocr.reprocessed",
    "invoice.lines.extracted",
    "invoice.extracted",
}


def _invoice_document_metadata(db: Session, document_ids: list[int]) -> dict[int, dict[str, Any]]:
    if not document_ids:
        return {}
    metadata: dict[int, dict[str, Any]] = {}
    events = db.scalars(
        select(DocumentEvent)
        .where(
            DocumentEvent.document_id.in_(document_ids),
            DocumentEvent.action.in_(INVOICE_EXTRACTION_ACTIONS),
        )
        .order_by(DocumentEvent.created_at.desc(), DocumentEvent.id.desc())
    ).all()
    for event in events:
        if event.document_id in metadata or not event.new_value:
            continue
        try:
            payload = json.loads(event.new_value)
        except (TypeError, ValueError):
            continue
        if isinstance(payload, dict):
            metadata[event.document_id] = payload
    return metadata


def _invoice_documents(db: Session) -> list[Document]:
    return db.scalars(
        select(Document).where(
            or_(
                Document.document_type.in_(
                    {"workshop_supplier_invoice", "finance_supplier_invoice"}
                ),
                Document.title.ilike("%fatura%"),
                Document.title.ilike("%factura%"),
            )
        )
    ).all()


def _document_invoice_identity(
    document: Document,
    metadata: dict[str, Any],
) -> tuple[str, str]:
    number = (
        metadata.get("document_number")
        or metadata.get("invoice_number")
        or document.contract_number
        or document.reservation_number
        or ""
    )
    supplier_nif = (
        metadata.get("supplier_nif")
        or metadata.get("nif_fornecedor")
        or metadata.get("tax_id")
        or document.supplier_name
        or ""
    )
    return _document_key(number), _supplier_nif(supplier_nif)


def _existing_invoice_keys(db: Session) -> tuple[set[str], set[str]]:
    exact_keys: set[str] = set()
    number_only_keys: set[str] = set()
    records = db.scalars(
        select(VehicleDocumentRecord).where(
            VehicleDocumentRecord.main_group == "invoices",
            VehicleDocumentRecord.external_reference.is_not(None),
        )
    ).all()
    for record in records:
        number_key = _document_key(record.external_reference)
        if not number_key:
            continue
        nif = _supplier_nif(
            (record.metadata_json or {}).get("supplier_nif")
            or record.supplier_name
        )
        if nif:
            exact_keys.add(_invoice_key(record.external_reference, nif))
        else:
            number_only_keys.add(number_key)

    documents = _invoice_documents(db)
    metadata_by_document = _invoice_document_metadata(
        db,
        [document.id for document in documents],
    )
    for document in documents:
        number_key, supplier_nif = _document_invoice_identity(
            document,
            metadata_by_document.get(document.id, {}),
        )
        if not number_key:
            continue
        if supplier_nif:
            exact_keys.add(_invoice_key(number_key, supplier_nif))
        else:
            number_only_keys.add(number_key)
    return exact_keys, number_only_keys


def _pending_vehicle(
    row: VehicleDocumentRecord,
    by_vin: dict[str, Vehicle],
    by_plate: dict[str, Vehicle],
    by_unit: dict[str, Vehicle],
) -> tuple[Vehicle | None, str | None]:
    metadata = row.metadata_json if isinstance(row.metadata_json, dict) else {}
    vin = _vin(row.vin or metadata.get("vin"))
    plate = _plate_key(row.plate or metadata.get("plate"))
    unit = _key(metadata.get("unit_number"))
    if vin and vin in by_vin:
        return by_vin[vin], "vin"
    if plate and plate in by_plate:
        return by_plate[plate], "plate"
    if unit and unit in by_unit:
        return by_unit[unit], "unit"
    return None, None


def reconcile_pending_invoices(
    db: Session,
    *,
    user_id: int | None,
    document_ids: set[int] | None = None,
    record_ids: set[int] | None = None,
) -> dict[str, int]:
    """Link expected invoices to real documents without accepting ambiguous matches."""
    result = {
        "reviewed": 0,
        "associated": 0,
        "fulfilled": 0,
        "ambiguous": 0,
        "unmatched": 0,
    }
    pending = db.scalars(
        select(VehicleDocumentRecord).where(
            VehicleDocumentRecord.source_record_type == "pending_import",
            VehicleDocumentRecord.main_group == "invoices",
            VehicleDocumentRecord.status == "pending",
            *(
                (VehicleDocumentRecord.id.in_(record_ids),)
                if record_ids is not None
                else ()
            ),
        )
    ).all()
    if not pending:
        return result

    by_vin, by_plate, by_unit = _vehicle_maps(db)
    for record in pending:
        if record.vehicle_id is not None:
            continue
        vehicle, method = _pending_vehicle(record, by_vin, by_plate, by_unit)
        if not vehicle:
            continue
        record.vehicle_id = vehicle.id
        record.plate = vehicle.plate
        record.vin = vehicle.vin
        metadata = dict(record.metadata_json or {})
        metadata["association_method"] = method
        metadata["association_reconciled_at"] = datetime.now(UTC).isoformat()
        record.metadata_json = metadata
        record.updated_by_id = user_id
        result["associated"] += 1

    documents = _invoice_documents(db)
    if document_ids is not None:
        documents = [document for document in documents if document.id in document_ids]
    metadata_by_document = _invoice_document_metadata(
        db,
        [document.id for document in documents],
    )
    by_number: dict[str, list[tuple[Document, str]]] = {}
    for document in documents:
        number_key, supplier_nif = _document_invoice_identity(
            document,
            metadata_by_document.get(document.id, {}),
        )
        if number_key:
            by_number.setdefault(number_key, []).append((document, supplier_nif))

    used_documents = {
        int(record.document_id)
        for record in db.scalars(
            select(VehicleDocumentRecord).where(
                VehicleDocumentRecord.document_id.is_not(None),
                VehicleDocumentRecord.source_record_type == "pending_import",
            )
        ).all()
        if record.document_id is not None
    }
    for record in pending:
        result["reviewed"] += 1
        number_key = _document_key(record.external_reference)
        metadata = dict(record.metadata_json or {})
        supplier_nif = _supplier_nif(metadata.get("supplier_nif"))
        candidates = [
            (document, nif)
            for document, nif in by_number.get(number_key, [])
            if document.id not in used_documents
            and (not supplier_nif or not nif or supplier_nif == nif)
            and (
                not record.vehicle_id
                or not document.vehicle_id
                or document.vehicle_id == record.vehicle_id
            )
        ]
        exact = [
            (document, nif)
            for document, nif in candidates
            if supplier_nif and nif and supplier_nif == nif
        ]
        eligible = exact if exact else candidates
        if len(eligible) != 1:
            if len(eligible) > 1:
                result["ambiguous"] += 1
                metadata["reconciliation_state"] = "ambiguous"
                metadata["reconciliation_candidates"] = [item[0].id for item in eligible]
                record.metadata_json = metadata
            else:
                result["unmatched"] += 1
            continue

        document = eligible[0][0]
        if record.vehicle_id and not document.vehicle_id:
            document.vehicle_id = record.vehicle_id
            vehicle = db.get(Vehicle, record.vehicle_id)
            document.plate = vehicle.plate if vehicle else record.plate
            state = get_or_create_workflow_state(db, document)
            state.association_status = "associated"
        elif document.vehicle_id and not record.vehicle_id:
            record.vehicle_id = document.vehicle_id
            record.plate = document.plate

        record.document_id = document.id
        record.status = "fulfilled"
        record.has_physical_file = True
        record.storage_path = document.storage_path
        record.updated_by_id = user_id
        metadata["reconciliation_state"] = "fulfilled"
        metadata["fulfilled_document_id"] = document.id
        metadata["fulfilled_at"] = datetime.now(UTC).isoformat()
        record.metadata_json = metadata
        db.add(
            DocumentEvent(
                document_id=document.id,
                action="pending_invoice.reconciled",
                old_value=None,
                new_value=json.dumps(
                    {
                        "pending_record_id": record.id,
                        "invoice_number": record.external_reference,
                    },
                    ensure_ascii=False,
                ),
                user_id=user_id,
            )
        )
        used_documents.add(document.id)
        result["fulfilled"] += 1
    db.flush()
    return result


def _parsed_pending_rows(
    db: Session,
    *,
    path: Path,
    original_name: str,
) -> list[dict[str, Any]]:
    by_vin, by_plate, by_unit = _vehicle_maps(db)
    exact_existing_keys, number_only_existing_keys = _existing_invoice_keys(db)
    parsed_rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for index, (sheet, row_number, row) in enumerate(_rows(path), start=1):
        number = _text(
            _first(
                row,
                (
                    "Nº fatura",
                    "Numero fatura",
                    "Número fatura",
                    "Fatura",
                    "Factura",
                    "Nº documento",
                    "Numero documento",
                    "Documento",
                    "Invoice number",
                ),
            )
        )
        number_key = _document_key(number)
        supplier_name = _text(_first(row, ("Fornecedor", "Nome fornecedor", "Supplier")))
        supplier_nif = _supplier_nif(
            _first(row, ("NIF fornecedor", "NIF Fornecedor", "Supplier NIF", "VAT"))
        )
        duplicate_key = _invoice_key(number, supplier_nif)
        if not number_key:
            status = "invalid"
            status_detail = "Sem número de fatura"
        elif duplicate_key in seen:
            status = "duplicate"
            status_detail = "Repetida no ficheiro"
        elif duplicate_key in exact_existing_keys:
            status = "duplicate"
            status_detail = "Já existe para este NIF e número"
        elif number_key in number_only_existing_keys:
            status = "duplicate"
            status_detail = "Possível duplicado antigo com este número"
        else:
            status = "ready"
            status_detail = "Pronta para criar"
            seen.add(duplicate_key)

        raw_vin = _text(_first(row, ("Chassi", "Chassis", "VIN", "Referência", "Referencia")))
        vin = _vin(raw_vin)
        plate = _text(_first(row, ("Matrícula", "Matricula", "Plate", "PlateNr")))
        unit = _text(
            _first(row, ("Unit", "Unit nr", "Unit number", "Nº viatura", "Numero viatura"))
        )
        vehicle = (
            by_vin.get(vin)
            or by_plate.get(_plate_key(plate))
            or by_unit.get(_key(unit))
        )
        document_date = _date(_first(row, ("Data", "Data fatura", "Data factura", "Invoice date")))
        total = _decimal(_first(row, ("Total", "Valor", "Total com IVA", "Total c/ IVA")))
        description = _text(_first(row, ("Observações", "Observacoes", "Descrição", "Descricao")))
        association_method = (
            "vin"
            if vehicle and vin and by_vin.get(vin) is vehicle
            else "plate"
            if vehicle and plate and by_plate.get(_plate_key(plate)) is vehicle
            else "unit"
            if vehicle
            else None
        )
        parsed_rows.append(
            {
                "row_id": str(index),
                "source_sheet": sheet,
                "source_row": row_number,
                "number": number,
                "supplier_name": supplier_name,
                "supplier_nif": supplier_nif,
                "document_date": document_date.isoformat() if document_date else None,
                "total": total,
                "description": description,
                "raw_vin": raw_vin,
                "vin": vin,
                "plate": plate,
                "unit_number": unit,
                "vehicle_id": vehicle.id if vehicle else None,
                "vehicle_plate": vehicle.plate if vehicle else None,
                "vehicle_label": (
                    f"{vehicle.plate} · Unit {vehicle.rentway_unit_nr}"
                    if vehicle and vehicle.rentway_unit_nr
                    else vehicle.plate
                    if vehicle
                    else None
                ),
                "association_method": association_method,
                "status": status,
                "status_detail": status_detail,
                "selected": status == "ready",
                "original_name": original_name,
            }
        )
    return parsed_rows


def preview_pending_documents(
    db: Session,
    *,
    path: Path,
    original_name: str,
) -> dict[str, Any]:
    rows = _parsed_pending_rows(db, path=path, original_name=original_name)
    return {
        "schema": "carfast.pending-invoice-preview.v1",
        "original_name": original_name,
        "rows": rows,
        "summary": {
            "total": len(rows),
            "ready": sum(row["status"] == "ready" for row in rows),
            "associated": sum(
                row["status"] == "ready" and row["vehicle_id"] is not None for row in rows
            ),
            "unmatched": sum(
                row["status"] == "ready" and row["vehicle_id"] is None for row in rows
            ),
            "duplicates": sum(row["status"] == "duplicate" for row in rows),
            "invalid": sum(row["status"] == "invalid" for row in rows),
        },
    }


def create_pending_documents_from_preview(
    db: Session,
    *,
    preview: dict[str, Any],
    selected_row_ids: set[str],
    user_id: int | None,
) -> dict[str, int]:
    result = {"created": 0, "associated": 0, "unmatched": 0, "duplicates": 0, "invalid": 0}
    exact_existing_keys, number_only_existing_keys = _existing_invoice_keys(db)
    selected_rows = [
        row
        for row in preview.get("rows", [])
        if str(row.get("row_id")) in selected_row_ids
    ]
    seen: set[str] = set()
    for row in selected_rows:
        number = _text(row.get("number"))
        number_key = _document_key(number)
        supplier_nif = _supplier_nif(row.get("supplier_nif"))
        duplicate_key = _invoice_key(number, supplier_nif)
        if not number_key or row.get("status") == "invalid":
            result["invalid"] += 1
            continue
        if (
            duplicate_key in seen
            or duplicate_key in exact_existing_keys
            or number_key in number_only_existing_keys
        ):
            result["duplicates"] += 1
            continue
        seen.add(duplicate_key)

        vehicle_id = row.get("vehicle_id")
        vehicle = db.get(Vehicle, int(vehicle_id)) if vehicle_id else None
        record = VehicleDocumentRecord(
            vehicle_id=vehicle.id if vehicle else None,
            source_record_type="pending_import",
            main_group="invoices",
            status="pending",
            external_reference=number,
            title=f"Fatura pendente {number}",
            plate=vehicle.plate if vehicle else _text(row.get("plate")) or None,
            vin=vehicle.vin if vehicle else _text(row.get("vin")) or None,
            supplier_name=_text(row.get("supplier_name")) or None,
            raw_description=_text(row.get("description")) or None,
            document_date=_date(row.get("document_date")),
            has_physical_file=False,
            source_system="pending_document_import",
            metadata_json={
                "supplier_nif": supplier_nif or None,
                "unit_number": _text(row.get("unit_number")) or None,
                "expected_total": row.get("total"),
                "source_file": preview.get("original_name"),
                "source_sheet": row.get("source_sheet"),
                "source_row": row.get("source_row"),
                "association_method": row.get("association_method"),
            },
            created_by_id=user_id,
            updated_by_id=user_id,
        )
        db.add(record)
        exact_existing_keys.add(duplicate_key)
        result["created"] += 1
        result["associated" if vehicle else "unmatched"] += 1
    db.flush()
    return result


def import_pending_documents(
    db: Session,
    *,
    path: Path,
    original_name: str,
    user_id: int | None,
) -> dict[str, int]:
    """Compatibility wrapper for non-interactive callers and existing tests."""
    preview = preview_pending_documents(db, path=path, original_name=original_name)
    selected = {
        str(row["row_id"])
        for row in preview["rows"]
        if row["status"] == "ready"
    }
    result = create_pending_documents_from_preview(
        db,
        preview=preview,
        selected_row_ids=selected,
        user_id=user_id,
    )
    result["duplicates"] += preview["summary"]["duplicates"]
    result["invalid"] += preview["summary"]["invalid"]
    return result
