from __future__ import annotations

import hashlib
import json
import shutil
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.imports import ImportBatch, ImportError, ImportFile, ImportRawRow
from app.models.vehicles import Vehicle, VehicleManualField
from app.services.audit import record_audit
from app.services.storage import persistent_import_storage_root


TRADE_DEBT_IMPORT_TYPE = "trade_debt"
TRADE_DEBT_SOURCE_SYSTEM = "carfast_sales_map"
TRADE_DEBT_SHEET = "Mapa_Base"
FIELD_DEBT_VALUE = "debt_value"


def trade_debt_storage_root() -> Path:
    return persistent_import_storage_root("trade_debt")


def store_trade_debt_upload(source_path: Path, original_name: str) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    safe_name = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in original_name)[:120]
    target = trade_debt_storage_root() / "pending" / f"{timestamp}_{safe_name}"
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source_path, target)
    return target


def normalize_plate(value: object) -> str:
    return str(value or "").strip().upper().replace(" ", "")


def compact_plate(value: object) -> str:
    return "".join(ch for ch in normalize_plate(value) if ch.isalnum())


def normalize_decimal(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float):
        return f"{Decimal(str(value)).quantize(Decimal('0.01'))}"
    if isinstance(value, int):
        return f"{Decimal(value).quantize(Decimal('0.01'))}"
    text = str(value).strip().replace(" ", "").replace("\u00a0", "")
    if not text:
        return ""
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return f"{Decimal(text).quantize(Decimal('0.01'))}"
    except InvalidOperation:
        return ""


def row_hash(raw: dict[str, Any]) -> str:
    raw_json = json.dumps(raw, ensure_ascii=False, default=str, sort_keys=True)
    return hashlib.sha1(raw_json.encode("utf-8")).hexdigest()


def load_trade_debt_rows(path: str | Path) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if TRADE_DEBT_SHEET not in workbook.sheetnames:
            raise ValueError(f"Folha em falta: {TRADE_DEBT_SHEET}")
        sheet = workbook[TRADE_DEBT_SHEET]
        headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
        index = {name: position + 1 for position, name in enumerate(headers) if name}
        required = {"matricula", "divida_com_iva"}
        missing_headers = sorted(required - set(index))
        if missing_headers:
            raise ValueError(f"Colunas em falta: {', '.join(missing_headers)}")

        rows: list[dict[str, Any]] = []
        for row_number in range(2, sheet.max_row + 1):
            raw = {header: sheet.cell(row_number, index[header]).value for header in index}
            plate = normalize_plate(raw.get("matricula"))
            if not plate:
                continue
            debt_value = normalize_decimal(raw.get("divida_com_iva"))
            finance_entity = str(raw.get("entidade_divida") or "").strip()
            rows.append(
                {
                    "row_number": row_number,
                    "raw": raw,
                    "hash": row_hash(raw),
                    "plate": plate,
                    "compact_plate": compact_plate(plate),
                    "debt_value": debt_value,
                    "finance_entity": finance_entity,
                }
            )
        return headers, rows
    finally:
        workbook.close()


def vehicle_lookup(db: Session) -> dict[str, Vehicle]:
    lookup: dict[str, Vehicle] = {}
    for vehicle in db.scalars(select(Vehicle).where(Vehicle.plate.is_not(None))).all():
        if vehicle.plate:
            lookup[normalize_plate(vehicle.plate)] = vehicle
            lookup[compact_plate(vehicle.plate)] = vehicle
    return lookup


def manual_values_for_vehicle(db: Session, vehicle_id: int) -> dict[str, Any]:
    rows = db.scalars(
        select(VehicleManualField).where(
            VehicleManualField.vehicle_id == vehicle_id,
            VehicleManualField.field_code == FIELD_DEBT_VALUE,
        )
    ).all()
    return {row.field_code: row.value_json for row in rows}


def preview_trade_debt_import(db: Session, path: str | Path) -> dict[str, Any]:
    headers, source_rows = load_trade_debt_rows(path)
    vehicles = vehicle_lookup(db)
    rows = []
    seen_plates: set[str] = set()
    for source_row in source_rows:
        vehicle = vehicles.get(source_row["plate"]) or vehicles.get(source_row["compact_plate"])
        state = "valid"
        message = ""
        if source_row["plate"] in seen_plates:
            state = "duplicate"
            message = "Matrícula repetida no ficheiro; a linha será ignorada."
        elif not vehicle:
            state = "error"
            message = "Viatura não encontrada na Frota."
        seen_plates.add(source_row["plate"])

        current = manual_values_for_vehicle(db, vehicle.id) if vehicle else {}
        debt_changed = vehicle is not None and str(current.get(FIELD_DEBT_VALUE) or "") != source_row["debt_value"]
        rows.append(
            {
                **source_row,
                "vehicle_id": vehicle.id if vehicle else None,
                "current_debt_value": current.get(FIELD_DEBT_VALUE) or "",
                "debt_changed": debt_changed,
                "state": state,
                "message": message,
            }
        )

    total = len(rows)
    valid = sum(1 for row in rows if row["state"] == "valid")
    errors = sum(1 for row in rows if row["state"] == "error")
    duplicates = sum(1 for row in rows if row["state"] == "duplicate")
    with_debt = sum(1 for row in rows if row["state"] == "valid" and row["debt_value"])
    changed = sum(1 for row in rows if row["state"] == "valid" and row["debt_changed"])
    return {
        "sheet_name": TRADE_DEBT_SHEET,
        "headers": headers,
        "rows": rows,
        "summary": {
            "total": total,
            "valid": valid,
            "errors": errors,
            "duplicates": duplicates,
            "with_debt": with_debt,
            "changed": changed,
        },
    }


def upsert_manual_field(db: Session, vehicle_id: int, field_code: str, value: str, user_id: int | None) -> bool:
    field = db.scalar(
        select(VehicleManualField).where(
            VehicleManualField.vehicle_id == vehicle_id,
            VehicleManualField.field_code == field_code,
        )
    )
    if field:
        if str(field.value_json or "") == value:
            return False
        field.value_json = value
        field.updated_by_id = user_id
        return True
    db.add(
        VehicleManualField(
            vehicle_id=vehicle_id,
            field_code=field_code,
            value_json=value,
            updated_by_id=user_id,
        )
    )
    return True


def apply_trade_debt_import(
    db: Session,
    path: str | Path,
    original_name: str,
    *,
    user_id: int | None,
) -> dict[str, Any]:
    preview = preview_trade_debt_import(db, path)
    batch = ImportBatch(
        source_system=TRADE_DEBT_SOURCE_SYSTEM,
        import_type=TRADE_DEBT_IMPORT_TYPE,
        status="running",
        imported_by_id=user_id,
        total_rows=preview["summary"]["total"],
        detail="Importação pontual de valor em dívida para Gestão CarFast.",
    )
    db.add(batch)
    db.flush()

    stored_path = trade_debt_storage_root() / f"batch_{batch.id}_{Path(path).name}"
    shutil.copyfile(path, stored_path)
    db.add(
        ImportFile(
            batch_id=batch.id,
            original_name=original_name,
            file_name=stored_path.name,
            storage_path=str(stored_path),
            sheet_name=preview["sheet_name"],
            columns_json=preview["headers"],
        )
    )

    updated_rows = 0
    skipped_rows = 0
    error_rows = 0
    updated_vehicle_ids: list[int] = []
    for row in preview["rows"]:
        db.add(
            ImportRawRow(
                batch_id=batch.id,
                row_number=row["row_number"],
                external_reference=row["plate"],
                raw_json=row["raw"],
                row_hash=row["hash"],
            )
        )
        if row["state"] == "error":
            error_rows += 1
            db.add(
                ImportError(
                    batch_id=batch.id,
                    row_number=row["row_number"],
                    entity_type="vehicle",
                    error_message=row["message"],
                    raw_json=row["raw"],
                )
            )
            continue
        if row["state"] == "duplicate":
            skipped_rows += 1
            db.add(
                ImportError(
                    batch_id=batch.id,
                    row_number=row["row_number"],
                    entity_type="trade_debt_duplicate",
                    error_message=row["message"],
                    raw_json=row["raw"],
                )
            )
            continue
        vehicle_id = row["vehicle_id"]
        debt_changed = upsert_manual_field(db, vehicle_id, FIELD_DEBT_VALUE, row["debt_value"], user_id)
        if debt_changed:
            updated_rows += 1
            updated_vehicle_ids.append(vehicle_id)

    batch.status = "completed" if error_rows == 0 else "completed_with_errors"
    batch.created_rows = 0
    batch.updated_rows = updated_rows
    batch.skipped_rows = skipped_rows
    batch.error_rows = error_rows
    batch.finished_at = datetime.now(UTC)
    batch.detail = f"{updated_rows} viaturas atualizadas; {preview['summary']['with_debt']} com dívida no ficheiro."
    record_audit(
        db,
        action="import.trade_debt.completed",
        entity_type="import_batch",
        entity_id=batch.id,
        detail=batch.detail,
        after_json={
            "updated_vehicle_ids": updated_vehicle_ids[:200],
            "with_debt": preview["summary"]["with_debt"],
            "skipped_rows": skipped_rows,
            "error_rows": error_rows,
        },
        user_id=user_id,
    )
    db.commit()
    return {
        "batch_id": batch.id,
        "updated_rows": updated_rows,
        "skipped_rows": skipped_rows,
        "error_rows": error_rows,
        "with_debt": preview["summary"]["with_debt"],
    }
