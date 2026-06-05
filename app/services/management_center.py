import csv
import hashlib
import io
import json
import re
import shutil
from collections import Counter
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.imports import ImportBatch, ImportError, ImportFile, ImportRawRow
from app.models.management_center import (
    ClaimIncident,
    ClaimRefstroLine,
    ClaimRentwayAR,
    ManagementAction,
    ManagementHistory,
    ManagementProcess,
    ManagementProcessAssociation,
    ManagementProcessType,
    ManagementRule,
)
from app.services.audit import record_audit
from app.services.spreadsheets import (
    build_column_lookup,
    clean_text,
    excel_date_to_iso,
    first_row_value,
    iter_xlsx_rows,
    normalize_header,
)

MANAGEMENT_CENTER_TYPE_CODE = "claims_ar"
MANAGEMENT_CENTER_TYPE_NAME = "Sinistros"
MANAGEMENT_CENTER_SOURCE_SYSTEM = "carfast_management_center"
AR_IMPORT_TYPE = "claims_ar_rentway_ar"
CRAR_PER_VEHICLE_IMPORT_TYPE = "claims_ar_rentway_per_vehicle"
REFSTRO_IMPORT_TYPE = "claims_ar_refstro"
MANAGEMENT_STORAGE_DIR = Path("data/imports/management_center")

CARFAST_CLAIM_PHASES = [
    "1.0 ABERTO",
    "1.1 ABERTO SEM DAAA",
    "1.2 ABERTO COM DAAA",
    "1.3 ABERTO RECL COMPANHIA",
    "1.4 AV VALID ACIDENTE",
    "2.0 PEND ENVIO PARTICIPAÇ",
    "2.1 PEND FEEDB MEDIADOR",
    "2.2 PEND MARCAÇÃO PERITAG",
    "2.3 PEND RELAT. PERITAGEM",
    "3.0 REPARAÇÃO POR AUTORIZ",
    "3.1 REPARAÇÃO EM CURSO",
    "3.2 REPAR. VALI FECHAR FO",
    "4.0 COBRAR INDEMNIZAÇÃO",
    "4.1 COBRAR PARALIZAÇÃO",
    "5.0 FECHADO SEM REPARAÇÃO",
    "5.1 FECHADO COM REPARAÇÃO",
    "5.2 FECHADO SEM PARTICIPA",
]

PROCESS_STATUS_LABELS = {
    "open": "Aberto",
    "waiting": "A aguardar",
    "information_request": "Pedido de informação",
    "analysis": "Em análise",
    "internal_closed": "Fechado internamente",
}

PROCESS_PHASE_LABELS = {
    "information_request": "Pedido de informação",
    "ar_missing": "AR em falta",
    "analysis": "Em análise",
    "rentway_closed_review": "Fechado Rentway em validação interna",
    "internal_closed": "Fechado internamente",
}

ACTION_STATUS_LABELS = {
    "open": "Aberta",
    "done": "Concluída",
    "cancelled": "Cancelada",
}

CLAIM_COMPONENT_ALIASES = {
    "rc": "RC",
    "responsabilidade civil": "RC",
    "dp": "DP",
    "danos proprios": "DP",
    "danos próprios": "DP",
    "ids credor": "IDS Credor",
    "ids": "IDS Credor",
    "vidros": "Vidros",
    "glass": "Vidros",
    "custos de gestao": "Custos de Gestão",
    "custos de gestão": "Custos de Gestão",
    "gestao": "Custos de Gestão",
    "gestão": "Custos de Gestão",
}

MANAGEMENT_RULE_DEFINITIONS = [
    (
        "internal_sin_reference",
        "Referência SIN obrigatória",
        "Cada sinistro acompanhado pela CarFast tem uma referência interna única independente do AR.",
        "info",
    ),
    (
        "missing_ar_action",
        "AR em falta exige ação",
        "Quando há sinistro identificado sem AR ativo associado, o SIN fica em AR em falta e abre ação obrigatória.",
        "critical",
    ),
    (
        "status_uses_ar_status",
        "Status operacional usa Status",
        "O campo Status dos ARs alimenta a fase operacional; o campo state do export fica apenas como bruto.",
        "warning",
    ),
    (
        "rentway_closed_not_internal",
        "Fecho Rentway não fecha internamente",
        "Um AR fechado no Rentway passa a validação interna, mas não fecha automaticamente o SIN CarFast.",
        "warning",
    ),
    (
        "same_plate_date_consolidates",
        "Matrícula e data consolidam",
        "Linhas REFSTRO com a mesma matrícula e data são consolidadas num único SIN com componentes separados.",
        "info",
    ),
    (
        "minimum_data_information_request",
        "Sem mínimos há pedido de informação",
        "Sem matrícula ou data suficientes não há alerta operacional; o SIN fica como pedido de informação.",
        "info",
    ),
]


def normalize_plate(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return text.upper().replace(" ", "").replace("-", "")


def normalize_component(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return CLAIM_COMPONENT_ALIASES.get(text.casefold(), text[:120])


def parse_date_value(value: Any) -> date | None:
    iso = excel_date_to_iso(value)
    if not iso:
        return None
    try:
        return date.fromisoformat(iso[:10])
    except ValueError:
        return None


def parse_decimal_value(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("€", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def row_hash(raw: dict[str, Any]) -> str:
    raw_json = json.dumps(raw, ensure_ascii=False, default=str, sort_keys=True)
    return hashlib.sha1(raw_json.encode("utf-8")).hexdigest()


def management_storage_root() -> Path:
    MANAGEMENT_STORAGE_DIR.mkdir(parents=True, exist_ok=True)
    return MANAGEMENT_STORAGE_DIR


def store_management_upload(source_path: Path, original_name: str) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    safe_name = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in original_name)[:120]
    target = management_storage_root() / "pending" / f"{timestamp}_{safe_name}"
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source_path, target)
    return target


def iter_management_rows(path: str | Path):
    file_path = Path(path)
    if file_path.suffix.lower() == ".csv":
        raw_bytes = file_path.read_bytes()
        try:
            text = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw_bytes.decode("cp1252")
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        for row_number, raw in enumerate(reader, start=2):
            row = tuple(raw.get(header) for header in headers)
            yield "CSV", headers, row_number, row, raw
        return
    yield from iter_xlsx_rows(file_path)


def preview_header_score(headers: list[Any], import_kind: str) -> int:
    normalized = {normalize_header(header) for header in headers}
    if import_kind == "ar":
        expected = {
            "accidentreportid",
            "accidentreport",
            "ar",
            "status",
            "requestdate",
            "accidentdate",
            "platenumber",
            "matricula",
            "documentno",
        }
    elif import_kind == "ar_rentway_per_vehicle":
        expected = {
            "accidentreport",
            "n",
            "nmanual",
            "ndocumento",
            "datadoacidente",
            "declaracaoamigavel",
            "condutor",
            "cliente",
        }
    else:
        expected = {
            "refstro",
            "matricula",
            "dtastro",
            "data sinistro",
            "rc",
            "dp",
            "idscredor",
            "vidros",
            "custosgestao",
            "custototal",
        }
    return sum(1 for key in expected if key in normalized)


def parse_crar_vehicle_line(value: Any) -> dict[str, str | None] | None:
    text = clean_text(value)
    if not text:
        return None
    match = re.match(r"(?P<unit>\d+)\s*-\s*(?P<plate>[A-Z0-9-]+)\s*:\s*(?P<model>.+)", text, flags=re.I)
    if not match:
        return None
    return {
        "vehicle_reference": clean_text(match.group("unit")),
        "plate": normalize_plate(match.group("plate")),
        "vehicle_model": clean_text(match.group("model")),
    }


def iter_crar_per_vehicle_rows(path: str | Path):
    file_path = Path(path)
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    headers = [
        "Accident report / Nº",
        "Accident report / Nº manual",
        "Nº documento / Tipo",
        "Nº documento / Nº",
        "Data do acidente",
        "Declaração amigável",
        "Condutor",
        "Condutor / Nº",
        "Cliente",
        "Veículo / Nº",
        "Veículo / Matrícula",
        "Veículo / Modelo",
    ]
    try:
        sheet = workbook[workbook.sheetnames[0]]
        current_vehicle = {"vehicle_reference": None, "plate": None, "vehicle_model": None}
        for row_number, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
            vehicle = parse_crar_vehicle_line(row[3] if len(row) > 3 else None)
            if vehicle and clean_text(row[2] if len(row) > 2 else None) == "Veículo :":
                current_vehicle = vehicle
                continue
            ar_reference = clean_text(row[0] if row else None)
            if not ar_reference or not ar_reference.isdigit():
                continue
            raw = {
                "Accident report / Nº": ar_reference,
                "Accident report / Nº manual": json_safe_preview_value(row[2] if len(row) > 2 else None),
                "Nº documento / Tipo": json_safe_preview_value(row[3] if len(row) > 3 else None),
                "Nº documento / Nº": json_safe_preview_value(row[4] if len(row) > 4 else None),
                "Data do acidente": json_safe_preview_value(row[5] if len(row) > 5 else None),
                "Declaração amigável": json_safe_preview_value(row[6] if len(row) > 6 else None),
                "Condutor": json_safe_preview_value(row[7] if len(row) > 7 else None),
                "Condutor / Nº": json_safe_preview_value(row[10] if len(row) > 10 else None),
                "Cliente": json_safe_preview_value(row[11] if len(row) > 11 else None),
                "Veículo / Nº": current_vehicle.get("vehicle_reference"),
                "Veículo / Matrícula": current_vehicle.get("plate"),
                "Veículo / Modelo": current_vehicle.get("vehicle_model"),
            }
            yield sheet.title, headers, row_number, tuple(raw.get(header) for header in headers), raw
    finally:
        workbook.close()


def json_safe_preview_value(value: Any) -> Any:
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def iter_management_preview_rows(path: str | Path, import_kind: str):
    file_path = Path(path)
    if import_kind == "ar_rentway_per_vehicle":
        yield from iter_crar_per_vehicle_rows(file_path)
        return
    if file_path.suffix.lower() == ".csv":
        yield from iter_management_rows(file_path)
        return
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        candidate_rows = []
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True),
            start=1,
        ):
            headers = [str(value).strip() if value is not None else "" for value in row]
            score = preview_header_score(headers, import_kind)
            if score:
                candidate_rows.append((score, row_number, headers))
        if candidate_rows:
            _, header_row, headers = sorted(candidate_rows, key=lambda item: (-item[0], item[1]))[0]
        else:
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
            header_row = 1
        for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not any(value not in (None, "") for value in row):
                continue
            raw = {
                headers[idx] or f"coluna_{idx + 1}": json_safe_preview_value(value)
                for idx, value in enumerate(row)
                if idx < len(headers)
            }
            yield sheet.title, headers, row_number, row, raw
    finally:
        workbook.close()


def mapped_value(row: tuple[Any, ...], col: dict[str, int], aliases: list[str]) -> Any:
    normalized_aliases = [normalize_header(alias) for alias in aliases]
    return first_row_value(row, col, [*aliases, *normalized_aliases])


def ensure_management_defaults(db: Session) -> ManagementProcessType:
    process_type = db.scalar(
        select(ManagementProcessType).where(ManagementProcessType.code == MANAGEMENT_CENTER_TYPE_CODE)
    )
    if not process_type:
        process_type = ManagementProcessType(
            code=MANAGEMENT_CENTER_TYPE_CODE,
            name=MANAGEMENT_CENTER_TYPE_NAME,
            description="Acompanhamento por sinistro/SIN; ARs Rentway e linhas REFSTRO são dados associados.",
            active=True,
        )
        db.add(process_type)
        db.flush()
    else:
        process_type.name = MANAGEMENT_CENTER_TYPE_NAME
        process_type.description = "Acompanhamento por sinistro/SIN; ARs Rentway e linhas REFSTRO são dados associados."
    existing_rules = {
        rule.code
        for rule in db.scalars(select(ManagementRule).where(ManagementRule.process_type_id == process_type.id))
    }
    for code, title, description, severity in MANAGEMENT_RULE_DEFINITIONS:
        if code not in existing_rules:
            db.add(
                ManagementRule(
                    process_type_id=process_type.id,
                    code=code,
                    title=title,
                    description=description,
                    severity=severity,
                    active=True,
                )
            )
    return process_type


def next_sin_reference(db: Session, reference_year: int) -> str:
    sequence = (
        db.scalar(
            select(func.count())
            .select_from(ClaimIncident)
            .where(ClaimIncident.sin_reference.like(f"SIN-{reference_year}-%"))
        )
        or 0
    ) + 1
    while True:
        reference = f"SIN-{reference_year}-{sequence:06d}"
        if not db.scalar(select(ClaimIncident).where(ClaimIncident.sin_reference == reference)):
            return reference
        sequence += 1


def add_history(
    db: Session,
    process_id: int,
    *,
    action: str,
    entity_type: str | None = None,
    entity_id: str | int | None = None,
    old_value: str | None = None,
    new_value: str | None = None,
    detail: str | None = None,
    user_id: int | None = None,
) -> None:
    db.add(
        ManagementHistory(
            process_id=process_id,
            user_id=user_id,
            action=action,
            entity_type=entity_type,
            entity_id=str(entity_id) if entity_id is not None else None,
            old_value=old_value,
            new_value=new_value,
            detail=detail,
        )
    )


def find_claim_by_plate_date(db: Session, plate: str | None, accident_date: date | None) -> ClaimIncident | None:
    if not plate or not accident_date:
        return None
    return db.scalar(
        select(ClaimIncident).where(
            ClaimIncident.plate == plate,
            ClaimIncident.accident_date == accident_date,
        )
    )


def create_claim_process(
    db: Session,
    process_type: ManagementProcessType,
    *,
    plate: str | None,
    accident_date: date | None,
    customer_name: str | None = None,
    driver_name: str | None = None,
    document_reference: str | None = None,
    user_id: int | None = None,
) -> ClaimIncident:
    reference_year = (accident_date or date.today()).year
    sin_reference = next_sin_reference(db, reference_year)
    title_subject = plate or document_reference or "sem dados mínimos"
    process = ManagementProcess(
        process_type_id=process_type.id,
        internal_reference=sin_reference,
        title=f"Sinistro / AR {title_subject}",
        status="information_request" if not plate or not accident_date else "open",
        phase="information_request" if not plate or not accident_date else "analysis",
        priority="normal",
        plate=plate,
        document_reference=document_reference,
        customer_name=customer_name,
        driver_name=driver_name,
        pending_reason="missing_minimum_data" if not plate or not accident_date else None,
        pending_detail="Completar matrícula e data do sinistro." if not plate or not accident_date else None,
        opened_on=accident_date or date.today(),
        sla_due_on=(accident_date or date.today()) + timedelta(days=2),
    )
    db.add(process)
    db.flush()
    claim = ClaimIncident(
        process_id=process.id,
        sin_reference=sin_reference,
        accident_date=accident_date,
        plate=plate,
        operational_status=process.phase,
        has_missing_minimum_data=not bool(plate and accident_date),
        has_missing_ar=False,
        components_json={},
    )
    db.add(claim)
    db.flush()
    add_history(
        db,
        process.id,
        action="process.created",
        entity_type="claim_incident",
        entity_id=claim.id,
        new_value=sin_reference,
        detail="Processo SIN criado.",
        user_id=user_id,
    )
    return claim


def get_or_create_claim(
    db: Session,
    process_type: ManagementProcessType,
    *,
    plate: str | None,
    accident_date: date | None,
    customer_name: str | None = None,
    driver_name: str | None = None,
    document_reference: str | None = None,
    user_id: int | None = None,
) -> ClaimIncident:
    claim = find_claim_by_plate_date(db, plate, accident_date)
    if claim:
        process = db.get(ManagementProcess, claim.process_id)
        changed = False
        if process and customer_name and not process.customer_name:
            process.customer_name = customer_name
            changed = True
        if process and driver_name and not process.driver_name:
            process.driver_name = driver_name
            changed = True
        if process and document_reference and not process.document_reference:
            process.document_reference = document_reference
            changed = True
        if changed and process:
            add_history(
                db,
                process.id,
                action="process.enriched",
                entity_type="claim_incident",
                entity_id=claim.id,
                detail="Dados do SIN enriquecidos por importação.",
                user_id=user_id,
            )
        return claim
    return create_claim_process(
        db,
        process_type,
        plate=plate,
        accident_date=accident_date,
        customer_name=customer_name,
        driver_name=driver_name,
        document_reference=document_reference,
        user_id=user_id,
    )


def active_association_exists(db: Session, process_id: int, entity_type: str, entity_id: int) -> bool:
    return bool(
        db.scalar(
            select(ManagementProcessAssociation).where(
                ManagementProcessAssociation.process_id == process_id,
                ManagementProcessAssociation.entity_type == entity_type,
                ManagementProcessAssociation.entity_id == entity_id,
                ManagementProcessAssociation.active.is_(True),
            )
        )
    )


def associate_to_process(
    db: Session,
    process_id: int,
    *,
    entity_type: str,
    entity_id: int,
    reason: str,
    user_id: int | None = None,
) -> None:
    if active_association_exists(db, process_id, entity_type, entity_id):
        return
    association = ManagementProcessAssociation(
        process_id=process_id,
        entity_type=entity_type,
        entity_id=entity_id,
        association_role="source",
        active=True,
        reason=reason,
        created_by_id=user_id,
    )
    db.add(association)
    db.flush()
    add_history(
        db,
        process_id,
        action="association.created",
        entity_type=entity_type,
        entity_id=entity_id,
        new_value="active",
        detail=reason,
        user_id=user_id,
    )


def end_association(
    db: Session,
    association: ManagementProcessAssociation,
    *,
    reason: str,
    user_id: int | None,
) -> None:
    association.active = False
    association.ended_at = datetime.now(UTC)
    association.ended_by_id = user_id
    association.reason = reason
    add_history(
        db,
        association.process_id,
        action="association.ended",
        entity_type=association.entity_type,
        entity_id=association.entity_id,
        old_value="active",
        new_value="inactive",
        detail=reason,
        user_id=user_id,
    )


def rule_by_code(db: Session, process_type_id: int, code: str) -> ManagementRule | None:
    return db.scalar(
        select(ManagementRule).where(
            ManagementRule.process_type_id == process_type_id,
            ManagementRule.code == code,
            ManagementRule.active.is_(True),
        )
    )


def ensure_action(
    db: Session,
    process: ManagementProcess,
    *,
    rule_code: str,
    title: str,
    description: str,
    mandatory: bool,
) -> None:
    rule = rule_by_code(db, process.process_type_id, rule_code)
    existing = db.scalar(
        select(ManagementAction).where(
            ManagementAction.process_id == process.id,
            ManagementAction.title == title,
            ManagementAction.status == "open",
        )
    )
    if existing:
        return
    db.add(
        ManagementAction(
            process_id=process.id,
            rule_id=rule.id if rule else None,
            title=title,
            description=description,
            status="open",
            mandatory=mandatory,
            due_on=date.today() + timedelta(days=1),
        )
    )
    add_history(
        db,
        process.id,
        action="action.created",
        entity_type="management_action",
        new_value=title,
        detail=description,
    )


def close_action_by_title(db: Session, process_id: int, title: str) -> None:
    actions = db.scalars(
        select(ManagementAction).where(
            ManagementAction.process_id == process_id,
            ManagementAction.title == title,
            ManagementAction.status == "open",
        )
    ).all()
    for action in actions:
        action.status = "done"
        action.completed_at = datetime.now(UTC)
        add_history(
            db,
            process_id,
            action="action.completed",
            entity_type="management_action",
            entity_id=action.id,
            old_value="open",
            new_value="done",
            detail=title,
        )


def refresh_claim_state(db: Session, claim: ClaimIncident) -> None:
    process = db.get(ManagementProcess, claim.process_id)
    if not process:
        return
    associations = db.scalars(
        select(ManagementProcessAssociation).where(
            ManagementProcessAssociation.process_id == process.id,
            ManagementProcessAssociation.active.is_(True),
        )
    ).all()
    ar_ids = [item.entity_id for item in associations if item.entity_type == "claim_rentway_ar"]
    refstro_ids = [item.entity_id for item in associations if item.entity_type == "claim_refstro_line"]
    ars = db.scalars(select(ClaimRentwayAR).where(ClaimRentwayAR.id.in_(ar_ids))).all() if ar_ids else []
    refstros = (
        db.scalars(select(ClaimRefstroLine).where(ClaimRefstroLine.id.in_(refstro_ids))).all()
        if refstro_ids
        else []
    )
    total_claim = sum((item.claim_value or Decimal("0.00")) for item in refstros)
    total_cost = sum((item.cost_value or Decimal("0.00")) for item in refstros)
    components = sorted({item.component for item in refstros if item.component})
    claim.components_json = {"components": components}
    claim.has_missing_minimum_data = not bool(claim.plate and claim.accident_date)
    claim.has_missing_ar = bool(not claim.has_missing_minimum_data and not ars)
    process.total_claim_value = total_claim
    process.total_cost_value = total_cost
    process.raw_summary_json = {
        "ar_count": len(ars),
        "refstro_count": len(refstros),
        "components": components,
    }

    if claim.has_missing_minimum_data:
        process.status = "information_request"
        process.phase = "information_request"
        process.pending_reason = "missing_minimum_data"
        process.pending_detail = "Completar matrícula e data do sinistro."
        claim.operational_status = "information_request"
        ensure_action(
            db,
            process,
            rule_code="minimum_data_information_request",
            title="Completar dados mínimos do sinistro",
            description="Pedir matrícula e data antes de criar alertas operacionais.",
            mandatory=False,
        )
    elif claim.has_missing_ar:
        process.status = "waiting"
        process.phase = "ar_missing"
        process.pending_reason = "missing_ar"
        process.pending_detail = "Criar ou associar AR Rentway ao SIN."
        claim.operational_status = "ar_missing"
        ensure_action(
            db,
            process,
            rule_code="missing_ar_action",
            title="Criar ou associar AR",
            description="Este sinistro tem dados mínimos, mas não tem AR associado.",
            mandatory=True,
        )
        close_action_by_title(db, process.id, "Completar dados mínimos do sinistro")
    else:
        latest_status = next((ar.status for ar in ars if ar.status), None)
        claim.rentway_status = latest_status
        claim.operational_status = "rentway_closed_review" if latest_status and "fech" in latest_status.casefold() else "analysis"
        process.status = "analysis"
        process.phase = claim.operational_status
        process.pending_reason = None
        process.pending_detail = None
        close_action_by_title(db, process.id, "Criar ou associar AR")
        close_action_by_title(db, process.id, "Completar dados mínimos do sinistro")


def build_ar_payload(
    row: tuple[Any, ...],
    col: dict[str, int],
    raw: dict[str, Any],
    *,
    source_file: str,
    row_number: int,
) -> dict[str, Any]:
    return {
        "ar_reference": clean_text(
            mapped_value(
                row,
                col,
                ["AR", "ar", "accidentReportID", "accidentReport", "accident_report", "reportNumber", "nrAR", "sinistro"],
            )
        ),
        "status": clean_text(mapped_value(row, col, ["Status", "status"])),
        "raw_state": clean_text(mapped_value(row, col, ["state", "State"])),
        "request_date": parse_date_value(mapped_value(row, col, ["requestDate", "request_date", "data", "data AR", "Data AR"])),
        "plate": normalize_plate(
            mapped_value(row, col, ["matricula", "matrícula", "plate", "plateNr", "plateNumber", "licensePlate"])
        ),
        "vehicle_reference": clean_text(mapped_value(row, col, ["viatura", "vehicle", "unit", "unitNo", "unitNr", "rentwayUnitNr"])),
        "driver_name": clean_text(mapped_value(row, col, ["condutor", "driver", "driverName"])),
        "customer_name": clean_text(mapped_value(row, col, ["cliente", "customer", "customerName"])),
        "ra_reference": clean_text(mapped_value(row, col, ["RA", "ra", "rentalAgreement", "reservation"])),
        "impro_reference": clean_text(mapped_value(row, col, ["IMPRO", "impro"])),
        "daaa_reference": clean_text(mapped_value(row, col, ["DAAA", "daaa"])),
        "insurance_policy": clean_text(mapped_value(row, col, ["insurancePolicy", "insurance_policy", "apolice", "apólice"])),
        "rental_station_out": clean_text(mapped_value(row, col, ["rentalStationOut", "stationOut", "estacaoSaida"])),
        "created_by_rental_station": clean_text(
            mapped_value(row, col, ["createdByRentalSation", "createdByRentalStation", "created_by_rental_station"])
        ),
        "source_file": source_file,
        "source_row_number": row_number,
        "raw_json": raw,
    }


def build_crar_per_vehicle_payload(
    row: tuple[Any, ...],
    col: dict[str, int],
    raw: dict[str, Any],
    *,
    source_file: str,
    row_number: int,
) -> dict[str, Any]:
    document_type = clean_text(mapped_value(row, col, ["Nº documento / Tipo", "documentType", "document_type"]))
    document_no = clean_text(mapped_value(row, col, ["Nº documento / Nº", "documentNo", "document_no"]))
    return {
        "ar_reference": clean_text(mapped_value(row, col, ["Accident report / Nº", "accidentReportID", "AR"])),
        "manual_reference": clean_text(mapped_value(row, col, ["Accident report / Nº manual", "manualNo", "Nº manual"])),
        "document_type": document_type,
        "document_reference": document_no,
        "ra_reference": document_no if document_type and document_type.upper() == "RA" else None,
        "impro_reference": document_no if document_type and document_type.upper() == "IMPRO" else None,
        "accident_date": parse_date_value(mapped_value(row, col, ["Data do acidente", "accidentDate", "data acidente"])),
        "daaa_reference": clean_text(mapped_value(row, col, ["Declaração amigável", "friendlyDeclaration", "DAAA"])),
        "driver_name": clean_text(mapped_value(row, col, ["Condutor", "driver", "driverName"])),
        "driver_reference": clean_text(mapped_value(row, col, ["Condutor / Nº", "driverNo"])),
        "customer_name": clean_text(mapped_value(row, col, ["Cliente", "customer", "customerName"])),
        "plate": normalize_plate(mapped_value(row, col, ["Veículo / Matrícula", "matricula", "plate"])),
        "vehicle_reference": clean_text(mapped_value(row, col, ["Veículo / Nº", "unitNo", "unit"])),
        "vehicle_model": clean_text(mapped_value(row, col, ["Veículo / Modelo", "vehicleModel", "modelo"])),
        "source_file": source_file,
        "source_row_number": row_number,
        "raw_json": raw,
    }


def build_refstro_payload(
    row: tuple[Any, ...],
    col: dict[str, int],
    raw: dict[str, Any],
    *,
    source_file: str,
    row_number: int,
) -> dict[str, Any]:
    return {
        "refstro_reference": clean_text(mapped_value(row, col, ["REFSTRO", "refstro", "referencia sinistro", "referência sinistro"])),
        "document_reference": clean_text(mapped_value(row, col, ["ADESAO", "adesao", "documento", "document", "doc", "fatura", "invoice"])),
        "plate": normalize_plate(mapped_value(row, col, ["matricula", "matrícula", "plate", "plateNr", "licensePlate"])),
        "accident_date": parse_date_value(
            mapped_value(row, col, ["DTASTRO", "data sinistro", "data_sinistro", "accidentDate", "data acidente", "data"])
        ),
        "component": normalize_component(mapped_value(row, col, ["componente", "component", "tipo", "rubrica"])),
        "status": clean_text(mapped_value(row, col, ["status", "estado", "situação", "situacao"])),
        "close_date": parse_date_value(
            mapped_value(row, col, ["DTAENCERR.", "DTAENCERR", "data fecho", "closeDate", "closedAt", "encerramento"])
        ),
        "customer_name": clean_text(mapped_value(row, col, ["cliente", "customer", "customerName"])),
        "driver_name": clean_text(mapped_value(row, col, ["condutor", "driver", "driverName"])),
        "claim_value": parse_decimal_value(
            mapped_value(row, col, ["CUSTO TOTAL", "valor", "claimValue", "valor sinistro", "montante"])
        ),
        "cost_value": parse_decimal_value(mapped_value(row, col, ["CUSTO TOTAL", "custo", "cost", "custos", "valor custo"])),
        "source_file": source_file,
        "source_row_number": row_number,
        "raw_json": raw,
    }


def preview_status_phase(status: str | None) -> str:
    text = (status or "").strip()
    if not text:
        return "Sem fase"
    normalized = text.casefold()
    for phase in CARFAST_CLAIM_PHASES:
        if normalized.startswith(phase[:3].casefold()) or normalized == phase.casefold():
            return phase
    if "fechado" in normalized or "fech" in normalized:
        return "5.x Fechado Rentway - validar internamente"
    if "canceled" in normalized or "cancel" in normalized:
        return "Cancelado - validar se deve acompanhar"
    if "repar" in normalized:
        return "3.x Reparação"
    if "pend" in normalized:
        return "2.x Pendente"
    if "aberto" in normalized:
        return "1.x Aberto"
    return "Por mapear"


def non_zero_refstro_components(raw: dict[str, Any]) -> list[str]:
    components = []
    for source_name, component in {
        "RC": "RC",
        "DP": "DP",
        "IDS Credor": "IDS Credor",
        "VIDROS": "Vidros",
        "CUSTOS GESTÃO": "Custos de Gestão",
    }.items():
        amount = parse_decimal_value(raw.get(source_name))
        if amount not in (None, Decimal("0.00")):
            components.append(component)
    return components


def preview_suggested_action(
    *,
    import_kind: str,
    payload: dict[str, Any],
    accident_date: date | None,
    duplicate_key: bool,
    grouped_count: int,
    components: list[str] | None = None,
    known_ar: dict[str, Any] | None = None,
    conflicts: list[str] | None = None,
) -> dict[str, str]:
    if duplicate_key:
        return {
            "status": "reconciliar depois",
            "action": "Duplicado provável",
            "reason": "A mesma referência/matrícula/data aparece mais do que uma vez no ficheiro.",
        }
    if import_kind == "ar_rentway_per_vehicle":
        useful_fields = [
            payload.get("ar_reference"),
            payload.get("plate"),
            payload.get("accident_date"),
            payload.get("document_reference"),
            payload.get("driver_name"),
            payload.get("customer_name"),
        ]
        if not any(useful_fields):
            return {
                "status": "importado por validar",
                "action": "Ignorar / sem ação",
                "reason": "Linha complementar sem dados úteis para enriquecer AR.",
            }
        if conflicts:
            return {
                "status": "importado por validar",
                "action": "Conflito entre fontes",
                "reason": f"Valores diferentes face ao AR principal: {', '.join(conflicts)}.",
            }
        if known_ar:
            return {
                "status": "importado por validar",
                "action": "Enriquecer AR existente",
                "reason": "AR encontrado no ficheiro principal; CRAR acrescenta dados por viatura/documento/condutor/cliente.",
            }
        return {
            "status": "reconciliar depois",
            "action": "AR complementar sem correspondência",
            "reason": "AR existe no CRAR, mas não foi encontrado no conjunto principal; validar antes de criar processo.",
        }
    if import_kind == "ar":
        if not payload.get("ar_reference"):
            return {
                "status": "importado por validar",
                "action": "Ignorar / sem ação",
                "reason": "Linha sem referência AR; só deve avançar se for confirmada manualmente.",
            }
        if not payload.get("plate") or not (accident_date or payload.get("request_date")):
            return {
                "status": "reconciliar depois",
                "action": "AR em falta",
                "reason": "AR identificado, mas faltam matrícula ou data para propor processo com segurança.",
            }
        status = (payload.get("status") or "").casefold()
        if "cancel" in status:
            return {
                "status": "importado por validar",
                "action": "Ignorar / sem ação",
                "reason": "Status sugere cancelamento; deve ser confirmado antes de criar processo.",
            }
        if grouped_count > 1:
            return {
                "status": "importado por validar",
                "action": "Associar a processo existente",
                "reason": "Há mais linhas com a mesma matrícula/data; pode pertencer ao mesmo processo.",
            }
        return {
            "status": "importado por validar",
            "action": "Criar novo processo",
            "reason": "AR oficial com referência, matrícula e data; requer validação humana antes de criar SIN/PROC.",
        }

    if not payload.get("refstro_reference") and not payload.get("plate"):
        return {
            "status": "importado por validar",
            "action": "Ignorar / sem ação",
            "reason": "Linha sem REFSTRO nem matrícula suficiente para reconciliação.",
        }
    if not payload.get("plate") or not payload.get("accident_date"):
        return {
            "status": "reconciliar depois",
            "action": "REFSTRO por reconciliar",
            "reason": "Participação sem matrícula/data suficientes; fica para tratamento manual.",
        }
    if grouped_count > 1 or (components and len(components) > 1):
        return {
            "status": "importado por validar",
            "action": "Associar a processo existente",
            "reason": "Matrícula/data ou componentes indicam possível associação a um processo já aberto.",
        }
    return {
        "status": "reconciliar depois",
        "action": "REFSTRO por reconciliar",
        "reason": "Participação identificada sem AR confirmado no preview; requer validação/associação.",
    }


def crar_conflicts(payload: dict[str, Any], known_ar: dict[str, Any] | None) -> list[str]:
    if not known_ar:
        return []
    conflicts = []
    payload_plate = normalize_plate(payload.get("plate"))
    known_plate = normalize_plate(known_ar.get("plate"))
    if payload_plate and known_plate and payload_plate != known_plate:
        conflicts.append("matrícula")
    if (
        payload.get("vehicle_reference")
        and known_ar.get("vehicle_reference")
        and str(payload["vehicle_reference"]) != str(known_ar["vehicle_reference"])
    ):
        conflicts.append("viatura/unit")
    return conflicts


def preview_claims_file(
    path: str | Path,
    original_name: str,
    *,
    import_kind: str,
    known_ar_records: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    rows = list(iter_management_preview_rows(path, import_kind))
    headers = rows[0][1] if rows else []
    sheet_name = rows[0][0] if rows else None
    duplicate_counter: Counter[tuple[Any, ...]] = Counter()
    status_counter: Counter[str] = Counter()
    phase_counter: Counter[str] = Counter()
    component_counter: Counter[str] = Counter()
    samples = []
    warnings = []
    missing_minimum = 0
    ar_refs = set()
    refstro_refs = set()
    grouped_by_plate_date = Counter()
    normalized_rows = []
    known_ar_records = known_ar_records or {}
    crar_known_ar_matches = 0
    crar_only_rows = 0
    crar_conflict_rows = 0
    enrichment_field_counter: Counter[str] = Counter()

    for _, headers, row_number, row, raw in rows:
        col = build_column_lookup(headers)
        if import_kind == "ar":
            payload = build_ar_payload(row, col, raw, source_file=original_name, row_number=row_number)
            accident_date = parse_date_value(mapped_value(row, col, ["accidentDate", "data acidente", "data sinistro"]))
            key = (payload["ar_reference"], payload["plate"], accident_date or payload["request_date"])
            duplicate_counter[key] += 1
            if payload["ar_reference"]:
                ar_refs.add(payload["ar_reference"])
            if payload["plate"] and (accident_date or payload["request_date"]):
                grouped_by_plate_date[(payload["plate"], accident_date or payload["request_date"])] += 1
            else:
                missing_minimum += 1
            status = payload["status"] or "Sem Status"
            status_counter[status] += 1
            phase_counter[preview_status_phase(payload["status"])] += 1
            normalized_rows.append(
                {
                    "row": row_number,
                    "kind": "AR",
                    "payload": payload,
                    "key": key,
                    "group_key": (payload["plate"], accident_date or payload["request_date"]),
                    "date": accident_date or payload["request_date"],
                    "phase": preview_status_phase(payload["status"]),
                    "components": [],
                }
            )
        elif import_kind == "ar_rentway_per_vehicle":
            payload = build_crar_per_vehicle_payload(row, col, raw, source_file=original_name, row_number=row_number)
            key = (payload["ar_reference"], payload["plate"], payload["accident_date"])
            duplicate_counter[key] += 1
            if payload["ar_reference"]:
                ar_refs.add(payload["ar_reference"])
            if payload["plate"] and payload["accident_date"]:
                grouped_by_plate_date[(payload["plate"], payload["accident_date"])] += 1
            elif not payload["ar_reference"]:
                missing_minimum += 1
            known_ar = known_ar_records.get(str(payload["ar_reference"])) if payload["ar_reference"] else None
            conflicts = crar_conflicts(payload, known_ar)
            if known_ar:
                crar_known_ar_matches += 1
            else:
                crar_only_rows += 1
            if conflicts:
                crar_conflict_rows += 1
            for field, label in {
                "manual_reference": "Nº manual",
                "document_reference": "Documento RA/IMPRO",
                "accident_date": "Data do acidente",
                "daaa_reference": "Declaração amigável",
                "driver_name": "Condutor",
                "driver_reference": "Nº condutor",
                "customer_name": "Cliente",
                "vehicle_model": "Modelo viatura",
            }.items():
                if payload.get(field):
                    enrichment_field_counter[label] += 1
            phase_counter["Complementar Rentway por viatura"] += 1
            normalized_rows.append(
                {
                    "row": row_number,
                    "kind": "CRAR",
                    "payload": payload,
                    "key": key,
                    "group_key": (payload["plate"], payload["accident_date"]),
                    "date": payload["accident_date"],
                    "phase": payload["document_type"] or "Complementar",
                    "components": [],
                    "known_ar": known_ar,
                    "conflicts": conflicts,
                }
            )
        else:
            payload = build_refstro_payload(row, col, raw, source_file=original_name, row_number=row_number)
            key = (payload["refstro_reference"], payload["plate"], payload["accident_date"])
            duplicate_counter[key] += 1
            if payload["refstro_reference"]:
                refstro_refs.add(payload["refstro_reference"])
            if payload["plate"] and payload["accident_date"]:
                grouped_by_plate_date[(payload["plate"], payload["accident_date"])] += 1
            else:
                missing_minimum += 1
            components = non_zero_refstro_components(raw)
            if not components and payload["component"]:
                components = [payload["component"]]
            if not components:
                components = ["Sem componente"]
            for component in components:
                component_counter[component] += 1
            normalized_rows.append(
                {
                    "row": row_number,
                    "kind": "REFSTRO",
                    "payload": payload,
                    "key": key,
                    "group_key": (payload["plate"], payload["accident_date"]),
                    "date": payload["accident_date"],
                    "phase": ", ".join(components),
                    "components": components,
                }
            )

    duplicate_keys = [key for key, total in duplicate_counter.items() if total > 1 and any(key)]
    grouped_candidates = [key for key, total in grouped_by_plate_date.items() if total > 1]
    action_counter: Counter[str] = Counter()
    validation_status_counter: Counter[str] = Counter()
    for item in normalized_rows:
        payload = item["payload"]
        suggestion = preview_suggested_action(
            import_kind=import_kind,
            payload=payload,
            accident_date=item["date"] if import_kind in {"ar", "ar_rentway_per_vehicle"} else payload.get("accident_date"),
            duplicate_key=duplicate_counter[item["key"]] > 1 and any(item["key"]),
            grouped_count=grouped_by_plate_date[item["group_key"]] if all(item["group_key"]) else 0,
            components=item["components"],
            known_ar=item.get("known_ar"),
            conflicts=item.get("conflicts"),
        )
        action_counter[suggestion["action"]] += 1
        validation_status_counter[suggestion["status"]] += 1
        if len(samples) < 10:
            samples.append(
                {
                    "row": item["row"],
                    "kind": item["kind"],
                    "reference": payload.get("ar_reference") or payload.get("refstro_reference") or "-",
                    "plate": payload.get("plate") or "-",
                    "date": str(item["date"] or "-"),
                    "status": payload.get("status") or "-",
                    "phase": item["phase"],
                    "suggested_action": suggestion["action"],
                    "validation_status": suggestion["status"],
                    "suggested_reason": suggestion["reason"],
                }
            )
    if missing_minimum:
        warnings.append(f"{missing_minimum} linhas sem matrícula/data suficientes devem ficar como pedido de informação.")
    if duplicate_keys:
        warnings.append(f"{len(duplicate_keys)} chaves duplicadas devem ser revistas no preview.")
    if grouped_candidates:
        warnings.append(f"{len(grouped_candidates)} grupos por matrícula/data podem representar associações ou componentes.")
    if import_kind == "ar_rentway_per_vehicle" and crar_only_rows:
        warnings.append(f"{crar_only_rows} ARs aparecem no CRAR sem correspondência no ficheiro principal.")
    if import_kind == "ar_rentway_per_vehicle" and crar_conflict_rows:
        warnings.append(f"{crar_conflict_rows} ARs têm conflitos de matrícula ou viatura entre fontes.")

    return {
        "original_name": original_name,
        "import_kind": import_kind,
        "sheet_name": sheet_name,
        "headers": headers,
        "total_rows": len(rows),
        "official_ar_count": len(ar_refs),
        "participation_count": len(refstro_refs),
        "component_count": sum(component_counter.values()),
        "missing_minimum": missing_minimum,
        "duplicate_keys": len(duplicate_keys),
        "grouped_candidates": len(grouped_candidates),
        "crar_known_ar_matches": crar_known_ar_matches,
        "crar_only_rows": crar_only_rows,
        "crar_conflict_rows": crar_conflict_rows,
        "status_counts": status_counter.most_common(12),
        "phase_counts": phase_counter.most_common(12),
        "component_counts": component_counter.most_common(12),
        "enrichment_field_counts": enrichment_field_counter.most_common(12),
        "action_counts": action_counter.most_common(),
        "validation_status_counts": validation_status_counter.most_common(),
        "warnings": warnings,
        "samples": samples,
        "apply_enabled": False,
    }


def import_claims_file(
    db: Session,
    path: str | Path,
    original_name: str,
    *,
    import_kind: str,
    user_id: int | None,
) -> dict[str, Any]:
    if import_kind == "ar_rentway_per_vehicle":
        raise ValueError("CRAR per vehicle é fonte complementar e está disponível apenas em preview/staging.")
    process_type = ensure_management_defaults(db)
    import_type = AR_IMPORT_TYPE if import_kind == "ar" else REFSTRO_IMPORT_TYPE
    rows = list(iter_management_rows(path))
    batch = ImportBatch(
        source_system=MANAGEMENT_CENTER_SOURCE_SYSTEM,
        import_type=import_type,
        status="running",
        imported_by_id=user_id,
        total_rows=len(rows),
        detail=f"Importação Centro de Gestão: {MANAGEMENT_CENTER_TYPE_NAME} / dados associados.",
    )
    db.add(batch)
    db.flush()

    stored_path = management_storage_root() / f"batch_{batch.id}_{Path(path).name}"
    shutil.copyfile(path, stored_path)
    headers = rows[0][1] if rows else []
    sheet_name = rows[0][0] if rows else None
    db.add(
        ImportFile(
            batch_id=batch.id,
            original_name=original_name,
            file_name=stored_path.name,
            storage_path=str(stored_path),
            sheet_name=sheet_name,
            columns_json=headers,
        )
    )

    created_rows = 0
    error_rows = 0
    touched_process_ids: set[int] = set()
    groups = Counter()
    for sheet_name, headers, row_number, row, raw in rows:
        col = build_column_lookup(headers)
        db.add(
            ImportRawRow(
                batch_id=batch.id,
                row_number=row_number,
                external_reference=clean_text(raw.get("REFSTRO") or raw.get("AR") or raw.get("Status")),
                raw_json=raw,
                row_hash=row_hash(raw),
            )
        )
        if import_kind == "ar":
            payload = build_ar_payload(row, col, raw, source_file=original_name, row_number=row_number)
            if not payload["ar_reference"] and not payload["plate"]:
                error_rows += 1
                db.add(
                    ImportError(
                        batch_id=batch.id,
                        row_number=row_number,
                        entity_type="claim_rentway_ar",
                        error_message="Linha AR sem referência nem matrícula.",
                        raw_json=raw,
                    )
                )
                continue
            ar = ClaimRentwayAR(**payload)
            db.add(ar)
            db.flush()
            claim = get_or_create_claim(
                db,
                process_type,
                plate=payload["plate"],
                accident_date=payload["request_date"],
                customer_name=payload["customer_name"],
                driver_name=payload["driver_name"],
                document_reference=payload["ar_reference"],
                user_id=user_id,
            )
            associate_to_process(
                db,
                claim.process_id,
                entity_type="claim_rentway_ar",
                entity_id=ar.id,
                reason=f"Associado por importação AR lote #{batch.id}.",
                user_id=user_id,
            )
            refresh_claim_state(db, claim)
            touched_process_ids.add(claim.process_id)
            groups[payload["status"] or "Sem Status"] += 1
            created_rows += 1
        else:
            payload = build_refstro_payload(row, col, raw, source_file=original_name, row_number=row_number)
            if not payload["plate"] and not payload["refstro_reference"]:
                error_rows += 1
                db.add(
                    ImportError(
                        batch_id=batch.id,
                        row_number=row_number,
                        entity_type="claim_refstro_line",
                        error_message="Linha REFSTRO sem matrícula nem REFSTRO.",
                        raw_json=raw,
                    )
                )
                continue
            refstro = ClaimRefstroLine(**payload)
            db.add(refstro)
            db.flush()
            claim = get_or_create_claim(
                db,
                process_type,
                plate=payload["plate"],
                accident_date=payload["accident_date"],
                customer_name=payload["customer_name"],
                driver_name=payload["driver_name"],
                document_reference=payload["document_reference"] or payload["refstro_reference"],
                user_id=user_id,
            )
            associate_to_process(
                db,
                claim.process_id,
                entity_type="claim_refstro_line",
                entity_id=refstro.id,
                reason=f"Associado por importação REFSTRO lote #{batch.id}.",
                user_id=user_id,
            )
            refresh_claim_state(db, claim)
            touched_process_ids.add(claim.process_id)
            groups[payload["component"] or "Sem componente"] += 1
            created_rows += 1

    batch.status = "completed" if error_rows == 0 else "completed_with_errors"
    batch.created_rows = created_rows
    batch.updated_rows = len(touched_process_ids)
    batch.error_rows = error_rows
    batch.finished_at = datetime.now(UTC)
    batch.detail = f"{created_rows} linhas guardadas; {len(touched_process_ids)} SIN atualizados."
    record_audit(
        db,
        action="management_center.import.completed",
        entity_type="import_batch",
        entity_id=batch.id,
        detail=batch.detail,
        after_json={
            "import_kind": import_kind,
            "process_ids": sorted(touched_process_ids),
            "groups": groups.most_common(),
        },
        user_id=user_id,
    )
    db.commit()
    return {
        "batch_id": batch.id,
        "created_rows": created_rows,
        "updated_rows": len(touched_process_ids),
        "error_rows": error_rows,
        "groups": groups.most_common(),
    }
