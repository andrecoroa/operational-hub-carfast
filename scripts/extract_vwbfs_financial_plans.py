import json
import re
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\andre\OneDrive\Документы\New project\_v2-report-fix")
SOURCE = ROOT / "tmp" / "plans_bpi_audit" / "Planos de renda 19-05"
VWBFS = SOURCE / "VWBFS"
OUTPUT = ROOT / "tmp" / "vwbfs_financial_plans.json"
REFERENCE_DATE = date(2026, 7, 31)
VAT = Decimal("1.23")


def money(value):
    if value in (None, ""):
        return None
    if isinstance(value, str):
        value = re.sub(r"[^\d,.-]", "", value).replace(".", "").replace(",", ".")
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def with_vat(value):
    amount = money(value)
    return (amount * VAT).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def parsed_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value), "%d/%m/%Y").date()


def contract_key(value):
    return re.sub(r"\D", "", str(value or ""))


def vehicle_mapping():
    path = SOURCE / "_GERAL" / "Frota_Bancos_NumContrato.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    mapping = {}
    try:
        for row in workbook["Vehicles"].iter_rows(min_row=2, values_only=True):
            if str(row[0] or "").strip().upper() != "VW":
                continue
            contract = contract_key(row[1])
            mapping[contract] = {
                "plate": str(row[13] or "").strip().upper(),
                "unit": str(row[12] or "").strip(),
                "vin": str(row[26] or "").strip().upper(),
                "brand": str(row[20] or "").strip(),
                "model": str(row[21] or "").strip(),
                "version": str(row[22] or "").strip(),
                "source": path.name,
            }
    finally:
        workbook.close()
    return mapping


mapping = vehicle_mapping()
workbook_path = VWBFS / "Planos_Financeiros_TODOS_vmbfs.xlsx"
workbook = load_workbook(workbook_path, read_only=True, data_only=True, keep_links=False)
records = []
unmatched = []
incomplete = []
try:
    for sheet_name in workbook.sheetnames:
        contract = contract_key(sheet_name)
        vehicle = mapping.get(contract)
        if not vehicle:
            unmatched.append(sheet_name)
            continue
        sheet = workbook[sheet_name]
        financed_amount = money(sheet["B5"].value)
        spread = str(sheet["B7"].value or "").strip()
        schedule = []
        for row in sheet.iter_rows(min_row=11, values_only=True):
            if not row[1] or money(row[6]) is None:
                continue
            due_date = parsed_date(row[1])
            label = str(row[12] or row[0] or "").strip()
            installment_with_vat = money(row[11])
            schedule.append(
                {
                    "period_number": len(schedule) + 1,
                    "period_label": label,
                    "period_start": due_date.replace(day=1).isoformat(),
                    "period_end": due_date.isoformat(),
                    "amortization_amount": str(money(row[3]) or ""),
                    "interest_amount": str(money(row[4]) or ""),
                    "financial_installment": str(money(row[5]) or ""),
                    "installment_amount": str(installment_with_vat or ""),
                    "outstanding_amount": str(money(row[6]) or ""),
                    "outstanding_amount_with_vat": str(with_vat(row[6]) or ""),
                    "tax_amount": str(money(row[7]) or ""),
                    "commission_amount": str(money(row[10]) or ""),
                }
            )
        if not schedule:
            incomplete.append({"contract": sheet_name, "plate": vehicle["plate"]})
            continue
        elapsed = [item for item in schedule if date.fromisoformat(item["period_end"]) <= REFERENCE_DATE]
        selected = elapsed[-1] if elapsed else schedule[0]
        residual = next(
            (money(item["installment_amount"]) for item in schedule if item["period_label"].upper() == "VRESIDUAL"),
            None,
        )
        regular = [item for item in schedule if item["period_label"].upper() != "VRESIDUAL"]
        records.append(
            {
                **vehicle,
                "entity": "Volkswagen Bank / VWBFS",
                "contract": sheet_name,
                "start_date": schedule[0]["period_end"],
                "end_date": schedule[-1]["period_end"],
                "term_months": len(regular),
                "initial_amount": str(financed_amount or ""),
                "outstanding_amount": selected["outstanding_amount"],
                "outstanding_amount_with_vat": selected["outstanding_amount_with_vat"],
                "reference_date": selected["period_end"],
                "installment_amount": selected["financial_installment"],
                "installment_with_vat": selected["installment_amount"],
                "residual_with_vat": str(residual or ""),
                "spread": spread,
                "source": workbook_path.name,
                "installments": schedule,
            }
        )
finally:
    workbook.close()

OUTPUT.write_text(
    json.dumps(
        {"records": records, "unmatched": unmatched, "incomplete": incomplete},
        ensure_ascii=False,
        indent=2,
    ),
    encoding="utf-8",
)
print(
    f"{len(records)} planos VWBFS associados; "
    f"{sum(len(record['installments']) for record in records)} períodos; "
    f"{len(unmatched)} sem matrícula; {len(incomplete)} incompletos"
)
