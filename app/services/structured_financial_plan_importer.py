import hashlib
import json
import re
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.imports import ImportBatch, ImportFile, ImportRawRow
from app.models.vehicles import Vehicle, VehicleFinancialPlan, VehicleIdentifier
from app.services.audit import record_audit


MAIN_SHEET = "Todos os contratos"
ASSOCIATIONS_SHEET = "Viaturas associadas"
IMPORT_TYPE = "vehicle_financial_plans"


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).strip())


def _key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", _text(value).upper())


def _money(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    clean = _text(value).replace("€", "").replace(" ", "")
    if "," in clean and "." in clean:
        clean = clean.replace(".", "").replace(",", ".")
    elif "," in clean:
        clean = clean.replace(",", ".")
    try:
        return Decimal(clean).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def _date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 1:
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    text = _text(value)
    for pattern in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], pattern).date()
        except ValueError:
            pass
    return None


def _reference_date(definition: Any) -> date | None:
    text = _text(definition)
    match = re.search(r"\b(\d{2}[/-]\d{2}[/-]\d{4}|\d{4}-\d{2}-\d{2})\b", text)
    return _date_value(match.group(1)) if match else None


def _plan_status(contract_row: dict[str, Any]) -> tuple[str, bool]:
    status_text = " ".join(
        [
            _text(contract_row.get("Estado associação")),
            _text(contract_row.get("Observações")),
        ]
    ).casefold()
    closed_tokens = ("encerrado", "vendida", "vendido", "liquidado", "cancelado", "terminado")
    if any(token in status_text for token in closed_tokens):
        return "closed", False
    start = _date_value(contract_row.get("Data início"))
    end = _date_value(contract_row.get("Data fim"))
    today = date.today()
    if start and start > today:
        return "future", False
    if end and end < today:
        return "expired", False
    return "active", True


def _sheet_rows(workbook: Any, name: str) -> list[dict[str, Any]]:
    if name not in workbook.sheetnames:
        raise ValueError(f"Falta o separador obrigatório: {name}.")
    sheet = workbook[name]
    values = sheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(values)]
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        for row in values
        if any(value not in (None, "") for value in row)
    ]


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _vehicle_indexes(db: Session) -> dict[str, dict[str, Vehicle]]:
    vehicles = db.scalars(select(Vehicle)).all()
    indexes = {
        "vin": {_key(vehicle.vin): vehicle for vehicle in vehicles if _key(vehicle.vin)},
        "plate": {_key(vehicle.plate): vehicle for vehicle in vehicles if _key(vehicle.plate)},
        "unit": {_key(vehicle.rentway_unit_nr): vehicle for vehicle in vehicles if _key(vehicle.rentway_unit_nr)},
    }
    for identifier, vehicle in db.execute(
        select(VehicleIdentifier, Vehicle).join(Vehicle, Vehicle.id == VehicleIdentifier.vehicle_id)
        .where(VehicleIdentifier.active.is_(True))
    ).all():
        kind = _text(identifier.identifier_type).lower()
        if kind in indexes:
            indexes[kind].setdefault(_key(identifier.identifier_value), vehicle)
    return indexes


def _match_vehicle(row: dict[str, Any], indexes: dict[str, dict[str, Vehicle]]) -> tuple[Vehicle | None, str]:
    for field, column in (("vin", "VIN/chassis"), ("plate", "Matrícula"), ("unit", "Unidade")):
        value = _key(row.get(column))
        if value and value in indexes[field]:
            return indexes[field][value], field
    return None, ""


def preview_financial_plan_workbook(db: Session, path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        contracts = _sheet_rows(workbook, MAIN_SHEET)
        associations = _sheet_rows(workbook, ASSOCIATIONS_SHEET)
    finally:
        workbook.close()

    contract_by_key = {
        (_key(row.get("Financeira")), _key(row.get("Contrato"))): row
        for row in contracts
        if _key(row.get("Financeira")) and _key(row.get("Contrato"))
    }
    indexes = _vehicle_indexes(db)
    preview_rows: list[dict[str, Any]] = []
    matched = conflicts = unmatched = 0
    seen: set[tuple[str, str, int]] = set()

    for row_number, association in enumerate(associations, start=2):
        entity = _text(association.get("Financeira"))
        contract = _text(association.get("Contrato"))
        contract_row = contract_by_key.get((_key(entity), _key(contract)))
        vehicle, method = _match_vehicle(association, indexes)
        status = "ready"
        message = ""
        if not contract_row:
            status, message = "conflict", "Contrato não encontrado no separador principal."
            conflicts += 1
        elif not vehicle:
            status, message = "unmatched", "Viatura não encontrada por VIN, matrícula ou Unit."
            unmatched += 1
        elif (entity.upper(), contract.upper(), vehicle.id) in seen:
            status, message = "conflict", "Associação repetida no ficheiro."
            conflicts += 1
        else:
            seen.add((entity.upper(), contract.upper(), vehicle.id))
            matched += 1
        definition = _text((contract_row or {}).get("Base temporal / definição"))
        installment_amount = _money((contract_row or {}).get("Renda financeira (€)"))
        installment_with_vat = _money((contract_row or {}).get("Encargos/renda c/IVA (€)"))
        installment_display = installment_with_vat or installment_amount
        installment_source = (
            "Encargos/renda c/IVA (€)"
            if installment_with_vat is not None
            else "Renda financeira (€)"
            if installment_amount is not None
            else ""
        )
        plan_status, active = _plan_status(contract_row or {})
        preview_rows.append(
            {
                "row_number": row_number,
                "status": status,
                "message": message,
                "vehicle_id": vehicle.id if vehicle else None,
                "vehicle_plate": vehicle.plate if vehicle else _text(association.get("Matrícula")),
                "vehicle_vin": vehicle.vin if vehicle else _text(association.get("VIN/chassis")),
                "vehicle_unit": vehicle.rentway_unit_nr if vehicle else _text(association.get("Unidade")),
                "match_method": method,
                "finance_entity": entity,
                "contract_number": contract,
                "association_status": _text(association.get("Estado viatura") or (contract_row or {}).get("Estado associação")),
                "association_confidence": _text(association.get("Confiança") or (contract_row or {}).get("Confiança")),
                "association_evidence": _text(association.get("Evidência")),
                "plan_status": plan_status,
                "active": active,
                "start_date": (_date_value((contract_row or {}).get("Data início")) or "").isoformat()
                if _date_value((contract_row or {}).get("Data início")) else "",
                "end_date": (_date_value((contract_row or {}).get("Data fim")) or "").isoformat()
                if _date_value((contract_row or {}).get("Data fim")) else "",
                "term_months": int((contract_row or {}).get("Prazo (meses)"))
                if isinstance((contract_row or {}).get("Prazo (meses)"), (int, float)) else None,
                "initial_amount": str(_money((contract_row or {}).get("Capital inicial (€)")) or ""),
                "outstanding_amount": str(_money((contract_row or {}).get("Saldo conhecido (€)")) or ""),
                "amount_reference_date": (_reference_date(definition) or "").isoformat()
                if _reference_date(definition) else "",
                "installment_amount": str(installment_amount or ""),
                "installment_with_vat": str(installment_display or ""),
                "installment_source": installment_source,
                "residual_amount": str(_money((contract_row or {}).get("Valor residual (€)")) or ""),
                "source_definition": definition,
                "source_references": _text((contract_row or {}).get("Fontes consolidadas")),
                "raw": {
                    "contract": {key: _text(value) for key, value in (contract_row or {}).items()},
                    "association": {key: _text(value) for key, value in association.items()},
                },
            }
        )

    return {
        "file_hash": file_sha256(path),
        "total_contracts": len(contracts),
        "total_associations": len(associations),
        "matched": matched,
        "conflicts": conflicts,
        "unmatched": unmatched,
        "rows": preview_rows,
    }


def _decimal_or_none(value: Any) -> Decimal | None:
    return Decimal(str(value)) if value not in (None, "") else None


def _iso_date(value: Any) -> date | None:
    return date.fromisoformat(str(value)) if value else None


def apply_financial_plan_preview(
    db: Session,
    preview: dict[str, Any],
    *,
    source_path: Path,
    original_name: str,
    user_id: int | None,
) -> dict[str, int]:
    ready_rows = [row for row in preview.get("rows", []) if row.get("status") == "ready"]
    batch = ImportBatch(
        source_system="structured_workbook",
        import_type=IMPORT_TYPE,
        status="completed",
        imported_by_id=user_id,
        finished_at=datetime.now(UTC),
        total_rows=int(preview.get("total_associations") or 0),
        created_rows=0,
        updated_rows=0,
        skipped_rows=int(preview.get("conflicts") or 0) + int(preview.get("unmatched") or 0),
        error_rows=int(preview.get("conflicts") or 0),
        detail=json.dumps({"sha256": preview.get("file_hash"), "confirmed_preview": True}, ensure_ascii=False),
    )
    db.add(batch)
    db.flush()
    db.add(
        ImportFile(
            batch_id=batch.id,
            original_name=original_name[:255],
            file_name=source_path.name[:255],
            storage_path=str(source_path),
            sheet_name=f"{MAIN_SHEET}; {ASSOCIATIONS_SHEET}",
            columns_json=[],
        )
    )
    created = updated = unchanged = 0
    for row in ready_rows:
        key = (
            _text(row["finance_entity"]),
            _text(row["contract_number"]),
            int(row["vehicle_id"]),
        )
        plan = db.scalar(
            select(VehicleFinancialPlan).where(
                VehicleFinancialPlan.finance_entity == key[0],
                VehicleFinancialPlan.contract_number == key[1],
                VehicleFinancialPlan.vehicle_id == key[2],
            )
        )
        values = {
            "import_batch_id": batch.id,
            "association_status": row.get("association_status") or None,
            "association_confidence": row.get("association_confidence") or None,
            "association_method": row.get("match_method") or None,
            "plan_status": row.get("plan_status") or "active",
            "start_date": _iso_date(row.get("start_date")),
            "end_date": _iso_date(row.get("end_date")),
            "term_months": row.get("term_months"),
            "initial_amount": _decimal_or_none(row.get("initial_amount")),
            "outstanding_amount": _decimal_or_none(row.get("outstanding_amount")),
            "amount_reference_date": _iso_date(row.get("amount_reference_date")),
            "installment_amount": _decimal_or_none(row.get("installment_amount")),
            "installment_with_vat": _decimal_or_none(row.get("installment_with_vat")),
            "residual_amount": _decimal_or_none(row.get("residual_amount")),
            "source_definition": row.get("source_definition") or None,
            "source_references": row.get("source_references") or None,
            "raw_json": row.get("raw") or {},
            "active": bool(row.get("active")),
            "human_confirmed": True,
            "confirmed_by_id": user_id,
            "confirmed_at": datetime.now(UTC),
        }
        if plan is None:
            plan = VehicleFinancialPlan(
                vehicle_id=key[2],
                finance_entity=key[0],
                contract_number=key[1],
                **values,
            )
            db.add(plan)
            created += 1
        else:
            changed = any(getattr(plan, field) != value for field, value in values.items() if field != "import_batch_id")
            for field, value in values.items():
                setattr(plan, field, value)
            updated += int(changed)
            unchanged += int(not changed)
        db.add(
            ImportRawRow(
                batch_id=batch.id,
                row_number=int(row["row_number"]),
                external_reference=f"{key[0]}:{key[1]}:{key[2]}",
                raw_json=row,
                row_hash=hashlib.sha256(json.dumps(row, sort_keys=True).encode()).hexdigest(),
            )
        )
    batch.created_rows = created
    batch.updated_rows = updated
    batch.skipped_rows += unchanged
    record_audit(
        db,
        action="financial_plans.import.confirmed",
        entity_type="import_batch",
        entity_id=batch.id,
        after_json={"created": created, "updated": updated, "unchanged": unchanged},
        user_id=user_id,
    )
    return {"batch_id": batch.id, "created": created, "updated": updated, "unchanged": unchanged}
