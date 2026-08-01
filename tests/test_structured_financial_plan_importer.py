from io import BytesIO
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from fastapi.testclient import TestClient
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.imports import ImportBatch
from app.models.vehicles import Vehicle, VehicleFinancialPlan
from app.services.structured_financial_plan_importer import (
    apply_financial_plan_preview,
    preview_financial_plan_workbook,
)


def _workbook(path: Path) -> None:
    wb = Workbook()
    contracts = wb.active
    contracts.title = "Todos os contratos"
    contracts.append(
        [
            "Financeira",
            "Contrato",
            "Estado associação",
            "N.º viaturas",
            "Matrículas",
            "VIN/chassis",
            "Confiança",
            "Data início",
            "Data fim",
            "Prazo (meses)",
            "Capital inicial (€)",
            "Saldo conhecido (€)",
            "Renda financeira (€)",
            "Taxa juro",
            "Spread",
            "Valor residual (€)",
            "Encargos/renda c/IVA (€)",
            "Qualidade financeira",
            "Base temporal / definição",
            "Fontes consolidadas",
            "Observações",
        ]
    )
    contracts.append(
        [
            "CGD",
            "400144083",
            "Associado",
            2,
            "AA-00-AA; BB-00-BB",
            "",
            "Exata",
            "01/01/2024",
            "31/12/2031",
            96,
            50000,
            32000,
            510,
            "",
            "",
            5000,
            627.30,
            "",
            "Saldo conhecido em 30/06/2026",
            "Fonte teste",
            "",
        ]
    )
    associations = wb.create_sheet("Viaturas associadas")
    associations.append(
        [
            "Financeira",
            "Contrato",
            "Matrícula",
            "VIN/chassis",
            "Unidade",
            "Marca",
            "Modelo",
            "Versão",
            "Estado viatura",
            "Confiança",
            "Evidência",
            "Fontes",
            "Observações",
        ]
    )
    associations.append(["CGD", "400144083", "AA-00-AA", "VIN00000000000001", "101", "", "", "", "Ativa", "Exata", "VIN", "", ""])
    associations.append(["CGD", "400144083", "BB-00-BB", "", "102", "", "", "", "Ativa", "Forte", "Unit", "", ""])
    associations.append(["CGD", "400144083", "ZZ-99-ZZ", "", "", "", "", "", "Ativa", "Provável", "", "", ""])
    wb.save(path)


def _workbook_bytes(tmp_path: Path) -> bytes:
    path = tmp_path / "plans-upload.xlsx"
    _workbook(path)
    return path.read_bytes()


def test_preview_matches_by_vin_then_unit_without_writing(db_session: Session, tmp_path: Path):
    path = tmp_path / "plans.xlsx"
    _workbook(path)
    first = Vehicle(plate="AA-00-AA", vin="VIN00000000000001", rentway_unit_nr="101")
    second = Vehicle(plate="BB-00-BB", vin="VIN00000000000002", rentway_unit_nr="102")
    db_session.add_all([first, second])
    db_session.commit()

    preview = preview_financial_plan_workbook(db_session, path)

    assert preview["total_contracts"] == 1
    assert preview["total_associations"] == 3
    assert preview["matched"] == 2
    assert preview["unmatched"] == 1
    assert preview["rows"][0]["match_method"] == "vin"
    assert preview["rows"][1]["match_method"] == "plate"
    assert db_session.scalar(select(func.count()).select_from(VehicleFinancialPlan)) == 0


def test_confirm_preserves_multivehicle_contract_and_96_month_term(db_session: Session, tmp_path: Path):
    path = tmp_path / "plans.xlsx"
    _workbook(path)
    db_session.add_all(
        [
            Vehicle(plate="AA-00-AA", vin="VIN00000000000001", rentway_unit_nr="101"),
            Vehicle(plate="BB-00-BB", vin="VIN00000000000002", rentway_unit_nr="102"),
        ]
    )
    db_session.commit()
    preview = preview_financial_plan_workbook(db_session, path)

    result = apply_financial_plan_preview(
        db_session,
        preview,
        source_path=path,
        original_name=path.name,
        user_id=None,
    )
    db_session.commit()

    plans = db_session.scalars(select(VehicleFinancialPlan).order_by(VehicleFinancialPlan.vehicle_id)).all()
    assert result["created"] == 2
    assert len(plans) == 2
    assert {plan.contract_number for plan in plans} == {"400144083"}
    assert {plan.term_months for plan in plans} == {96}
    assert {plan.installment_with_vat for plan in plans} == {Decimal("627.30")}
    assert db_session.scalar(select(func.count()).select_from(ImportBatch)) == 1


def test_confirmed_import_deactivates_previous_plan_from_same_finance_entity(
    db_session: Session,
    tmp_path: Path,
):
    path = tmp_path / "plans.xlsx"
    _workbook(path)
    vehicle = Vehicle(plate="AA-00-AA", vin="VIN00000000000001", rentway_unit_nr="101")
    db_session.add(vehicle)
    db_session.flush()
    previous = VehicleFinancialPlan(
        vehicle_id=vehicle.id,
        finance_entity="CGD",
        contract_number="OLD-CONTRACT",
        active=True,
    )
    db_session.add(previous)
    db_session.commit()

    preview = preview_financial_plan_workbook(db_session, path)
    apply_financial_plan_preview(
        db_session,
        preview,
        source_path=path,
        original_name=path.name,
        user_id=None,
    )
    db_session.commit()

    db_session.refresh(previous)
    imported = db_session.scalar(
        select(VehicleFinancialPlan).where(
            VehicleFinancialPlan.vehicle_id == vehicle.id,
            VehicleFinancialPlan.contract_number == "400144083",
        )
    )
    assert previous.active is False
    assert imported is not None
    assert imported.active is True


def test_preview_uses_financial_rent_when_vat_rent_is_missing(db_session: Session, tmp_path: Path):
    path = tmp_path / "plans-without-vat-rent.xlsx"
    _workbook(path)
    workbook = load_workbook(path)
    contracts = workbook["Todos os contratos"]
    contracts["Q2"] = None
    workbook.save(path)
    db_session.add(Vehicle(plate="AA-00-AA", vin="VIN00000000000001", rentway_unit_nr="101"))
    db_session.commit()

    preview = preview_financial_plan_workbook(db_session, path)

    assert preview["rows"][0]["installment_amount"] == "510.00"
    assert preview["rows"][0]["installment_with_vat"] == "510.00"
    assert preview["rows"][0]["installment_source"] == "Renda financeira (€)"


def test_preview_prefers_vehicle_financial_values_from_association_sheet(
    db_session: Session,
    tmp_path: Path,
):
    path = tmp_path / "plans-per-vehicle-values.xlsx"
    _workbook(path)
    workbook = load_workbook(path)
    associations = workbook["Viaturas associadas"]
    extra_headers = [
        "Saldo conhecido (€)",
        "Encargos/renda c/IVA (€)",
        "Valor residual (€)",
        "Base temporal / definição",
    ]
    for column, header in enumerate(extra_headers, start=14):
        associations.cell(1, column, header)
    associations.cell(2, 14, 1200.50)
    associations.cell(2, 15, 320.10)
    associations.cell(2, 16, 800.25)
    associations.cell(2, 17, "Saldo conhecido em 15/07/2026")
    workbook.save(path)
    db_session.add(Vehicle(plate="AA-00-AA", vin="VIN00000000000001", rentway_unit_nr="101"))
    db_session.commit()

    preview = preview_financial_plan_workbook(db_session, path)

    row = preview["rows"][0]
    assert row["outstanding_amount"] == "1200.50"
    assert row["installment_with_vat"] == "320.10"
    assert row["residual_amount"] == "800.25"
    assert row["amount_reference_date"] == "2026-07-15"


def test_clean_financial_plan_import_requires_preview_before_confirm(
    authenticated_client: TestClient,
    db_session: Session,
    tmp_path: Path,
):
    db_session.add(
        Vehicle(
            plate="AA-00-AA",
            vin="VIN00000000000001",
            rentway_unit_nr="101",
        )
    )
    db_session.commit()

    response = authenticated_client.post(
        "/v2-clean/documentation/financial-plans/preview",
        files={
            "file": (
                "plans.xlsx",
                BytesIO(_workbook_bytes(tmp_path)),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert "/v2-clean/documentation/financial-plans/preview/" in response.headers["location"]
    assert db_session.scalar(select(func.count()).select_from(VehicleFinancialPlan)) == 0
    preview_page = authenticated_client.get(response.headers["location"])
    assert preview_page.status_code == 200
    assert "Confirmar 1 associações" in preview_page.text
