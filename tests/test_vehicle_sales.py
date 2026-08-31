import io
from datetime import date
from decimal import Decimal
from types import SimpleNamespace

from sqlalchemy import select
from openpyxl import load_workbook
from PyPDF2 import PdfReader

import app.web.router as base_router
from app.core.config import settings
from app.models import (
    AuditLog,
    Document,
    Vehicle,
    VehicleExternalSnapshot,
    VehicleFinancialPlan,
    VehicleFinancialPlanInstallment,
    VehicleImage,
    VehicleManualField,
    VehicleSaleLead,
    VehicleSaleProfile,
    VehicleSaleProposal,
    VehicleSaleProposalLine,
    VehicleSalePublication,
)
from app.services.users import create_user
from app.services.vehicle_financials import canonical_vehicle_financial_values
from app.web.vehicle_sales import (
    _filter_rows,
    _financial_audit_rows,
    _media_root,
    _sale_row,
    compact_finance_entity,
)


def test_compact_finance_entity_labels():
    assert compact_finance_entity("Caixa Geral de Depósitos, S.A.") == "CGD"
    assert compact_finance_entity("Santander Consumer Finance") == "Santander"
    assert compact_finance_entity("Banco BPI, S.A.") == "BPI"
    assert compact_finance_entity("CGD Locação Corrente") == "CGD Locação"
    assert compact_finance_entity("LeasePlan Portugal") == "LeasePlan"
    assert compact_finance_entity("Mercedes-Benz Financial") == "Mercedes"


def test_amortization_is_effective_on_first_day_of_current_month():
    reference = date(2026, 8, 5)
    for purchase_date, installment_end in (
        (date(2023, 5, 30), date(2026, 6, 30)),
        (date(2025, 7, 15), date(2026, 7, 15)),
        (date(2026, 8, 3), date(2026, 7, 31)),
    ):
        values = canonical_vehicle_financial_values(
            cost_context={
                "purchase_date": purchase_date,
                "initial_cost_with_vat": Decimal("24000"),
            },
            plan=SimpleNamespace(
                start_date=purchase_date,
                initial_amount=Decimal("24000"),
                outstanding_amount=Decimal("18000"),
                amount_reference_date=installment_end,
            ),
            installments=[
                SimpleNamespace(
                    period_end=installment_end,
                    period_number=1,
                    amortization_amount=Decimal("250"),
                )
            ],
            current_value_calculator=lambda *_args: Decimal("10000"),
            reference=reference,
        )

        assert values["current_value_date"] == date(2026, 8, 1)


def test_vehicle_sale_media_defaults_to_persistent_document_archive(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "vehicle_sale_media_root", None)
    monkeypatch.setattr(base_router, "document_archive_root", lambda: tmp_path)

    assert _media_root() == (tmp_path / "Venda de viaturas" / "imagens").resolve()


def test_cgd_filter_includes_all_cgd_name_variants():
    rows = []
    for entity in ("CGD", "Caixa Geral de Depósitos, S.A."):
        rows.append(
            {
                "vehicle": SimpleNamespace(
                    plate=entity,
                    rentway_unit_nr=None,
                    vin=None,
                    brand=None,
                    model=None,
                ),
                "finance_entity": entity,
                "finance_entity_key": compact_finance_entity(entity).casefold(),
                "status": "candidate",
                "vehicle_state": "free",
                "registration": None,
                "return_on": None,
                "financial_margin": None,
                "commercial_margin": None,
                "market_trade": None,
                "market_retail": None,
            }
        )
    filters = {
        "q": "",
        "sale_status": "",
        "finance_entity": "CGD",
        "vehicle_state": "",
        "rentway_group": "",
        "registration_from": "",
        "registration_to": "",
        "return_from": "",
        "return_to": "",
        "financial_margin_min": "",
        "financial_margin_max": "",
        "commercial_margin_min": "",
        "commercial_margin_max": "",
        "market_state": "",
    }

    assert len(_filter_rows(rows, filters)) == 2


def test_sale_proposal_keeps_vehicle_values_independent(authenticated_client, db_session):
    vehicle = create_sale_vehicle(db_session)
    profile = VehicleSaleProfile(
        vehicle_id=vehicle.id,
        status="candidate",
        market_trade_value=Decimal("21000.00"),
    )
    db_session.add(profile)
    db_session.add(
        VehicleFinancialPlan(
            vehicle_id=vehicle.id,
            finance_entity="Santander",
            contract_number="PROP-DEBT-1",
            outstanding_amount=Decimal("17000.00"),
            amount_reference_date=date(2026, 8, 1),
            active=True,
        )
    )
    db_session.commit()

    created = authenticated_client.post(
        "/v2-clean/fleet/sales/bulk",
        data={"vehicle_ids": [str(vehicle.id)], "action": "proposal"},
        follow_redirects=False,
    )
    assert created.status_code == 303
    proposal = db_session.scalar(select(VehicleSaleProposal))
    assert proposal is not None
    line = db_session.scalar(
        select(VehicleSaleProposalLine).where(
            VehicleSaleProposalLine.proposal_id == proposal.id
        )
    )
    assert line is not None
    assert line.base_price == Decimal("21000.00")
    assert line.snapshot_json["registration"] == "15/01/2022"
    assert line.snapshot_json["colour"] == "Azul"
    assert line.snapshot_json["fuel"] == "Diesel"
    assert line.snapshot_json["gearbox"] == "Automática"
    assert line.snapshot_json["debt"] == "20910.00"
    assert line.snapshot_json["finance_entity"] == "Santander"
    assert line.snapshot_json["contract_number"] == "PROP-DEBT-1"

    saved = authenticated_client.post(
        f"/v2-clean/fleet/sales/proposals/{proposal.id}",
        data={
            "title": "Lote agosto",
            "recipient": "Comerciante teste",
            "included_line_ids": [str(line.id)],
            f"price_{line.id}": "19800",
            f"counteroffer_{line.id}": "19400",
            f"notes_{line.id}": "Preço exclusivo do lote",
        },
        follow_redirects=False,
    )
    assert saved.status_code == 303
    db_session.expire_all()
    assert db_session.get(VehicleSaleProposalLine, line.id).proposed_price == Decimal("19800.00")
    assert db_session.get(VehicleSaleProposalLine, line.id).customer_counteroffer == Decimal("19400.00")
    assert db_session.get(VehicleSaleProfile, profile.id).market_trade_value == Decimal("21000.00")

    sent = authenticated_client.post(
        f"/v2-clean/fleet/sales/proposals/{proposal.id}/send",
        follow_redirects=False,
    )
    assert sent.status_code == 303
    plan = db_session.scalar(
        select(VehicleFinancialPlan).where(
            VehicleFinancialPlan.vehicle_id == vehicle.id,
            VehicleFinancialPlan.active.is_(True),
            VehicleFinancialPlan.contract_number == "PROP-DEBT-1",
        )
    )
    plan.outstanding_amount = Decimal("16000.00")
    db_session.commit()
    reopened = authenticated_client.post(
        f"/v2-clean/fleet/sales/proposals/{proposal.id}/reopen",
        data={"refresh_financials": "1"},
        follow_redirects=False,
    )
    assert reopened.status_code == 303
    db_session.expire_all()
    proposals = db_session.scalars(
        select(VehicleSaleProposal).order_by(VehicleSaleProposal.version)
    ).all()
    assert [item.version for item in proposals] == [1, 2]
    assert proposals[0].status == "sent"
    assert proposals[1].status == "draft"
    first_version_line = db_session.scalar(
        select(VehicleSaleProposalLine).where(
            VehicleSaleProposalLine.proposal_id == proposals[0].id
        )
    )
    second_version_line = db_session.scalar(
        select(VehicleSaleProposalLine).where(
            VehicleSaleProposalLine.proposal_id == proposals[1].id
        )
    )
    assert first_version_line.snapshot_json["debt"] == "20910.00"
    assert second_version_line.snapshot_json["debt"] == "19680.00"
    assert second_version_line.proposed_price == Decimal("19800.00")
    assert second_version_line.customer_counteroffer == Decimal("19400.00")

    listing = authenticated_client.get("/v2-clean/fleet/sales/proposals")
    assert listing.status_code == 200
    assert proposal.reference in listing.text
    detail = authenticated_client.get(
        f"/v2-clean/fleet/sales/proposals/{proposals[1].id}"
    )
    assert detail.status_code == 200
    assert "Preço exclusivo do lote" in detail.text
    export = authenticated_client.get(
        f"/v2-clean/fleet/sales/proposals/{proposals[1].id}/xlsx"
    )
    assert export.status_code == 200
    assert export.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    sheet = load_workbook(io.BytesIO(export.content), data_only=True).active
    headers = [cell.value for cell in sheet[3]]
    assert "Data matrícula" in headers
    assert "Cor" in headers
    assert "Combustível" in headers
    assert "Caixa" in headers
    assert "Valor em dívida" in headers
    assert "Entidade financeira" in headers
    assert "N.º contrato" in headers
    assert "Margem CarFast" in headers
    assert "Contraproposta cliente" in headers
    assert "Margem contraproposta" in headers
    assert "Custo" in headers
    assert "Situação" in headers
    assert "Cliente" in headers
    assert "Data de devolução" in headers
    assert "Ano" not in headers
    assert sheet.cell(row=4, column=headers.index("Data matrícula") + 1).value == "15/01/2022"
    assert sheet.cell(row=4, column=headers.index("Cor") + 1).value == "Azul"
    assert sheet.cell(row=4, column=headers.index("Combustível") + 1).value == "Diesel"
    assert sheet.cell(row=4, column=headers.index("Caixa") + 1).value == "Automática"
    assert sheet.cell(row=4, column=headers.index("Valor em dívida") + 1).value == 19680
    assert sheet.cell(row=4, column=headers.index("Entidade financeira") + 1).value == "Santander"
    contract_cell = sheet.cell(row=4, column=headers.index("N.º contrato") + 1)
    assert contract_cell.value == "PROP-DEBT-1"
    assert contract_cell.number_format == "@"
    assert sheet.cell(row=4, column=headers.index("Margem CarFast") + 1).value == 120
    assert sheet.cell(row=4, column=headers.index("Contraproposta cliente") + 1).value == 19400
    assert sheet.cell(row=4, column=headers.index("Margem contraproposta") + 1).value == -280
    assert sheet.cell(row=4, column=headers.index("Custo") + 1).value is not None
    customer_export = authenticated_client.get(
        f"/v2-clean/fleet/sales/proposals/{proposals[1].id}/customer.xlsx"
    )
    assert customer_export.status_code == 200
    customer_sheet = load_workbook(
        io.BytesIO(customer_export.content), data_only=True
    ).active
    customer_headers = [cell.value for cell in customer_sheet[3]]
    assert "Proposto CarFast" in customer_headers
    assert "Contraproposta cliente" not in customer_headers
    assert "Valor em dívida" not in customer_headers
    assert "Entidade financeira" not in customer_headers
    assert "N.º contrato" not in customer_headers
    assert "Margem CarFast" not in customer_headers
    assert "Valor em dívida" not in customer_headers
    assert "Margem negocial" not in customer_headers
    pdf = authenticated_client.get(
        f"/v2-clean/fleet/sales/proposals/{proposals[1].id}/pdf"
    )
    assert pdf.status_code == 200
    assert pdf.headers["content-type"] == "application/pdf"
    assert pdf.content.startswith(b"%PDF")
    pdf_text = "\n".join(
        page.extract_text() or "" for page in PdfReader(io.BytesIO(pdf.content)).pages
    )
    assert "Data matrícula" in pdf_text
    assert "15/01/2022" in pdf_text
    assert "Azul" in pdf_text
    assert "Em dívida" not in pdf_text
    assert "Margem" not in pdf_text

    completed = authenticated_client.post(
        f"/v2-clean/fleet/sales/proposals/{proposals[1].id}/status",
        data={"status": "completed"},
        follow_redirects=False,
    )
    assert completed.status_code == 303
    db_session.expire_all()
    assert db_session.get(VehicleSaleProposal, proposals[1].id).status == "completed"

    cancelled = authenticated_client.post(
        f"/v2-clean/fleet/sales/proposals/{proposal.id}/status",
        data={"status": "cancelled"},
        follow_redirects=False,
    )
    assert cancelled.status_code == 303
    db_session.expire_all()
    assert db_session.get(VehicleSaleProposal, proposal.id).status == "cancelled"


def test_proposal_vehicle_status_updates_only_included_and_blocks_other_lots(
    authenticated_client, db_session
):
    included_vehicle = create_sale_vehicle(db_session)
    excluded_vehicle = Vehicle(
        plate="98-ZZ-76",
        vin="VF3TESTVEHICLE0002",
        brand="Peugeot",
        model="208",
        lifecycle_status="active",
        operational_status="free",
        active=True,
    )
    db_session.add(excluded_vehicle)
    db_session.flush()
    proposal = VehicleSaleProposal(
        reference="PC-TEST-LOT-1-V2",
        version=2,
        status="draft",
        title="Lote protegido",
    )
    db_session.add(proposal)
    db_session.flush()
    db_session.add_all(
        [
            VehicleSaleProposalLine(
                proposal_id=proposal.id,
                vehicle_id=included_vehicle.id,
                snapshot_json={"plate": included_vehicle.plate},
                included=True,
                sort_order=0,
            ),
            VehicleSaleProposalLine(
                proposal_id=proposal.id,
                vehicle_id=excluded_vehicle.id,
                snapshot_json={"plate": excluded_vehicle.plate},
                included=False,
                sort_order=1,
            ),
        ]
    )
    db_session.commit()

    reserved = authenticated_client.post(
        f"/v2-clean/fleet/sales/proposals/{proposal.id}/vehicle-status",
        data={"status": "reserved"},
        follow_redirects=False,
    )
    assert reserved.status_code == 303
    assert "vehicle_status_saved=reserved" in reserved.headers["location"]
    db_session.expire_all()
    included_profile = db_session.scalar(
        select(VehicleSaleProfile).where(
            VehicleSaleProfile.vehicle_id == included_vehicle.id
        )
    )
    excluded_profile = db_session.scalar(
        select(VehicleSaleProfile).where(
            VehicleSaleProfile.vehicle_id == excluded_vehicle.id
        )
    )
    assert included_profile.status == "reserved"
    assert included_profile.status_reference == "PC-TEST-LOT-1"
    assert excluded_profile is None
    listing = authenticated_client.get(
        "/v2-clean/fleet/sales",
        params={"search": "1", "q": included_vehicle.plate},
    )
    assert listing.status_code == 200
    assert "Reservada · lote PC-TEST-LOT-1" in listing.text

    other_proposal = VehicleSaleProposal(
        reference="PC-TEST-LOT-2",
        version=1,
        status="draft",
        title="Outro lote",
    )
    db_session.add(other_proposal)
    db_session.flush()
    db_session.add(
        VehicleSaleProposalLine(
            proposal_id=other_proposal.id,
            vehicle_id=included_vehicle.id,
            snapshot_json={"plate": included_vehicle.plate},
            included=True,
            sort_order=0,
        )
    )
    db_session.commit()

    conflict = authenticated_client.post(
        f"/v2-clean/fleet/sales/proposals/{other_proposal.id}/vehicle-status",
        data={"status": "sold"},
        follow_redirects=False,
    )
    assert conflict.status_code == 303
    assert "vehicle_status_error=lot_conflict" in conflict.headers["location"]
    db_session.expire_all()
    included_profile = db_session.get(VehicleSaleProfile, included_profile.id)
    assert included_profile.status == "reserved"
    assert included_profile.status_reference == "PC-TEST-LOT-1"

    manual = authenticated_client.post(
        f"/v2-clean/fleet/sales/{included_vehicle.id}",
        data={"status": "for_sale"},
        follow_redirects=False,
    )
    assert manual.status_code == 303
    db_session.expire_all()
    included_profile = db_session.get(VehicleSaleProfile, included_profile.id)
    assert included_profile.status == "for_sale"
    assert included_profile.status_reference is None

    audit = db_session.scalar(
        select(AuditLog).where(
            AuditLog.action == "vehicle.sale.proposal_vehicles_reserved"
        )
    )
    assert audit is not None


def test_unfinanced_vehicle_ignores_legacy_manual_debt():
    vehicle = Vehicle(
        plate="BS-13-UU",
        brand="Mercedes",
        model="Classe A",
        operational_status="active",
        lifecycle_status="active",
        active=True,
    )
    snapshot = VehicleExternalSnapshot(
        data_json={
            "purchase_date": "2025-04-30",
            "acquisition_value": "30.902,40",
            "value_with_tax": "38.010,00",
        }
    )

    row = _sale_row(
        vehicle,
        snapshot,
        {"debt_value": "36.982,78"},
        None,
        None,
    )

    assert row["debt"] is None
    assert row["financial_margin"] is None


def test_sale_row_uses_same_amortized_current_cost_as_vehicle_sheet(monkeypatch):
    vehicle = Vehicle(plate="12-AA-34", active=True)
    snapshot = VehicleExternalSnapshot(data_json={})
    monkeypatch.setattr(
        base_router,
        "current_cost_from_snapshot",
        lambda _snapshot: {
            "initial_cost_with_vat": Decimal("9600.00"),
            "current_cost_with_vat": Decimal("9000.00"),
            "amortization_month": 24,
        },
    )

    row = _sale_row(vehicle, snapshot, {}, None, None)

    assert row["cost"] == Decimal("7200.00")


def create_sale_vehicle(db_session) -> Vehicle:
    vehicle = Vehicle(
        plate="12-AB-34",
        vin="VF3TESTVEHICLE0001",
        rentway_unit_nr="UNIT123",
        brand="Peugeot",
        model="3008",
        version="1.5 BlueHDi",
        year=2022,
        lifecycle_status="active",
        operational_status="in_contract",
        active=True,
    )
    db_session.add(vehicle)
    db_session.flush()
    db_session.add(
        VehicleExternalSnapshot(
            vehicle_id=vehicle.id,
            source_system="rentway",
            data_json={
                "plate_date": "2022-01-15",
                "colour": "Azul",
                "fuel": "Diesel",
                "transmission_type": "EAT8",
                "purchase_date": "2026-07-01",
                "acquisition_value": "15609.76",
                "value_with_tax": "19200",
                "km": "98450",
                "current_status": "Contrato",
                "groupid": "C1",
                "document_nr": "CONT-100",
                "return_date": "2026-10-18",
                "finance_entity": "Santander",
            },
        )
    )
    db_session.add_all(
        [
            VehicleManualField(
                vehicle_id=vehicle.id,
                field_code="finance_entity",
                value_json="Santander",
            ),
            VehicleManualField(
                vehicle_id=vehicle.id,
                field_code="debt_value",
                value_json="15000",
            ),
        ]
    )
    db_session.add(
        VehicleFinancialPlan(
            vehicle_id=vehicle.id,
            finance_entity="Santander",
            contract_number="CONT-100",
            outstanding_amount=Decimal("15000"),
            active=True,
        )
    )
    db_session.commit()
    return vehicle


def test_vehicle_sales_filters_bulk_values_and_price_rule(authenticated_client, db_session):
    vehicle = create_sale_vehicle(db_session)

    page = authenticated_client.get(
        "/v2-clean/fleet/sales",
        params={
            "finance_entity": "Santander",
            "vehicle_state": "contract",
            "rentway_group": "C1",
            "registration_from": "2022-01-01",
            "registration_to": "2022-12-31",
            "return_from": "2026-10-01",
            "return_to": "2026-10-31",
            "financial_margin_min": "-100000",
        },
    )
    assert page.status_code == 200
    assert "Venda de viaturas" in page.text
    assert "12-AB-34" in page.text
    assert "Dev. 18/10/2026" in page.text
    assert "Custo CarFast − valor em dívida" in page.text
    assert "Valor comércio − custo CarFast" in page.text
    assert "18 450,00 €" in page.text
    assert 'name="rentway_group"' in page.text
    assert 'value="C1" checked' in page.text
    assert "Valor de custo" in page.text
    assert 'name="price_target"' in page.text
    assert 'value="trade">Valor comércio' in page.text
    assert 'value="retail">Valor cliente final' in page.text

    other_group = authenticated_client.get(
        "/v2-clean/fleet/sales",
        params={"rentway_group": "C2"},
    )
    assert other_group.status_code == 200
    assert "12-AB-34" not in other_group.text

    detail = authenticated_client.get(f"/v2-clean/fleet/sales/{vehicle.id}")
    assert detail.status_code == 200
    assert "18 450,00 €" in detail.text

    values = authenticated_client.post(
        "/v2-clean/fleet/sales/bulk",
        data={
            "vehicle_ids": [str(vehicle.id)],
            "action": "market_values",
            "market_trade_value": "21000",
            "market_retail_value": "25200",
            "market_value_source": "Avaliação interna",
            "return_url": "/v2-clean/fleet/sales",
        },
        follow_redirects=False,
    )
    assert values.status_code == 303

    rule = authenticated_client.post(
        "/v2-clean/fleet/sales/bulk",
        data={
            "vehicle_ids": [str(vehicle.id)],
            "action": "price_rule",
            "price_base": "trade",
            "margin_mode": "percentage",
            "margin_value": "5",
            "rounding_mode": "up",
            "rounding_increment": "100",
            "return_url": "/v2-clean/fleet/sales",
        },
        follow_redirects=False,
    )
    assert rule.status_code == 303

    status = authenticated_client.post(
        "/v2-clean/fleet/sales/bulk",
        data={
            "vehicle_ids": [str(vehicle.id)],
            "action": "status",
            "bulk_status": "reserved",
            "return_url": "/v2-clean/fleet/sales",
        },
        follow_redirects=False,
    )
    assert status.status_code == 303

    db_session.expire_all()
    profile = db_session.scalar(
        select(VehicleSaleProfile).where(VehicleSaleProfile.vehicle_id == vehicle.id)
    )
    assert profile is not None
    assert profile.status == "reserved"
    assert profile.market_trade_value == Decimal("21000.00")
    assert profile.market_retail_value == Decimal("25200.00")
    assert profile.selling_price == Decimal("22100.00")
    assert profile.market_value_source == "Avaliação interna"
    assert profile.market_valued_on == date.today()
    assert (
        db_session.scalar(select(AuditLog).where(AuditLog.action == "vehicle.sale.bulk_price_rule"))
        is not None
    )

    cost_rule = authenticated_client.post(
        "/v2-clean/fleet/sales/bulk",
        data={
            "vehicle_ids": [str(vehicle.id)],
            "action": "price_rule",
            "price_base": "cost",
            "margin_mode": "value",
            "margin_value": "500",
            "rounding_mode": "none",
            "rounding_increment": "100",
            "return_url": "/v2-clean/fleet/sales",
        },
        follow_redirects=False,
    )
    assert cost_rule.status_code == 303
    db_session.expire_all()
    profile = db_session.scalar(
        select(VehicleSaleProfile).where(VehicleSaleProfile.vehicle_id == vehicle.id)
    )
    assert profile.selling_price == Decimal("19300.00")
    assert profile.price_base == "cost"

    target_cost = authenticated_client.post(
        "/v2-clean/fleet/sales/bulk",
        data={
            "vehicle_ids": [str(vehicle.id)],
            "action": "price_rule",
            "price_target": "trade",
            "price_base": "cost",
            "margin_mode": "value",
            "margin_value": "0",
            "rounding_mode": "none",
            "rounding_increment": "100",
            "return_url": "/v2-clean/fleet/sales",
        },
        follow_redirects=False,
    )
    assert target_cost.status_code == 303
    db_session.expire_all()
    profile = db_session.scalar(
        select(VehicleSaleProfile).where(VehicleSaleProfile.vehicle_id == vehicle.id)
    )
    assert profile.market_trade_value == Decimal("18800.00")


def test_sales_starts_empty_but_filters_return_vehicles(authenticated_client, db_session):
    vehicle = create_sale_vehicle(db_session)

    empty = authenticated_client.get("/v2-clean/fleet/sales")
    assert empty.status_code == 200
    assert "A tabela começa vazia" in empty.text
    assert vehicle.plate not in empty.text

    searched = authenticated_client.get(
        "/v2-clean/fleet/sales", params={"search": "1", "q": vehicle.plate}
    )
    assert searched.status_code == 200
    assert vehicle.plate in searched.text


def test_draft_proposal_can_add_and_remove_vehicles(authenticated_client, db_session):
    first = create_sale_vehicle(db_session)
    second = Vehicle(
        plate="34-CD-56",
        vin="VF3SECONDVEHICLE1",
        rentway_unit_nr="UNIT456",
        brand="Citroen",
        model="C4",
        lifecycle_status="active",
        operational_status="free",
        active=True,
    )
    db_session.add(second)
    db_session.flush()
    db_session.add(
        VehicleExternalSnapshot(
            vehicle_id=second.id,
            source_system="rentway",
            data_json={"groupid": "C2", "km": "12345"},
        )
    )
    db_session.commit()

    created = authenticated_client.post(
        "/v2-clean/fleet/sales/bulk",
        data={"vehicle_ids": [str(first.id)], "action": "proposal"},
        follow_redirects=False,
    )
    assert created.status_code == 303
    proposal = db_session.scalar(select(VehicleSaleProposal))
    assert proposal is not None

    added = authenticated_client.post(
        f"/v2-clean/fleet/sales/proposals/{proposal.id}/vehicles",
        data={"vehicle_ids": [str(second.id)]},
        follow_redirects=False,
    )
    assert added.status_code == 303
    db_session.expire_all()
    lines = db_session.scalars(
        select(VehicleSaleProposalLine).where(
            VehicleSaleProposalLine.proposal_id == proposal.id
        )
    ).all()
    assert {line.vehicle_id for line in lines} == {first.id, second.id}
    assert next(line for line in lines if line.vehicle_id == second.id).snapshot_json["plate"] == "34-CD-56"

    second_line = next(line for line in lines if line.vehicle_id == second.id)
    removed = authenticated_client.post(
        f"/v2-clean/fleet/sales/proposals/{proposal.id}",
        data={"remove_line_id": str(second_line.id)},
        follow_redirects=False,
    )
    assert removed.status_code == 303
    db_session.expire_all()
    remaining = db_session.scalars(
        select(VehicleSaleProposalLine).where(
            VehicleSaleProposalLine.proposal_id == proposal.id
        )
    ).all()
    assert [line.vehicle_id for line in remaining] == [first.id]


def test_vehicle_financial_audit_exports_missing_fields_and_latest_rentway_cost(
    authenticated_client,
    db_session,
):
    vehicle = create_sale_vehicle(db_session)
    db_session.add(
        VehicleFinancialPlan(
            vehicle_id=vehicle.id,
            finance_entity="Santander",
            contract_number="FIN-123",
            start_date=date(2026, 7, 1),
            installment_with_vat=Decimal("325.03"),
            outstanding_amount=Decimal("10000"),
            active=True,
        )
    )
    db_session.commit()

    response = authenticated_client.get("/v2-clean/fleet/financial-audit.csv")

    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]
    assert "12-AB-34" in response.text
    assert "19200" in response.text
    assert "FIN-123" in response.text
    assert "fim" in response.text
    assert "valor residual" in response.text

    page = authenticated_client.get("/v2-clean/fleet/financial-audit")
    assert page.status_code == 200
    assert "Auditoria dos planos financeiros" in page.text
    assert "12-AB-34" in page.text
    assert "Plano importado sem referência" in page.text
    assert "Sem plano mensal" in page.text
    assert "Valor residual" in page.text


def test_financial_audit_uses_same_current_value_as_vehicle_sales(db_session):
    vehicle = create_sale_vehicle(db_session)
    snapshot = db_session.scalar(
        select(VehicleExternalSnapshot).where(VehicleExternalSnapshot.vehicle_id == vehicle.id)
    )
    plan = db_session.scalar(
        select(VehicleFinancialPlan).where(
            VehicleFinancialPlan.vehicle_id == vehicle.id,
            VehicleFinancialPlan.active.is_(True),
        )
    )

    sale_row = _sale_row(vehicle, snapshot, {}, None, plan)
    audit_row = next(
        row
        for row in _financial_audit_rows(db_session)
        if row["vehicle_id"] == vehicle.id
    )

    assert audit_row["current_value_with_vat"] == sale_row["cost"]


def test_sheet_audit_and_sale_share_value_and_last_amortization_date(db_session):
    vehicle = Vehicle(
        plate="FV-10-AA",
        rentway_unit_nr="FV10",
        brand="Peugeot",
        model="208",
        lifecycle_status="active",
        operational_status="free",
        active=True,
    )
    db_session.add(vehicle)
    db_session.flush()
    snapshot = VehicleExternalSnapshot(
        vehicle_id=vehicle.id,
        source_system="rentway",
        data_json={
            "purchase_date": "2025-01-01",
            "acquisition_value": "20000",
            "value_with_tax": "24600",
        },
    )
    plan = VehicleFinancialPlan(
        vehicle_id=vehicle.id,
        finance_entity="Santander",
        contract_number="FV-PLAN",
        start_date=date(2025, 1, 1),
        outstanding_amount=Decimal("15000"),
        amount_reference_date=date(2026, 7, 31),
        active=True,
    )
    db_session.add_all([snapshot, plan])
    db_session.flush()
    db_session.add_all(
        [
            VehicleFinancialPlanInstallment(
                financial_plan_id=plan.id,
                period_number=17,
                period_end=date(2026, 5, 31),
                amortization_amount=Decimal("256.25"),
                outstanding_amount=Decimal("15256.25"),
            ),
            VehicleFinancialPlanInstallment(
                financial_plan_id=plan.id,
                period_number=18,
                period_end=date(2026, 6, 30),
                amortization_amount=Decimal("256.25"),
                outstanding_amount=Decimal("15000"),
            ),
        ]
    )
    db_session.commit()
    installments = db_session.scalars(
        select(VehicleFinancialPlanInstallment)
        .where(VehicleFinancialPlanInstallment.financial_plan_id == plan.id)
        .order_by(VehicleFinancialPlanInstallment.period_number)
    ).all()

    sale_row = _sale_row(vehicle, snapshot, {}, None, plan, installments)
    audit_row = next(
        row for row in _financial_audit_rows(db_session) if row["vehicle_id"] == vehicle.id
    )
    sheet = base_router.clean_vehicle_display_context(db_session, vehicle)

    assert sale_row["cost"] == audit_row["current_value_with_vat"]
    assert sheet["finance"]["current_cost_with_vat"] == base_router.format_eur(
        sale_row["cost"]
    )
    expected_amortization_date = date.today().replace(day=1)
    assert sale_row["current_value_date"] == expected_amortization_date
    assert audit_row["current_value_date"] == expected_amortization_date.isoformat()
    assert sheet["finance"]["current_value_date"] == expected_amortization_date.strftime(
        "%d/%m/%Y"
    )
    assert sheet["finance"]["debt_reference_date"] == "31/07/2026"


def test_vehicle_sale_images_public_snapshot_and_leads(
    authenticated_client,
    db_session,
    tmp_path,
    monkeypatch,
):
    monkeypatch.setattr(settings, "vehicle_sale_media_root", str(tmp_path))
    base_router.EXTERNAL_PORTAL_RATE_LIMIT.clear()
    vehicle = create_sale_vehicle(db_session)
    authorized_document = Document(
        title="Certificado comercial autorizado",
        document_type="certificate",
        original_name="certificado.pdf",
        file_name="certificado.pdf",
        storage_provider="local",
        storage_path="synthetic/certificado.pdf",
        vehicle_id=vehicle.id,
        archived=False,
    )
    db_session.add(authorized_document)
    other_vehicle = Vehicle(plate="DOC-OTHER", active=True)
    db_session.add(other_vehicle)
    db_session.flush()
    foreign_document = Document(
        title="Documento de outra viatura",
        original_name="foreign.pdf",
        file_name="foreign.pdf",
        storage_provider="local",
        storage_path="synthetic/foreign.pdf",
        vehicle_id=other_vehicle.id,
        archived=False,
    )
    archived_document = Document(
        title="Documento arquivado",
        original_name="archived.pdf",
        file_name="archived.pdf",
        storage_provider="local",
        storage_path="synthetic/archived.pdf",
        vehicle_id=vehicle.id,
        archived=True,
    )
    db_session.add_all([foreign_document, archived_document])
    db_session.commit()

    saved = authenticated_client.post(
        f"/v2-clean/fleet/sales/{vehicle.id}",
        data={
            "status": "for_sale",
            "market_trade_value": "21000",
            "market_retail_value": "25200",
            "selling_price": "24900",
            "market_value_source": "Avaliação interna",
            "market_valued_on": "2026-07-31",
            "price_base": "retail",
            "margin_mode": "value",
            "margin_value": "-300",
            "rounding_mode": "nearest",
            "rounding_increment": "100",
            "sale_notes": "Nota interna confidencial sobre custo e preparação.",
            "public_notes": "Viatura revista, disponível para entrega.",
        },
        follow_redirects=False,
    )
    assert saved.status_code == 303
    detail = authenticated_client.get(f"/v2-clean/fleet/sales/{vehicle.id}")
    assert detail.status_code == 200
    assert "Nota interna confidencial" in detail.text
    assert "Galeria permanente" in detail.text

    png = b"\x89PNG\r\n\x1a\n" + b"carfast-test-image"
    uploaded = authenticated_client.post(
        f"/v2-clean/fleet/sales/{vehicle.id}/images",
        data={"category": "exterior", "caption": "Vista dianteira"},
        files={"image": ("frente.png", png, "image/png")},
        follow_redirects=False,
    )
    assert uploaded.status_code == 303
    db_session.expire_all()
    image = db_session.scalar(select(VehicleImage).where(VehicleImage.vehicle_id == vehicle.id))
    assert image is not None
    assert image.category == "exterior"
    assert (tmp_path / image.storage_path).is_file()

    published = authenticated_client.post(
        f"/v2-clean/fleet/sales/{vehicle.id}/publish",
        data={
            "audience": "retail",
            "expires_on": "2026-12-31",
            "image_ids": [str(image.id)],
            "document_ids": [
                str(authorized_document.id),
                str(foreign_document.id),
                str(archived_document.id),
            ],
        },
        follow_redirects=False,
    )
    assert published.status_code == 303
    db_session.expire_all()
    publication = db_session.scalar(
        select(VehicleSalePublication).where(VehicleSalePublication.vehicle_id == vehicle.id)
    )
    assert publication is not None
    assert publication.snapshot_json["sale"]["price"] == "24900.00"
    assert publication.snapshot_json["documents"] == [
        {
            "id": authorized_document.id,
            "title": "Certificado comercial autorizado",
            "type": "certificate",
            "date": None,
        }
    ]
    assert "Documento de outra viatura" not in str(publication.snapshot_json)
    assert "Documento arquivado" not in str(publication.snapshot_json)
    serialized_snapshot = str(publication.snapshot_json)
    assert "debt" not in serialized_snapshot
    assert "margin" not in serialized_snapshot
    assert "Nota interna confidencial" not in serialized_snapshot
    internal_page = authenticated_client.get(f"/v2-clean/fleet/sales/{vehicle.id}")
    assert internal_page.status_code == 200
    assert publication.token in internal_page.text

    public_page = authenticated_client.get(f"/portal/viaturas/{publication.token}")
    assert public_page.status_code == 200
    assert "Peugeot 3008" in public_page.text
    assert "Viatura revista, disponível para entrega." in public_page.text
    assert "24 900,00 €" in public_page.text
    assert "Custo CarFast" not in public_page.text
    assert "Valor em dívida" not in public_page.text
    assert "Nota interna confidencial" not in public_page.text
    assert "Certificado comercial autorizado" in public_page.text

    publications_page = authenticated_client.get(
        "/v2-clean/fleet/sales/publications"
    )
    assert publications_page.status_code == 200
    assert "Relatórios comerciais publicados" in publications_page.text
    assert vehicle.plate in publications_page.text

    archived_image = authenticated_client.post(
        f"/v2-clean/fleet/sales/{vehicle.id}/images/{image.id}/archive",
        follow_redirects=False,
    )
    assert archived_image.status_code == 303
    public_image = authenticated_client.get(
        f"/portal/viaturas/{publication.token}/imagens/{image.id}"
    )
    assert public_image.status_code == 200
    assert public_image.content == png

    question = authenticated_client.post(
        f"/portal/viaturas/{publication.token}/interesse",
        data={
            "kind": "question",
            "name": "Cliente Externo",
            "email": "cliente@example.com",
            "message": "A viatura tem segunda chave?",
            "consent": "1",
        },
        follow_redirects=False,
    )
    assert question.status_code == 303
    assert "sent=1" in question.headers["location"]

    offer = authenticated_client.post(
        f"/portal/viaturas/{publication.token}/interesse",
        data={
            "kind": "offer",
            "name": "Comerciante",
            "phone": "910000000",
            "buyer_company": "Comércio Auto, Lda.",
            "offer_value": "23500",
            "message": "Proposta válida durante cinco dias.",
            "consent": "1",
        },
        follow_redirects=False,
    )
    assert offer.status_code == 303

    purchase = authenticated_client.post(
        f"/portal/viaturas/{publication.token}/interesse",
        data={
            "kind": "purchase",
            "name": "Comprador Final",
            "email": "comprador@example.com",
            "message": "Pretendo avançar com o pedido de compra.",
            "consent": "1",
        },
        follow_redirects=False,
    )
    assert purchase.status_code == 303

    db_session.expire_all()
    leads = db_session.scalars(
        select(VehicleSaleLead).where(VehicleSaleLead.vehicle_id == vehicle.id)
    ).all()
    assert {lead.kind for lead in leads} == {"question", "offer", "purchase"}
    assert next(lead for lead in leads if lead.kind == "offer").offer_value == Decimal("23500.00")

    opportunities = authenticated_client.get("/v2-clean/fleet/sales/opportunities")
    assert opportunities.status_code == 200
    assert "Cliente Externo" in opportunities.text
    assert "Comércio Auto, Lda." in opportunities.text
    assert "12-AB-34" in opportunities.text

    question_lead = next(lead for lead in leads if lead.kind == "question")
    updated_lead = authenticated_client.post(
        f"/v2-clean/fleet/sales/opportunities/{question_lead.id}",
        data={
            "status": "in_review",
            "page": "2",
            "status_filter": "new",
            "kind_filter": "question",
            "q": "Cliente Externo",
        },
        follow_redirects=False,
    )
    assert updated_lead.status_code == 303
    assert "page=2" in updated_lead.headers["location"]
    assert "status=new" in updated_lead.headers["location"]
    assert "kind=question" in updated_lead.headers["location"]
    assert "q=Cliente+Externo" in updated_lead.headers["location"]
    db_session.expire_all()
    assert db_session.get(VehicleSaleLead, question_lead.id).status == "in_review"

    revoked = authenticated_client.post(
        f"/v2-clean/fleet/sales/{vehicle.id}/publications/{publication.id}/revoke",
        follow_redirects=False,
    )
    assert revoked.status_code == 303
    unavailable = authenticated_client.get(f"/portal/viaturas/{publication.token}")
    assert unavailable.status_code == 410
    assert "Link indisponível" in unavailable.text


def test_vehicle_sale_internal_routes_require_authentication(client):
    response = client.get("/v2-clean/fleet/sales", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"].startswith("/login")


def test_vehicle_sales_financial_data_requires_management_permission(client, db_session):
    create_user(
        db_session,
        name="Consulta Frota",
        email="viewer.sales@carfast.local",
        password="Secret123!",
        role_codes=["viewer"],
        organizational_unit_codes=["carfast"],
    )
    db_session.commit()
    login = client.post(
        "/login",
        data={"email": "viewer.sales@carfast.local", "password": "Secret123!"},
        follow_redirects=False,
    )
    assert login.status_code == 303
    client.post("/change-notice", data={"next_url": "/v2-clean"}, follow_redirects=False)

    response = client.get("/v2-clean/fleet/sales", follow_redirects=False)

    assert response.status_code == 303
    assert response.headers["location"] == "/v2-clean?error=forbidden"
