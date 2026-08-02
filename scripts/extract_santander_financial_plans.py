import json
import re
from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(r"C:\Users\andre\OneDrive\Документы\New project\_v2-report-fix")
SOURCE = ROOT / "tmp" / "plans_bpi_audit" / "Planos de renda 19-05" / "SANTANDER"
OUTPUT = ROOT / "tmp" / "santander_financial_plans.json"
REFERENCE_DATE = date(2026, 7, 31)
VAT = Decimal("1.23")


def money(value):
    if value in (None, ""):
        return None
    if isinstance(value, str):
        value = value.replace(".", "").replace(",", ".")
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def with_vat(value):
    amount = money(value)
    return (amount * VAT).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def canonical_contract(value):
    digits = re.sub(r"\D", "", str(value or ""))
    return digits if len(digits) == 12 and digits.startswith("20") else None


def add_months(value, months):
    month_index = value.month - 1 + months
    year = value.year + month_index // 12
    month = month_index % 12 + 1
    return date(year, month, min(value.day, monthrange(year, month)[1]))


def contract_catalog():
    path = SOURCE / "SANTANDER - Contratos Ativos Carfast.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    catalog = {}
    try:
        for row in workbook["Folha1"].iter_rows(min_row=2, values_only=True):
            contract = canonical_contract(row[2] if len(row) > 2 else None)
            plate = str(row[14] or "").strip().upper() if len(row) > 14 else ""
            if not contract or not plate:
                continue
            catalog[contract] = {
                "plate": plate,
                "start_date": row[6].date() if isinstance(row[6], datetime) else row[6],
                "pvp": money(row[8]),
                "financed_amount": money(row[9]),
                "listed_installment": money(row[15]),
                "term_months": int(row[16] or 0),
                "listed_residual": money(row[17]),
                "end_date": row[18].date() if isinstance(row[18], datetime) else row[18],
                "source": path.name,
            }
    finally:
        workbook.close()
    return catalog


def pdf_contract_sections():
    sections = {}
    sources = {}
    for path in sorted(SOURCE.glob("*.pdf")):
        reader = PdfReader(path)
        for page in reader.pages:
            text = page.extract_text() or ""
            matches = re.findall(r"20\d{2}[.]?\d{6}[.]?\d{2}", text)
            contract = canonical_contract(matches[0]) if matches else None
            if not contract:
                continue
            sections[contract] = sections.get(contract, "") + "\n" + text
            sources[contract] = path.name
    return sections, sources


ROW = re.compile(
    r"(?m)^(\d+)\s+(?:(\d{4}-\d{2}-\d{2})\s+)?"
    r"\d+[,.]\d+%\s+([\d.]+,\d{2})\s+([\d.]+,\d{2})\s+"
    r"([\d.]+,\d{2})\s+([\d.]+,\d{2})\s+([\d.]+,\d{2})"
)
INITIAL = re.compile(r"(?m)^Início\s+0,00\s+0,00\s+0,00\s+([\d.]+,\d{2})")
RESIDUAL = re.compile(
    r"(?m)^VR\s+(?:(\d{4}-\d{2}-\d{2})\s+)?\d+[,.]\d+%\s+"
    r"([\d.]+,\d{2})\s+0,00\s+([\d.]+,\d{2})"
)


catalog = contract_catalog()
sections, pdf_sources = pdf_contract_sections()
records = []
unmatched = []
incomplete = []
for contract, text in sorted(sections.items()):
    meta = catalog.get(contract)
    if not meta:
        unmatched.append({"contract": contract, "source": pdf_sources[contract]})
        continue
    matches = list(ROW.finditer(text))
    if not matches:
        incomplete.append({"contract": contract, "plate": meta["plate"], "reason": "sem mensalidades"})
        continue
    schedule = []
    for match in matches:
        period_number = int(match.group(1))
        explicit_date = date.fromisoformat(match.group(2)) if match.group(2) else None
        due_date = explicit_date or add_months(meta["start_date"], period_number - 1)
        amortization = money(match.group(3))
        interest = money(match.group(4))
        installment_with_vat = money(match.group(5))
        outstanding = money(match.group(6))
        tax = money(match.group(7))
        schedule.append(
            {
                "period_number": period_number,
                "period_start": due_date.replace(day=1).isoformat(),
                "period_end": due_date.isoformat(),
                "amortization_amount": str(amortization),
                "interest_amount": str(interest),
                "installment_amount": str(installment_with_vat),
                "outstanding_amount": str(outstanding),
                "outstanding_amount_with_vat": str(with_vat(outstanding)),
                "tax_amount": str(tax),
                "date_source": "PDF" if explicit_date else "reconstruída do início contratual",
            }
        )
    residual_match = RESIDUAL.search(text)
    residual_with_vat = money(residual_match.group(3)) if residual_match else meta["listed_residual"]
    initial_match = INITIAL.search(text)
    initial_amount = money(initial_match.group(1)) if initial_match else meta["financed_amount"]
    elapsed = [item for item in schedule if date.fromisoformat(item["period_end"]) <= REFERENCE_DATE]
    selected = elapsed[-1] if elapsed else schedule[0]
    records.append(
        {
            "plate": meta["plate"],
            "entity": "Santander",
            "contract": contract,
            "start_date": meta["start_date"].isoformat(),
            "end_date": meta["end_date"].isoformat(),
            "term_months": len(schedule),
            "initial_amount": str(initial_amount),
            "outstanding_amount": selected["outstanding_amount"],
            "outstanding_amount_with_vat": selected["outstanding_amount_with_vat"],
            "reference_date": selected["period_end"],
            "installment_amount": selected["installment_amount"],
            "residual_with_vat": str(residual_with_vat or ""),
            "source": pdf_sources[contract],
            "catalog_source": meta["source"],
            "installments": schedule,
        }
    )

OUTPUT.write_text(
    json.dumps(
        {
            "records": records,
            "unmatched": unmatched,
            "incomplete": incomplete,
            "catalog_without_pdf": sorted(set(catalog) - set(sections)),
        },
        ensure_ascii=False,
        indent=2,
    ),
    encoding="utf-8",
)
print(
    f"{len(records)} planos Santander associados; {sum(len(r['installments']) for r in records)} mensalidades; "
    f"{len(unmatched)} sem matrícula; {len(incomplete)} incompletos; "
    f"{len(set(catalog) - set(sections))} contratos do mapa sem PDF"
)
