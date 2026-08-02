import json
import re
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\andre\OneDrive\Документы\New project\_v2-report-fix")
SOURCE = ROOT / "tmp" / "plans_bpi_audit" / "Planos de renda 19-05"
BPI = SOURCE / "BPI"
OUTPUT = ROOT / "tmp" / "bpi_financial_plans.json"
REFERENCE_DATE = date(2026, 7, 31)
VAT = Decimal("1.23")


def money(value):
    if value in (None, ""):
        return None
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def with_vat(value):
    amount = money(value)
    return (amount * VAT).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def iso_date(value):
    return value.date() if isinstance(value, datetime) else value


def contract_plate_map():
    path = SOURCE / "_GERAL" / "Planos_ativos_planos_em_falta.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True)
    mapping = {}
    try:
        for row in workbook["Planos existentes"].iter_rows(values_only=True):
            if len(row) <= 5 or not row[3] or "BPI" not in str(row[3]).upper() or not row[5]:
                continue
            mapping[str(row[5]).replace(".0", "").strip()] = str(row[4]).strip()
    finally:
        workbook.close()
    return mapping


active_contracts = {
    match.group(1)
    for path in BPI.glob("*.pdf")
    if "Vendido" not in path.name
    if (match := re.search(r"BPI - (\d+)", path.name))
}
mapping = contract_plate_map()
workbook_path = BPI / "Planos_Financeiros_TODOS_BPI.xlsx"
workbook = load_workbook(workbook_path, read_only=True, data_only=True)
records = []
unmatched = []
incomplete = []
try:
    for contract in sorted(active_contracts):
        plate = mapping.get(contract)
        if not plate:
            unmatched.append(contract)
            continue
        sheet = workbook[contract]
        schedule = []
        for row in sheet.iter_rows(min_row=9, values_only=True):
            if not row[1] or not row[2] or money(row[5]) is None:
                continue
            due_date = iso_date(row[2])
            schedule.append(
                {
                    "period_number": len(schedule) + 1,
                    "period_label": str(row[1]),
                    "period_start": due_date.replace(day=1).isoformat(),
                    "period_end": due_date.isoformat(),
                    "opening_amount": str(money(row[3])),
                    "installment_amount": str(money(row[4])),
                    "outstanding_amount": str(money(row[5])),
                    "outstanding_amount_with_vat": str(with_vat(row[5])),
                    "interest_amount": str(money(row[6])),
                    "amortization_amount": str(money(row[7])),
                    "type": str(row[8] or ""),
                }
            )
        if not schedule:
            incomplete.append({"contract": contract, "plate": plate})
            continue
        regular = [item for item in schedule if item["period_label"].upper().startswith("R")]
        selected_candidates = [
            item for item in regular if date.fromisoformat(item["period_end"]) <= REFERENCE_DATE
        ]
        selected = selected_candidates[-1] if selected_candidates else regular[0]
        residual = next(
            (with_vat(item["opening_amount"]) for item in schedule if item["period_label"].upper() == "VR"),
            None,
        )
        records.append(
            {
                "plate": plate,
                "entity": "BPI",
                "contract": contract,
                "start_date": regular[0]["period_start"],
                "end_date": schedule[-1]["period_end"],
                "term_months": len(regular),
                "initial_amount": schedule[0]["opening_amount"],
                "outstanding_amount": selected["outstanding_amount"],
                "outstanding_amount_with_vat": selected["outstanding_amount_with_vat"],
                "reference_date": selected["period_end"],
                "installment_amount": selected["installment_amount"],
                "installment_with_vat": str(with_vat(selected["installment_amount"])),
                "residual_with_vat": str(residual or ""),
                "source": workbook_path.name,
                "installments": schedule,
            }
        )
finally:
    workbook.close()

OUTPUT.write_text(
    json.dumps(
        {"records": records, "unmatched": unmatched, "incomplete": incomplete},
        ensure_ascii=False,
        indent=2,
    ),
    encoding="utf-8",
)
print(
    f"{len(records)} planos BPI associados; "
    f"{len(unmatched)} sem matrícula; {len(incomplete)} incompletos"
)
