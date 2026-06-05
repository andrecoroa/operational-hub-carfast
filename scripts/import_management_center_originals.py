from __future__ import annotations

import argparse
import json
import shutil
import sys
from collections import Counter
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.database import SessionLocal
from app.models.admin import User
from app.models.imports import ImportBatch, ImportError, ImportFile, ImportRawRow
from app.models.management_center import (
    ClaimIncident,
    ClaimRefstroLine,
    ClaimRentwayAR,
    ManagementAction,
    ManagementEvidence,
    ManagementHistory,
    ManagementProcess,
    ManagementProcessAssociation,
    ManagementProcessType,
    ManagementRule,
)
from app.services.audit import record_audit
from app.services.management_center import (
    AR_IMPORT_TYPE,
    MANAGEMENT_CENTER_SOURCE_SYSTEM,
    REFSTRO_IMPORT_TYPE,
    associate_to_process,
    ensure_action,
    ensure_management_defaults,
    get_or_create_claim,
    management_storage_root,
    normalize_component,
    normalize_plate,
    refresh_claim_state,
    row_hash,
)

DEFAULT_SOURCE_DIR = Path(r"C:\Users\andre\OneDrive - D'accord Invest - Serviços Partilhados SA\Descargas OneDrive")

REFSTRO_OLD = "documento (1) (3).xlsx"
REFSTRO_RECENT = "CARFAST SINISTROS (8).xlsx"
CRAR = "crar_pervehicle.xlsx"
ACCIDENT_REPORT = "accident_report_2026-06-04_01_07_41.xlsx"

REFSTRO_COMPONENT_COLUMNS = {
    "RC": "RC",
    "DP": "DP",
    "IDS Credor": "IDS Credor",
    "VIDROS": "Vidros",
    "CUSTOS GESTÃO": "Custos de Gestão",
}


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        except Exception:
            return None
    text = str(value).strip()
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_cutoff(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("Use uma data no formato YYYY-MM-DD.") from exc


def parse_decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = str(value).strip().replace("€", "").replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def decimal_or_zero(value: Any) -> Decimal:
    return parse_decimal(value) or Decimal("0.00")


def raw_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def source_paths(source_dir: Path) -> dict[str, Path]:
    return {
        "refstro_old": source_dir / REFSTRO_OLD,
        "refstro_recent": source_dir / REFSTRO_RECENT,
        "crar": source_dir / CRAR,
        "accident_report": source_dir / ACCIDENT_REPORT,
    }


def reset_management_center(db: Session) -> None:
    batch_ids = db.scalars(
        select(ImportBatch.id).where(ImportBatch.import_type.in_((AR_IMPORT_TYPE, REFSTRO_IMPORT_TYPE)))
    ).all()
    if batch_ids:
        db.execute(delete(ImportRawRow).where(ImportRawRow.batch_id.in_(batch_ids)))
        db.execute(delete(ImportError).where(ImportError.batch_id.in_(batch_ids)))
        db.execute(delete(ImportFile).where(ImportFile.batch_id.in_(batch_ids)))
        db.execute(delete(ImportBatch).where(ImportBatch.id.in_(batch_ids)))
    for model in (
        ManagementHistory,
        ManagementEvidence,
        ManagementAction,
        ManagementProcessAssociation,
        ClaimRefstroLine,
        ClaimRentwayAR,
        ClaimIncident,
        ManagementProcess,
        ManagementRule,
        ManagementProcessType,
    ):
        db.execute(delete(model))
    db.flush()


def find_admin_user_id(db: Session) -> int | None:
    user = db.scalar(select(User).where(User.email == "admin@carfast.local"))
    if not user:
        user = db.scalar(select(User).order_by(User.id))
    return user.id if user else None


def create_batch(
    db: Session,
    *,
    import_type: str,
    source_path: Path,
    sheet_name: str,
    headers: list[str],
    total_rows: int,
    user_id: int | None,
) -> ImportBatch:
    batch = ImportBatch(
        source_system=MANAGEMENT_CENTER_SOURCE_SYSTEM,
        import_type=import_type,
        status="running",
        imported_by_id=user_id,
        total_rows=total_rows,
        detail=f"Importação tratada de {source_path.name}.",
    )
    db.add(batch)
    db.flush()

    stored_path = management_storage_root() / f"batch_{batch.id}_{source_path.name}"
    shutil.copyfile(source_path, stored_path)
    db.add(
        ImportFile(
            batch_id=batch.id,
            original_name=source_path.name,
            file_name=stored_path.name,
            storage_path=str(stored_path),
            sheet_name=sheet_name,
            columns_json=headers,
        )
    )
    return batch


def add_raw_row_once(
    db: Session,
    batch: ImportBatch,
    *,
    row_number: int,
    external_reference: str | None,
    raw: dict[str, Any],
) -> None:
    db.add(
        ImportRawRow(
            batch_id=batch.id,
            row_number=row_number,
            external_reference=external_reference,
            raw_json=raw,
            row_hash=row_hash(raw),
        )
    )


def add_import_error(
    db: Session,
    batch: ImportBatch,
    *,
    row_number: int | None,
    entity_type: str,
    message: str,
    raw: dict[str, Any] | None,
) -> None:
    db.add(
        ImportError(
            batch_id=batch.id,
            row_number=row_number,
            entity_type=entity_type,
            error_message=message,
            raw_json=raw,
        )
    )


def existing_ar(db: Session, ar_reference: str | None, plate: str | None) -> ClaimRentwayAR | None:
    if not ar_reference:
        return None
    if plate:
        exact = db.scalar(
            select(ClaimRentwayAR)
            .where(
                ClaimRentwayAR.ar_reference == ar_reference,
                ClaimRentwayAR.plate == plate,
            )
            .order_by(ClaimRentwayAR.id)
        )
        if exact:
            return exact
    return db.scalar(select(ClaimRentwayAR).where(ClaimRentwayAR.ar_reference == ar_reference).order_by(ClaimRentwayAR.id))


def enrich_ar(ar: ClaimRentwayAR, payload: dict[str, Any]) -> bool:
    changed = False
    for field in (
        "status",
        "raw_state",
        "request_date",
        "plate",
        "vehicle_reference",
        "driver_name",
        "customer_name",
        "ra_reference",
        "impro_reference",
        "daaa_reference",
        "insurance_policy",
        "rental_station_out",
        "created_by_rental_station",
    ):
        value = payload.get(field)
        if value and not getattr(ar, field):
            setattr(ar, field, value)
            changed = True
    if payload.get("raw_json"):
        raw = ar.raw_json or {}
        sources = raw.get("_sources", []) if isinstance(raw, dict) else []
        sources.append(
            {
                "source_file": payload.get("source_file"),
                "source_row_number": payload.get("source_row_number"),
                "raw": payload["raw_json"],
            }
        )
        ar.raw_json = {**raw, "_sources": sources}
        changed = True
    return changed


def create_or_update_ar(
    db: Session,
    *,
    process_type,
    payload: dict[str, Any],
    accident_date: date | None,
    batch: ImportBatch,
    user_id: int | None,
    doubtful_reason: str | None = None,
) -> tuple[ClaimIncident, bool, bool]:
    ar = existing_ar(db, payload["ar_reference"], payload["plate"])
    created = False
    updated = False
    if ar:
        updated = enrich_ar(ar, payload)
        existing_association = db.scalar(
            select(ManagementProcessAssociation).where(
                ManagementProcessAssociation.entity_type == "claim_rentway_ar",
                ManagementProcessAssociation.entity_id == ar.id,
                ManagementProcessAssociation.active.is_(True),
            )
        )
        if existing_association:
            claim = db.scalar(select(ClaimIncident).where(ClaimIncident.process_id == existing_association.process_id))
            if not claim:
                claim = get_or_create_claim(
                    db,
                    process_type,
                    plate=payload["plate"],
                    accident_date=accident_date or payload["request_date"],
                    customer_name=payload["customer_name"],
                    driver_name=payload["driver_name"],
                    document_reference=payload["ar_reference"],
                    user_id=user_id,
                )
            if doubtful_reason:
                process = db.get(ManagementProcess, claim.process_id)
                if process:
                    ensure_action(
                        db,
                        process,
                        rule_code="minimum_data_information_request",
                        title="Validar dados AR importados",
                        description=doubtful_reason,
                        mandatory=False,
                    )
            refresh_claim_state(db, claim)
            return claim, created, updated
    else:
        ar = ClaimRentwayAR(**payload)
        db.add(ar)
        db.flush()
        created = True

    claim = get_or_create_claim(
        db,
        process_type,
        plate=payload["plate"],
        accident_date=accident_date or payload["request_date"],
        customer_name=payload["customer_name"],
        driver_name=payload["driver_name"],
        document_reference=payload["ar_reference"],
        user_id=user_id,
    )
    associate_to_process(
        db,
        claim.process_id,
        entity_type="claim_rentway_ar",
        entity_id=ar.id,
        reason=f"Associado por importação tratada AR lote #{batch.id}.",
        user_id=user_id,
    )
    if doubtful_reason:
        process = db.get(ManagementProcess, claim.process_id)
        if process:
            ensure_action(
                db,
                process,
                rule_code="minimum_data_information_request",
                title="Validar dados AR importados",
                description=doubtful_reason,
                mandatory=False,
            )
    refresh_claim_state(db, claim)
    return claim, created, updated


def refstro_components(raw: dict[str, Any]) -> list[tuple[str, Decimal]]:
    components: list[tuple[str, Decimal]] = []
    for source_name, component in REFSTRO_COMPONENT_COLUMNS.items():
        amount = parse_decimal(raw.get(source_name))
        if amount not in (None, Decimal("0.00")):
            components.append((component, amount))
    return components or [("Sem componente", Decimal("0.00"))]


def existing_refstro_line(
    db: Session,
    *,
    refstro_reference: str | None,
    component: str | None,
) -> ClaimRefstroLine | None:
    if not refstro_reference:
        return None
    return db.scalar(
        select(ClaimRefstroLine)
        .where(
            ClaimRefstroLine.refstro_reference == refstro_reference,
            ClaimRefstroLine.component == component,
        )
        .order_by(ClaimRefstroLine.id)
    )


def find_near_ar_claim(
    db: Session,
    *,
    plate: str | None,
    accident_date: date | None,
    max_days: int = 30,
) -> tuple[ClaimIncident, int] | None:
    if not plate or not accident_date:
        return None
    candidates = db.scalars(
        select(ClaimIncident).where(
            ClaimIncident.plate == plate,
            ClaimIncident.accident_date.is_not(None),
        )
    ).all()
    best: tuple[ClaimIncident, int] | None = None
    for candidate in candidates:
        if not candidate.accident_date:
            continue
        has_ar = db.scalar(
            select(ManagementProcessAssociation.id).where(
                ManagementProcessAssociation.process_id == candidate.process_id,
                ManagementProcessAssociation.entity_type == "claim_rentway_ar",
                ManagementProcessAssociation.active.is_(True),
            )
        )
        if not has_ar:
            continue
        delta_days = abs((candidate.accident_date - accident_date).days)
        if delta_days <= max_days and (best is None or delta_days < best[1]):
            best = (candidate, delta_days)
    return best


def create_or_update_refstro_component(
    db: Session,
    *,
    process_type,
    raw: dict[str, Any],
    component: str,
    amount: Decimal,
    source_file: str,
    row_number: int,
    batch: ImportBatch,
    user_id: int | None,
    prefer_recent: bool,
) -> tuple[ClaimIncident, bool, bool]:
    refstro_reference = clean_text(raw.get("REFSTRO"))
    accident_date = parse_date(raw.get("DTASTRO"))
    close_date = parse_date(raw.get("DTAENCERR."))
    plate = normalize_plate(raw.get("MATRICULA"))
    normalized_component = normalize_component(component)
    total_cost = decimal_or_zero(raw.get("CUSTO TOTAL"))
    management_cost = decimal_or_zero(raw.get("CUSTOS GESTÃO"))
    line = existing_refstro_line(
        db,
        refstro_reference=refstro_reference,
        component=normalized_component,
    )
    payload = {
        "refstro_reference": refstro_reference,
        "document_reference": clean_text(raw.get("ADESAO")),
        "plate": plate,
        "accident_date": accident_date,
        "component": normalized_component,
        "status": "Encerrado" if close_date else "Em acompanhamento",
        "close_date": close_date,
        "customer_name": None,
        "driver_name": None,
        "claim_value": amount,
        "cost_value": amount if normalized_component == "Custos de Gestão" else Decimal("0.00"),
        "source_file": source_file,
        "source_row_number": row_number,
        "raw_json": {
            **raw,
            "_total_cost": str(total_cost),
            "_management_cost": str(management_cost),
            "_component_source": component,
            "_prefer_recent": prefer_recent,
        },
    }

    created = False
    updated = False
    if line:
        if prefer_recent:
            for key, value in payload.items():
                setattr(line, key, value)
            updated = True
    else:
        line = ClaimRefstroLine(**payload)
        db.add(line)
        db.flush()
        created = True

    exact_claim = db.scalar(
        select(ClaimIncident).where(
            ClaimIncident.plate == plate,
            ClaimIncident.accident_date == accident_date,
        )
    ) if plate and accident_date else None
    near_match = None if exact_claim else find_near_ar_claim(db, plate=plate, accident_date=accident_date)
    claim = exact_claim or (near_match[0] if near_match else None)
    if not claim:
        claim = get_or_create_claim(
            db,
            process_type,
            plate=plate,
            accident_date=accident_date,
            document_reference=refstro_reference,
            user_id=user_id,
        )
    associate_to_process(
        db,
        claim.process_id,
        entity_type="claim_refstro_line",
        entity_id=line.id,
        reason=f"Associado por importação tratada REFSTRO lote #{batch.id}.",
        user_id=user_id,
    )
    process = db.get(ManagementProcess, claim.process_id)
    if process and near_match:
        ensure_action(
            db,
            process,
            rule_code="minimum_data_information_request",
            title="Validar associação REFSTRO/AR",
            description=(
                f"REFSTRO {refstro_reference or '-'} associada a AR da mesma matrícula "
                f"com diferença de {near_match[1]} dias. Confirmar se é o mesmo sinistro."
            ),
            mandatory=False,
        )
    if process and (not refstro_reference or not plate or not accident_date):
        ensure_action(
            db,
            process,
            rule_code="minimum_data_information_request",
            title="Validar dados REFSTRO importados",
            description="Linha de sinistro importada com referência, matrícula ou data em falta.",
            mandatory=False,
        )
    refresh_claim_state(db, claim)
    return claim, created, updated


def read_refstro_rows(path: Path) -> tuple[str, list[str], list[tuple[int, dict[str, Any]]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in row):
            continue
        raw = {headers[idx] or f"coluna_{idx + 1}": raw_value(value) for idx, value in enumerate(row)}
        rows.append((row_number, raw))
    wb.close()
    return ws.title, headers, rows


def import_refstro_sources(
    db: Session,
    paths: dict[str, Path],
    user_id: int | None,
    *,
    cutoff_date: date,
) -> dict[str, Any]:
    process_type = ensure_management_defaults(db)
    old_sheet, old_headers, old_rows = read_refstro_rows(paths["refstro_old"])
    recent_sheet, recent_headers, recent_rows = read_refstro_rows(paths["refstro_recent"])
    eligible_old = [(row_number, raw) for row_number, raw in old_rows if (parse_date(raw.get("DTASTRO")) or date.min) >= cutoff_date]
    eligible_recent = [
        (row_number, raw)
        for row_number, raw in recent_rows
        if (parse_date(raw.get("DTASTRO")) or date.min) >= cutoff_date
    ]
    recent_refs = {clean_text(raw.get("REFSTRO")) for _, raw in eligible_recent if clean_text(raw.get("REFSTRO"))}
    selected_old = [(row_number, raw) for row_number, raw in old_rows if clean_text(raw.get("REFSTRO")) not in recent_refs]
    selected_old = [(row_number, raw) for row_number, raw in selected_old if (parse_date(raw.get("DTASTRO")) or date.min) >= cutoff_date]
    batches = [
        (paths["refstro_old"], old_sheet, old_headers, selected_old, False),
        (paths["refstro_recent"], recent_sheet, recent_headers, eligible_recent, True),
    ]
    summary = {
        "cutoff_date": cutoff_date.isoformat(),
        "source_rows_total": len(old_rows) + len(recent_rows),
        "source_rows_skipped_before_cutoff_or_without_date": len(old_rows) + len(recent_rows) - len(eligible_old) - len(eligible_recent),
        "source_rows_selected": sum(len(item[3]) for item in batches),
        "source_rows_old_skipped_by_recent_refstro": len(eligible_old) - len(selected_old),
        "created_components": 0,
        "updated_components": 0,
        "errors": 0,
        "touched_process_ids": set(),
        "batches": [],
    }
    for source_path, sheet_name, headers, rows, prefer_recent in batches:
        batch = create_batch(
            db,
            import_type=REFSTRO_IMPORT_TYPE,
            source_path=source_path,
            sheet_name=sheet_name,
            headers=headers,
            total_rows=len(rows),
            user_id=user_id,
        )
        batch_created = 0
        batch_updated = 0
        batch_errors = 0
        for row_number, raw in rows:
            refstro_reference = clean_text(raw.get("REFSTRO"))
            add_raw_row_once(
                db,
                batch,
                row_number=row_number,
                external_reference=refstro_reference,
                raw=raw,
            )
            if not normalize_plate(raw.get("MATRICULA")) and not refstro_reference:
                batch_errors += 1
                add_import_error(
                    db,
                    batch,
                    row_number=row_number,
                    entity_type="claim_refstro_line",
                    message="Linha REFSTRO sem matrícula nem REFSTRO; criada apenas como exceção de importação.",
                    raw=raw,
                )
                continue
            for component, amount in refstro_components(raw):
                claim, created, updated = create_or_update_refstro_component(
                    db,
                    process_type=process_type,
                    raw=raw,
                    component=component,
                    amount=amount,
                    source_file=source_path.name,
                    row_number=row_number,
                    batch=batch,
                    user_id=user_id,
                    prefer_recent=prefer_recent,
                )
                summary["touched_process_ids"].add(claim.process_id)
                batch_created += int(created)
                batch_updated += int(updated)
        batch.status = "completed_with_errors" if batch_errors else "completed"
        batch.created_rows = batch_created
        batch.updated_rows = batch_updated
        batch.error_rows = batch_errors
        batch.finished_at = datetime.now(UTC)
        batch.detail = f"{batch_created} componentes criados; {batch_updated} atualizados; {batch_errors} exceções."
        record_audit(
            db,
            action="management_center.import.refstro_originals.completed",
            entity_type="import_batch",
            entity_id=batch.id,
            detail=batch.detail,
            user_id=user_id,
        )
        summary["created_components"] += batch_created
        summary["updated_components"] += batch_updated
        summary["errors"] += batch_errors
        summary["batches"].append({"id": batch.id, "file": source_path.name, "detail": batch.detail})
    summary["touched_process_ids"] = sorted(summary["touched_process_ids"])
    return summary


def read_accident_report_rows(path: Path) -> tuple[str, list[str], list[tuple[int, dict[str, Any]]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row = 5
    headers = [str(value).strip() if value is not None else "" for value in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(value not in (None, "") for value in row):
            continue
        rows.append((row_number, {headers[idx] or f"coluna_{idx + 1}": raw_value(value) for idx, value in enumerate(row)}))
    wb.close()
    return ws.title, headers, rows


def import_accident_report(db: Session, path: Path, user_id: int | None, *, cutoff_date: date) -> dict[str, Any]:
    process_type = ensure_management_defaults(db)
    sheet_name, headers, rows = read_accident_report_rows(path)
    source_total_rows = len(rows)
    rows = [
        (row_number, raw)
        for row_number, raw in rows
        if (parse_date(raw.get("accidentDate")) or date.min) >= cutoff_date
    ]
    batch = create_batch(
        db,
        import_type=AR_IMPORT_TYPE,
        source_path=path,
        sheet_name=sheet_name,
        headers=headers,
        total_rows=len(rows),
        user_id=user_id,
    )
    seen_ar_refs: set[str] = set()
    status_counter: Counter[str] = Counter()
    created = updated = errors = 0
    touched_process_ids: set[int] = set()
    for row_number, raw in rows:
        ar_reference = clean_text(raw.get("accidentReportID"))
        plate = normalize_plate(raw.get("plateNumber"))
        accident_date = parse_date(raw.get("accidentDate"))
        status = clean_text(raw.get("Status"))
        add_raw_row_once(db, batch, row_number=row_number, external_reference=ar_reference, raw=raw)
        if ar_reference in seen_ar_refs:
            errors += 1
            duplicate_claim = get_or_create_claim(
                db,
                process_type,
                plate=plate,
                accident_date=accident_date,
                document_reference=ar_reference,
                user_id=user_id,
            )
            duplicate_process = db.get(ManagementProcess, duplicate_claim.process_id)
            if duplicate_process:
                ensure_action(
                    db,
                    duplicate_process,
                    rule_code="minimum_data_information_request",
                    title="Validar dados AR importados",
                    description="AR duplicado dentro do export Rentway; confirmar se é duplicado técnico ou ocorrência distinta.",
                    mandatory=False,
                )
            refresh_claim_state(db, duplicate_claim)
            touched_process_ids.add(duplicate_claim.process_id)
            add_import_error(
                db,
                batch,
                row_number=row_number,
                entity_type="claim_rentway_ar",
                message="AR duplicado dentro do export Rentway; associar/validar manualmente se for intencional.",
                raw=raw,
            )
            continue
        if ar_reference:
            seen_ar_refs.add(ar_reference)
        document_type = clean_text(raw.get("documentType"))
        document_no = clean_text(raw.get("documentNo"))
        payload = {
            "ar_reference": ar_reference,
            "status": status,
            "raw_state": clean_text(raw.get("state")),
            "request_date": parse_date(raw.get("requestDate")),
            "plate": plate,
            "vehicle_reference": clean_text(raw.get("unitNo")),
            "driver_name": None,
            "customer_name": None,
            "ra_reference": document_no if document_type == "Rental Agreement" else None,
            "impro_reference": document_no if document_type == "Impro" else None,
            "daaa_reference": clean_text(raw.get("friendlyDeclaration")),
            "insurance_policy": clean_text(raw.get("insurancePolicy")),
            "rental_station_out": clean_text(raw.get("rentalStationOut")),
            "created_by_rental_station": clean_text(raw.get("createdByRentalSation")),
            "source_file": path.name,
            "source_row_number": row_number,
            "raw_json": raw,
        }
        doubtful = None
        if not ar_reference or not plate or not accident_date:
            doubtful = "AR Rentway importado com referência, matrícula ou data de acidente em falta."
        claim, was_created, was_updated = create_or_update_ar(
            db,
            process_type=process_type,
            payload=payload,
            accident_date=accident_date,
            batch=batch,
            user_id=user_id,
            doubtful_reason=doubtful,
        )
        created += int(was_created)
        updated += int(was_updated)
        touched_process_ids.add(claim.process_id)
        status_counter[status or "Sem Status"] += 1
    batch.status = "completed_with_errors" if errors else "completed"
    batch.created_rows = created
    batch.updated_rows = updated
    batch.error_rows = errors
    batch.finished_at = datetime.now(UTC)
    batch.detail = f"{created} ARs criados; {updated} enriquecidos; {errors} exceções."
    record_audit(
        db,
        action="management_center.import.ar_original.completed",
        entity_type="import_batch",
        entity_id=batch.id,
        detail=batch.detail,
        user_id=user_id,
        after_json={"status": status_counter.most_common()},
    )
    return {
        "batch_id": batch.id,
        "cutoff_date": cutoff_date.isoformat(),
        "source_total_rows": source_total_rows,
        "source_rows_skipped_before_cutoff_or_without_date": source_total_rows - len(rows),
        "created": created,
        "updated": updated,
        "errors": errors,
        "statuses": status_counter.most_common(),
        "touched_process_ids": sorted(touched_process_ids),
    }


def read_crar_rows(path: Path) -> tuple[str, list[str], list[tuple[int, dict[str, Any]]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[tuple[int, dict[str, Any]]] = []
    current_vehicle: dict[str, str | None] = {"plate": None, "vehicle": None}
    headers = [
        "accident_report_no",
        "plate",
        "vehicle",
        "document_type",
        "document_no",
        "accident_date",
        "friendly_declaration",
        "driver",
        "customer_code",
        "customer",
    ]
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        text = " ".join(str(value) for value in row if value not in (None, ""))
        if "Veículo" in text or "Veiculo" in text:
            vehicle_text = clean_text(text)
            parts = vehicle_text.split(" - ", 1) if vehicle_text else []
            plate = None
            if len(parts) > 1:
                plate = parts[1].split(":", 1)[0].strip()
            current_vehicle = {"plate": plate, "vehicle": vehicle_text}
        document_type = clean_text(row[3] if len(row) > 3 else None)
        if document_type not in {"RA", "IMPRO"} or not clean_text(row[0] if row else None):
            continue
        raw = {
            "accident_report_no": raw_value(row[0]),
            "plate": current_vehicle.get("plate"),
            "vehicle": current_vehicle.get("vehicle"),
            "document_type": document_type,
            "document_no": raw_value(row[4] if len(row) > 4 else None),
            "accident_date": raw_value(row[5] if len(row) > 5 else None),
            "friendly_declaration": raw_value(row[6] if len(row) > 6 else None),
            "driver": raw_value(row[7] if len(row) > 7 else None),
            "customer_code": raw_value(row[10] if len(row) > 10 else None),
            "customer": raw_value(row[11] if len(row) > 11 else None),
        }
        rows.append((row_number, raw))
    wb.close()
    return ws.title, headers, rows


def import_crar(db: Session, path: Path, user_id: int | None, *, cutoff_date: date) -> dict[str, Any]:
    process_type = ensure_management_defaults(db)
    sheet_name, headers, rows = read_crar_rows(path)
    source_total_rows = len(rows)
    rows = [
        (row_number, raw)
        for row_number, raw in rows
        if (parse_date(raw.get("accident_date")) or date.min) >= cutoff_date
    ]
    batch = create_batch(
        db,
        import_type=AR_IMPORT_TYPE,
        source_path=path,
        sheet_name=sheet_name,
        headers=headers,
        total_rows=len(rows),
        user_id=user_id,
    )
    created = updated = errors = skipped_without_primary_ar = 0
    touched_process_ids: set[int] = set()
    for row_number, raw in rows:
        ar_reference = clean_text(raw.get("accident_report_no"))
        plate = normalize_plate(raw.get("plate"))
        accident_date = parse_date(raw.get("accident_date"))
        add_raw_row_once(db, batch, row_number=row_number, external_reference=ar_reference, raw=raw)
        if not ar_reference and not plate:
            errors += 1
            add_import_error(
                db,
                batch,
                row_number=row_number,
                entity_type="claim_rentway_ar",
                message="Linha histórico AR sem referência nem matrícula.",
                raw=raw,
            )
            continue
        document_type = clean_text(raw.get("document_type"))
        document_no = clean_text(raw.get("document_no"))
        payload = {
            "ar_reference": ar_reference,
            "status": "Histórico crAR",
            "raw_state": None,
            "request_date": accident_date,
            "plate": plate,
            "vehicle_reference": clean_text(raw.get("vehicle")),
            "driver_name": clean_text(raw.get("driver")),
            "customer_name": clean_text(raw.get("customer")),
            "ra_reference": document_no if document_type == "RA" else None,
            "impro_reference": document_no if document_type == "IMPRO" else None,
            "daaa_reference": clean_text(raw.get("friendly_declaration")),
            "insurance_policy": None,
            "rental_station_out": None,
            "created_by_rental_station": None,
            "source_file": path.name,
            "source_row_number": row_number,
            "raw_json": raw,
        }
        if not existing_ar(db, payload["ar_reference"], None):
            skipped_without_primary_ar += 1
            add_import_error(
                db,
                batch,
                row_number=row_number,
                entity_type="claim_rentway_ar",
                message="Histórico crAR sem AR principal no export Rentway; não cria SIN novo.",
                raw=raw,
            )
            continue
        doubtful = None if plate and accident_date else "Histórico AR importado com matrícula ou data em falta."
        claim, was_created, was_updated = create_or_update_ar(
            db,
            process_type=process_type,
            payload=payload,
            accident_date=accident_date,
            batch=batch,
            user_id=user_id,
            doubtful_reason=doubtful,
        )
        created += int(was_created)
        updated += int(was_updated)
        touched_process_ids.add(claim.process_id)
    batch.status = "completed_with_errors" if errors or skipped_without_primary_ar else "completed"
    batch.created_rows = created
    batch.updated_rows = updated
    batch.skipped_rows = skipped_without_primary_ar
    batch.error_rows = errors
    batch.finished_at = datetime.now(UTC)
    batch.detail = (
        f"{created} ARs históricos criados; {updated} ARs enriquecidos; "
        f"{skipped_without_primary_ar} sem AR principal; {errors} exceções."
    )
    record_audit(
        db,
        action="management_center.import.crar_original.completed",
        entity_type="import_batch",
        entity_id=batch.id,
        detail=batch.detail,
        user_id=user_id,
    )
    return {
        "batch_id": batch.id,
        "cutoff_date": cutoff_date.isoformat(),
        "source_total_rows": source_total_rows,
        "source_rows_skipped_before_cutoff_or_without_date": source_total_rows - len(rows),
        "created": created,
        "updated": updated,
        "skipped_without_primary_ar": skipped_without_primary_ar,
        "errors": errors,
        "touched_process_ids": sorted(touched_process_ids),
    }


def final_counts(db: Session) -> dict[str, Any]:
    return {
        "processes": db.scalar(select(func.count()).select_from(ManagementProcess)) or 0,
        "claims": db.scalar(select(func.count()).select_from(ClaimIncident)) or 0,
        "ars": db.scalar(select(func.count()).select_from(ClaimRentwayAR)) or 0,
        "refstro_lines": db.scalar(select(func.count()).select_from(ClaimRefstroLine)) or 0,
        "open_actions": db.scalar(
            select(func.count()).select_from(ManagementAction).where(ManagementAction.status == "open")
        )
        or 0,
        "information_requests": db.scalar(
            select(func.count()).select_from(ManagementProcess).where(ManagementProcess.phase == "information_request")
        )
        or 0,
        "missing_ar": db.scalar(
            select(func.count()).select_from(ManagementProcess).where(ManagementProcess.phase == "ar_missing")
        )
        or 0,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa originais Sinistros/AR para o Centro de Gestão.")
    parser.add_argument("--source-dir", type=Path, default=DEFAULT_SOURCE_DIR)
    parser.add_argument("--from-date", type=parse_cutoff, default=date(2024, 1, 1))
    parser.add_argument("--reset-management-center", action="store_true")
    parser.add_argument("--summary-path", type=Path, default=None)
    args = parser.parse_args()

    paths = source_paths(args.source_dir)
    missing = [str(path) for path in paths.values() if not path.exists()]
    if missing:
        raise SystemExit(f"Ficheiros em falta: {missing}")

    with SessionLocal() as db:
        if args.reset_management_center:
            reset_management_center(db)
        user_id = find_admin_user_id(db)
        summary = {
            "started_at": datetime.now(UTC).isoformat(),
            "source_dir": str(args.source_dir),
            "from_date": args.from_date.isoformat(),
            "policy": "AR-first: o processo interno nasce preferencialmente do AR; REFSTRO/companhia é associada depois.",
            "accident_report": import_accident_report(db, paths["accident_report"], user_id, cutoff_date=args.from_date),
            "crar": import_crar(db, paths["crar"], user_id, cutoff_date=args.from_date),
            "refstro": import_refstro_sources(db, paths, user_id, cutoff_date=args.from_date),
        }
        db.commit()
        summary["final_counts"] = final_counts(db)
        summary["finished_at"] = datetime.now(UTC).isoformat()

    summary_path = args.summary_path or Path("exports") / f"management_center_originals_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps({"summary_path": str(summary_path), **summary["final_counts"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
