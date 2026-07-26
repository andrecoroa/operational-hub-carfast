from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
import unicodedata

from openpyxl import load_workbook


def normalize_header(name: Any) -> str:
    text = unicodedata.normalize("NFKD", str(name or "").lower())
    return "".join(ch for ch in text if ch.isalnum() and not unicodedata.combining(ch))


def build_column_lookup(headers: list[str]) -> dict[str, int]:
    lookup: dict[str, int] = {}
    for idx, name in enumerate(headers):
        lookup[name] = idx
        lookup[normalize_header(name)] = idx
    return lookup


def row_value(row: tuple[Any, ...], col: dict[str, int], name: str) -> Any:
    idx = col.get(name)
    return row[idx] if idx is not None and idx < len(row) else None


def first_row_value(row: tuple[Any, ...], col: dict[str, int], candidates: list[str]) -> Any:
    for name in candidates:
        value = row_value(row, col, name)
        if value not in (None, ""):
            return value
        value = row_value(row, col, normalize_header(name))
        if value not in (None, ""):
            return value
    return None


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def clean_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def excel_date_to_iso(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
        except (TypeError, ValueError, OverflowError):
            return None
    text = str(value).strip()
    if not text:
        return None
    normalized = text.replace("T", " ").replace("Z", "").strip()
    if normalized.isdigit() and len(normalized) == 8:
        for fmt in ("%Y%m%d", "%d%m%Y"):
            try:
                return datetime.strptime(normalized, fmt).date().isoformat()
            except ValueError:
                pass
    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%d.%m.%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(normalized, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def json_safe_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _non_empty_header_score(row: tuple[Any, ...]) -> int:
    score = 0
    for value in row:
        text = str(value or "").strip()
        if text:
            score += 1
    return score


def _row_is_empty(row: tuple[Any, ...]) -> bool:
    for value in row:
        if str(value or "").strip():
            return False
    return True


def _detect_header_row(ws, search_rows: int = 25) -> tuple[int, list[str]]:
    best_row_idx = 1
    best_headers: list[str] = []
    best_score = -1
    max_row = min(ws.max_row or 1, search_rows)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1):
        score = _non_empty_header_score(row)
        if score > best_score:
            best_score = score
            best_row_idx = row_idx
            best_headers = [str(value).strip() if value is not None else "" for value in row]
    return best_row_idx, best_headers


def iter_xlsx_rows(path: str | Path, preferred_sheet: str | None = None):
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[preferred_sheet] if preferred_sheet and preferred_sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        header_row, headers = _detect_header_row(ws)
        for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if _row_is_empty(row):
                continue
            raw = {
                headers[idx] or f"coluna_{idx + 1}": json_safe_value(value)
                for idx, value in enumerate(row)
                if idx < len(headers)
            }
            yield ws.title, headers, row_number, row, raw
    finally:
        wb.close()
