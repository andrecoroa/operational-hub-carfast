import json
import re
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl


SOURCE = Path(r"C:\Users\andre\OneDrive\Документы\New project\_v2-report-fix\tmp\leaseplan_20260802")
OUTPUT = Path(r"C:\Users\andre\OneDrive\Документы\New project\_v2-report-fix\tmp\leaseplan_updates.json")
REFERENCE_DATE = date(2026, 7, 31)
VAT = Decimal("1.23")
TARGETS = set(
    "AU-85-DN AU-86-DJ AU-87-XZ AV-13-IZ AV-49-OC AV-65-FX AV-71-XD AV-99-UJ "
    "AX-30-EF AX-37-FI AX-46-DJ AZ-17-QH AZ-68-QG BA-33-MD BB-74-FP BC-00-FB "
    "BC-10-EB BC-11-EG BC-12-CX BC-21-BJ BC-27-AC BC-47-BI BC-53-XM BC-58-CU "
    "BC-67-BI BC-68-BG BC-91-EE BC-92-EE".split()
)


def money(value):
    if value in (None, ""):
        return None
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def with_vat(value):
    amount = money(value)
    return (amount * VAT).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) if amount is not None else None


def period_dates(value):
    match = re.search(r"(\d{2})-(\d{2})/(\d{2})/(\d{4})$", str(value or ""))
    if not match:
        return None, None
    year, month = int(match.group(4)), int(match.group(3))
    return date(year, month, int(match.group(1))), date(year, month, int(match.group(2)))


def plate_from_sheet(name):
    text = re.sub(r"[^A-Z0-9]", "", name.upper())
    return f"{text[:2]}-{text[2:4]}-{text[4:6]}" if len(text) >= 6 else ""


records = {}
for source in sorted(SOURCE.glob("*.xlsx")):
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        plate = plate_from_sheet(sheet_name)
        if plate not in TARGETS:
            continue
        sheet = workbook[sheet_name]
        rows = {row_number: [cell for cell in row if cell not in (None, "")] for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1)}
        values = {str(items[0]).strip(): items[1] for items in rows.values() if len(items) >= 2}
        schedule = []
        for row_number in range(44, sheet.max_row + 1):
            items = rows.get(row_number, [])
            if len(items) >= 4:
                start, end = period_dates(items[0])
                if end:
                    schedule.append((start, end, items))
        if not schedule:
            raise RuntimeError(f"Sem plano de amortizacao: {source.name} / {sheet_name}")
        selected = min(schedule, key=lambda item: abs((item[1] - REFERENCE_DATE).days))
        if selected[1] != REFERENCE_DATE:
            raise RuntimeError(f"Sem saldo em {REFERENCE_DATE}: {plate}; mais proximo {selected[1]}")
        record = {
            "plate": plate,
            "entity": "LEASEPLAN",
            "contract": f"55743 / {plate}",
            "start_date": schedule[0][0].isoformat(),
            "end_date": values["Data Final"].date().isoformat() if isinstance(values["Data Final"], datetime) else str(values["Data Final"]),
            "term_months": len(schedule),
            "initial_amount": str(money(values["Capital Inícial"])),
            "outstanding_amount": str(money(selected[2][3])),
            "outstanding_amount_with_vat": str(with_vat(selected[2][3])),
            "reference_date": REFERENCE_DATE.isoformat(),
            "installment_amount": str(money(values["Renda Mensal"])),
            "installment_with_vat": str(with_vat(values["Renda Mensal"])),
            "residual_with_vat": str(with_vat(values["Valor Residual"])),
            "source": source.name,
            "source_document_date": values["Data do Documento:"].date().isoformat(),
            "vehicle_description": str(values["Veículo"]).strip(),
            "installments": [
                {
                    "period_number": index,
                    "period_start": start.isoformat(),
                    "period_end": end.isoformat(),
                    "amortization_amount": str(money(items[1])),
                    "interest_amount": str(money(items[2])),
                    "installment_amount": str(money(items[1]) + money(items[2])),
                    "outstanding_amount": str(money(items[3])),
                    "outstanding_amount_with_vat": str(with_vat(items[3])),
                }
                for index, (start, end, items) in enumerate(schedule, start=1)
            ],
        }
        previous = records.get(plate)
        if previous:
            comparable_previous = {key: value for key, value in previous.items() if key != "source"}
            comparable_record = {key: value for key, value in record.items() if key != "source"}
            if comparable_previous != comparable_record:
                raise RuntimeError(f"Duplicado divergente para {plate}: {previous} != {record}")
            previous["source"] = f'{previous["source"]}; {record["source"]}'
        else:
            records[plate] = record

missing = sorted(TARGETS - records.keys())
if missing:
    raise RuntimeError(f"Matriculas em falta: {missing}")

OUTPUT.write_text(json.dumps([records[key] for key in sorted(records)], ensure_ascii=False, indent=2), encoding="utf-8")
print(f"{len(records)} registos escritos")
